# Django
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import logout, login
from django.contrib.auth import update_session_auth_hash
from django.utils import timezone
from django.urls import reverse
from django.http import HttpResponse
from datetime import timedelta, datetime, time, date
from collections import Counter
from django.db.models import Count, Q
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import json
import calendar  # Importar para el calendario mensual
import logging
import holidays  # Feriados de Chile
import csv
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import matplotlib
matplotlib.use('Agg')  # Usar backend sin GUI
import matplotlib.pyplot as plt
import os

# Django REST Framework
from rest_framework import viewsets, mixins, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.authentication import SessionAuthentication
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

# APP
from .serializer import (
    UsuarioSerializer, PerfilUsuarioSerializer, RolesSerializer, AreasSerializer, CategoriasAjustesSerializer, CarrerasSerializer,
    EstudiantesSerializer, SolicitudesSerializer, EvidenciasSerializer, AsignaturasSerializer, AsignaturasEnCursoSerializer, 
    AjusteRazonableSerializer, AjusteAsignadoSerializer, EntrevistasSerializer, PublicaSolicitudSerializer
)
from .validators import validar_rut_chileno, validar_contraseña, traducir_feriado_chileno
from .models import(
    Usuario, PerfilUsuario, Roles, Areas, CategoriasAjustes, Carreras, Estudiantes, Solicitudes, Evidencias,
    Asignaturas, AsignaturasEnCurso, Entrevistas, AjusteRazonable, AjusteAsignado, HorarioBloqueado, SEMESTRE_CHOICES
)  

# Permisos personalizados
from .permissions import (
    IsAsesorPedagogico, IsDocente, IsDirectorCarrera, 
    IsCoordinadora, IsAsesorTecnico, IsAdminOrReadOnly
)

# ------------ CONSTANTES ------------
ROL_ASESOR = 'Asesor Pedagógico'
ROL_DIRECTOR = 'Director de Carrera'
ROL_DOCENTE = 'Docente'
ROL_ADMIN = 'Administrador'
ROL_COORDINADORA = 'Encargado de Inclusión'
ROL_COORDINADOR_TECNICO_PEDAGOGICO = 'Coordinador Técnico Pedagógico'


# ------------ FUNCIONES UTILITARIAS ------------

def desactivar_asignaturas_semestre_vencido():
    """
    Desactiva automáticamente las asignaturas cuyo semestre ha terminado.
    
    Semestres:
    - Otoño (otono): Marzo - Julio (termina el 31 de julio)
    - Primavera (primavera): Agosto - Diciembre (termina el 31 de diciembre)
    
    Esta función se ejecuta automáticamente cuando se accede al sistema.
    """
    hoy = timezone.localtime().date()
    mes_actual = hoy.month
    anio_actual = hoy.year
    
    asignaturas_desactivadas = 0
    
    # Si estamos en agosto o después, desactivar Otoño del año actual
    if mes_actual >= 8:
        count = Asignaturas.objects.filter(
            is_active=True,
            semestre='otono',
            anio=anio_actual
        ).update(is_active=False)
        asignaturas_desactivadas += count
    
    # Si estamos en enero o febrero, desactivar Primavera del año anterior
    if mes_actual <= 2:
        count = Asignaturas.objects.filter(
            is_active=True,
            semestre='primavera',
            anio=anio_actual - 1
        ).update(is_active=False)
        asignaturas_desactivadas += count
    
    # Desactivar cualquier asignatura de años anteriores que siga activa
    # (Otoño de años pasados)
    count = Asignaturas.objects.filter(
        is_active=True,
        semestre='otono',
        anio__lt=anio_actual
    ).update(is_active=False)
    asignaturas_desactivadas += count
    
    # (Primavera de años pasados, excepto el año actual si estamos en enero-febrero)
    anio_limite = anio_actual - 1 if mes_actual <= 2 else anio_actual
    count = Asignaturas.objects.filter(
        is_active=True,
        semestre='primavera',
        anio__lt=anio_limite
    ).update(is_active=False)
    asignaturas_desactivadas += count
    
    return asignaturas_desactivadas


# ----------------------------------------------
#           Vistas Públicas del Sistema
# ----------------------------------------------

class PublicSolicitudCreateView(APIView):
    """
    Endpoint público para que el Estudiante
    pueda enviar un formulario de solicitud de ajuste.
    (La lógica de creación ahora está en el Serializer)
    """
    authentication_classes = []  # Deshabilitar autenticación para endpoint público
    permission_classes = [AllowAny]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
            # Usamos el serializer modificado que maneja la creación de la cita
            serializer = PublicaSolicitudSerializer(data=request.data)
            if serializer.is_valid():
                solicitud = serializer.save()

                return Response(
                    {
                        "message": "Solicitud creada con éxito.",
                        "solicitud_id": solicitud.id
                    },
                    status=status.HTTP_201_CREATED
                )
            # Los errores de validación (incluyendo "hora tomada") se devuelven aquí
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def buscar_estudiante_por_rut(request):
    """
    Endpoint público para buscar un estudiante por su RUT.
    Devuelve los datos del estudiante si existe, o un 404 si no.
    Usado en el formulario de solicitud para autocompletar datos.
    """
    import re
    rut = request.query_params.get('rut', '').strip()
    
    if not rut:
        return Response(
            {'error': 'Debe proporcionar un RUT'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Validar RUT
    es_valido, mensaje_error = validar_rut_chileno(rut)
    if not es_valido:
        return Response(
            {'error': mensaje_error},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Normalizar el RUT (remover puntos y espacios, mantener guión)
    rut_normalizado = re.sub(r'[.\s]', '', rut).upper()
    
    # Buscar estudiante por RUT exacto o normalizado
    estudiante = Estudiantes.objects.filter(
        Q(rut=rut) | Q(rut=rut_normalizado)
    ).select_related('carreras').first()
    
    if not estudiante:
        return Response(
            {'error': 'RUT no encontrado en el sistema'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Obtener asignaturas activas del estudiante
    asignaturas_activas = AsignaturasEnCurso.objects.filter(
        estudiantes=estudiante,
        estado=True
    ).select_related('asignaturas', 'asignaturas__docente__usuario').values(
        'asignaturas__id',
        'asignaturas__nombre',
        'asignaturas__seccion',
        'asignaturas__docente__usuario__first_name',
        'asignaturas__docente__usuario__last_name'
    )
    
    asignaturas_list = [
        {
            'id': a['asignaturas__id'],
            'nombre': a['asignaturas__nombre'],
            'seccion': a['asignaturas__seccion'],
            'docente': f"{a['asignaturas__docente__usuario__first_name']} {a['asignaturas__docente__usuario__last_name']}" if a['asignaturas__docente__usuario__first_name'] else 'Sin docente'
        }
        for a in asignaturas_activas
    ]
    
    return Response({
        'encontrado': True,
        'estudiante': {
            'id': estudiante.id,
            'nombres': estudiante.nombres,
            'apellidos': estudiante.apellidos,
            'rut': estudiante.rut,
            'email': estudiante.email,
            'numero': estudiante.numero,
            'carrera_id': estudiante.carreras.id,
            'carrera_nombre': estudiante.carreras.nombre,
        },
        'asignaturas': asignaturas_list
    }, status=status.HTTP_200_OK)

    
def vista_formulario_solicitud(request):
    """
    Esta vista pública solo muestra la página HTML
    con el formulario de solicitud (formulario_solicitud.html).
    """

    try:
        carreras = Carreras.objects.all().order_by('nombre')
    except Carreras.DoesNotExist:
        carreras = []

    context = {
        'carreras': carreras,
    }

    return render(request, 'SIAPE/formulario_solicitud.html', context)


@api_view(['GET'])
@permission_classes([AllowAny])
def get_horarios_disponibles(request):
    """
    Endpoint público para obtener los slots de 1 hora disponibles
    para una fecha específica, basado en las citas del Encargado de Inclusión.
    Recibe un parámetro GET: ?date=YYYY-MM-DD
    """
    
    # 1. Obtener la fecha de la consulta con validación de seguridad
    selected_date_str = request.GET.get('date')
    if not selected_date_str:
        return Response({"error": "Debe proporcionar una fecha (date=YYYY-MM-DD)."}, status=status.HTTP_400_BAD_REQUEST)
    
    # Validar longitud para prevenir ataques
    if len(selected_date_str) > 20:
        return Response({"error": "Formato de fecha inválido."}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        # Validar que la fecha no sea demasiado lejana en el futuro (máximo 1 año)
        from datetime import date as date_class
        max_future_date = date_class.today() + timedelta(days=365)
        if selected_date > max_future_date:
            return Response({"error": "No se pueden agendar citas más de un año en el futuro."}, status=status.HTTP_400_BAD_REQUEST)
    except ValueError:
        return Response({"error": "Formato de fecha inválido. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

    # Validación extra: No permitir agendar fines de semana
    if selected_date.weekday() >= 5: # 5 = Sábado, 6 = Domingo
         return Response([], status=status.HTTP_200_OK) # Retorna lista vacía

    # Validación: No permitir agendar en feriados chilenos
    feriados_chile = holidays.Chile(years=selected_date.year)
    if selected_date in feriados_chile:
        return Response([], status=status.HTTP_200_OK)  # Retorna lista vacía si es feriado

    # 2. Definir todos los slots posibles (9:00 a 17:00, ya que 18:00 es el fin)
    possible_hours = range(9, 18)
    all_slots = [f"{hour:02d}:00" for hour in possible_hours]

    try:
        # 3. Encontrar la(s) coordinadora(s)
        coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
        if not coordinadoras.exists():
            return Response({"error": "No hay coordinadoras configuradas."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # 4. Calcular horarios disponibles: un horario está disponible si AL MENOS UNA coordinadora lo tiene libre
        start_of_day = timezone.make_aware(datetime.combine(selected_date, time.min))
        end_of_day = timezone.make_aware(datetime.combine(selected_date, time.max))

        # Para cada slot, verificar si al menos una coordinadora lo tiene disponible
        available_slots = []
        for slot in all_slots:
            # Convertir el slot (ej: "10:00") a datetime
            hora_obj = datetime.strptime(slot, '%H:%M').time()
            slot_datetime = timezone.make_aware(datetime.combine(selected_date, hora_obj))
            
            # Verificar si al menos una coordinadora tiene este horario libre
            # Un horario está disponible si AL MENOS UNA coordinadora NO tiene cita ni horario bloqueado en ese horario
            slot_disponible = False
            for coord in coordinadoras:
                tiene_cita = Entrevistas.objects.filter(
                    coordinadora=coord,
                    fecha_entrevista=slot_datetime
                ).exclude(coordinadora__isnull=True).exists()
                
                # Verificar si el horario está bloqueado para esta coordinadora
                tiene_horario_bloqueado = HorarioBloqueado.objects.filter(
                    coordinadora=coord,
                    fecha_hora=slot_datetime
                ).exists()
                
                if not tiene_cita and not tiene_horario_bloqueado:
                    slot_disponible = True
                    break
            
            if slot_disponible:
                available_slots.append(slot)
        
        return Response(available_slots, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": f"Error interno del servidor: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def get_calendario_disponible(request):
    """
    Endpoint público para obtener la disponibilidad de un mes completo.
    Recibe un parámetro GET: ?month=YYYY-MM
    """
    
    # 1. Obtener el mes y año de la consulta con validación de seguridad
    month_str = request.GET.get('month')
    try:
        # Si no se provee mes, usa el mes actual (en zona horaria de Chile)
        if not month_str:
            target_date = timezone.localtime(timezone.now()).date()
        else:
            # Validar longitud para prevenir ataques
            if len(month_str) > 10:
                return Response({"error": "Formato de mes inválido."}, status=status.HTTP_400_BAD_REQUEST)
            target_date = datetime.strptime(month_str, '%Y-%m').date()
            # Validar que el mes no sea demasiado lejano (máximo 1 año)
            from datetime import date as date_class
            max_future_date = date_class.today() + timedelta(days=365)
            if target_date > max_future_date:
                return Response({"error": "No se pueden consultar meses más de un año en el futuro."}, status=status.HTTP_400_BAD_REQUEST)
    except ValueError:
        return Response({"error": "Formato de mes inválido. Use YYYY-MM."}, status=status.HTTP_400_BAD_REQUEST)

    year = target_date.year
    month = target_date.month

    # 2. Encontrar a las coordinadoras
    coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
    if not coordinadoras.exists():
        return Response({"error": "No hay coordinadoras configuradas."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # 3. Obtener todas las citas ya tomadas en ese mes, agrupadas por coordinadora y día
    # Estructura: {coordinadora_id: {dia_str: set([hora1, hora2, ...])}}
    citas_por_coordinadora_dia = {}
    horarios_bloqueados_por_coordinadora_dia = {}
    
    # Obtener TODAS las citas del mes de TODAS las coordinadoras de una vez
    # Usar rango de fechas en lugar de __year y __month para evitar problemas de zona horaria
    # Calcular el primer y último día del mes en la zona horaria local
    primer_dia_mes = timezone.make_aware(datetime(year, month, 1, 0, 0, 0))
    if month == 12:
        ultimo_dia_mes = timezone.make_aware(datetime(year + 1, 1, 1, 0, 0, 0))
    else:
        ultimo_dia_mes = timezone.make_aware(datetime(year, month + 1, 1, 0, 0, 0))
    
    print(f"[DEBUG] Buscando citas entre {primer_dia_mes} y {ultimo_dia_mes}")
    
    # Filtrar por rango de fechas (esto funciona correctamente con zonas horarias)
    todas_las_citas = Entrevistas.objects.filter(
        coordinadora__in=coordinadoras,
        fecha_entrevista__gte=primer_dia_mes,
        fecha_entrevista__lt=ultimo_dia_mes
    ).exclude(coordinadora__isnull=True).select_related('coordinadora').values_list('coordinadora_id', 'fecha_entrevista')
    
    # Obtener TODOS los horarios bloqueados del mes de TODAS las coordinadoras
    todos_los_horarios_bloqueados = HorarioBloqueado.objects.filter(
        coordinadora__in=coordinadoras,
        fecha_hora__gte=primer_dia_mes,
        fecha_hora__lt=ultimo_dia_mes
    ).select_related('coordinadora').values_list('coordinadora_id', 'fecha_hora')
    
    # Debug: mostrar todas las citas encontradas
    total_citas = todas_las_citas.count()
    print(f"[DEBUG] Total de citas encontradas en {year}-{month:02d}: {total_citas}")
    for coord_id, dt in todas_las_citas:
        dt_local = timezone.localtime(dt)
        print(f"[DEBUG] Cita encontrada: Coord={coord_id}, Fecha UTC={dt}, Fecha Local={dt_local.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.debug(f"Total de citas encontradas en {year}-{month:02d}: {total_citas}")
    
    # Inicializar el diccionario para todas las coordinadoras
    for coord in coordinadoras:
        citas_por_coordinadora_dia[coord.id] = {}
        horarios_bloqueados_por_coordinadora_dia[coord.id] = {}
    
    # Procesar cada cita
    for coord_id, dt in todas_las_citas:
        # Convertir a la zona horaria local de forma consistente
        dt_local = timezone.localtime(dt)
        dia_str = dt_local.strftime('%Y-%m-%d')
        # Normalizar la hora a solo la hora en punto (ej: "10:30" -> "10:00")
        # Esto es necesario porque los slots son de hora en punto
        hora_str = f"{dt_local.hour:02d}:00"
        
        if coord_id not in citas_por_coordinadora_dia:
            citas_por_coordinadora_dia[coord_id] = {}
        if dia_str not in citas_por_coordinadora_dia[coord_id]:
            citas_por_coordinadora_dia[coord_id][dia_str] = set()
        citas_por_coordinadora_dia[coord_id][dia_str].add(hora_str)
        
        print(f"[DEBUG] Cita encontrada: Coordinadora {coord_id}, Día {dia_str}, Hora {hora_str} (original: {dt_local.strftime('%Y-%m-%d %H:%M:%S')})")
        logger.debug(f"Cita encontrada: Coordinadora {coord_id}, Día {dia_str}, Hora {hora_str} (original: {dt_local.strftime('%Y-%m-%d %H:%M:%S')})")
    
    # Procesar cada horario bloqueado
    for coord_id, dt in todos_los_horarios_bloqueados:
        # Convertir a la zona horaria local de forma consistente
        dt_local = timezone.localtime(dt)
        dia_str = dt_local.strftime('%Y-%m-%d')
        # Normalizar la hora a solo la hora en punto
        hora_str = f"{dt_local.hour:02d}:00"
        
        if coord_id not in horarios_bloqueados_por_coordinadora_dia:
            horarios_bloqueados_por_coordinadora_dia[coord_id] = {}
        if dia_str not in horarios_bloqueados_por_coordinadora_dia[coord_id]:
            horarios_bloqueados_por_coordinadora_dia[coord_id][dia_str] = set()
        horarios_bloqueados_por_coordinadora_dia[coord_id][dia_str].add(hora_str)
        
        logger.debug(f"Horario bloqueado encontrado: Encargado de Inclusión {coord_id}, Día {dia_str}, Hora {hora_str}")
    
    # Debug: Log de citas encontradas por coordinadora
    for coord in coordinadoras:
        if coord.id in citas_por_coordinadora_dia:
            logger.debug(f"Encargado de Inclusión {coord.id}: {sum(len(horas) for horas in citas_por_coordinadora_dia[coord.id].values())} horas ocupadas")
            for dia, horas in citas_por_coordinadora_dia[coord.id].items():
                logger.debug(f"  Día {dia}: horas ocupadas {sorted(horas)}")

    # 4. Definir los slots base y preparar la respuesta
    slots_base_por_hora = range(9, 18) # 9:00 a 17:00
    
    # Cargar feriados de Chile para el año actual
    feriados_chile = holidays.Chile(years=year)
    
    # Obtener lista de feriados del mes para enviar al frontend
    feriados_mes = []
    for dia in range(1, calendar.monthrange(year, month)[1] + 1):
        fecha_dia = date(year, month, dia)
        if fecha_dia in feriados_chile:
            nombre_feriado = feriados_chile.get(fecha_dia)
            nombre_espanol = traducir_feriado_chileno(nombre_feriado)
            feriados_mes.append({
                "fecha": fecha_dia.strftime('%Y-%m-%d'),
                "nombre": nombre_espanol
            })
    
    respuesta_api = {
        "fechasConDisponibilidad": [], # Días con al menos 1 hora libre
        "diasCompletos": [],           # Días sin horas libres
        "slotsDetallados": {},         # { "2025-11-13": ["09:00", "14:00"], ... }
        "slotsNoDisponibles": {},      # { "2025-11-13": ["10:00", ...], ... }
        "feriados": feriados_mes       # Lista de feriados del mes con nombre
    }

    # 5. Iterar por cada día del mes
    _, num_dias_mes = calendar.monthrange(year, month)

    now = timezone.localtime(timezone.now())
    hoy_str = now.date().strftime('%Y-%m-%d')

    for dia in range(1, num_dias_mes + 1):
        dia_actual_date = date(year, month, dia)
        dia_actual_str = dia_actual_date.strftime('%Y-%m-%d')

        # Omitir fines de semana, días pasados y feriados (usar fecha actual en zona horaria de Chile)
        hoy_chile = timezone.localtime(timezone.now()).date()
        es_feriado = dia_actual_date in feriados_chile
        if dia_actual_date.weekday() >= 5 or dia_actual_date < hoy_chile or es_feriado:
            continue

        slots_libres = []
        slots_no_disponibles = []
        
        for h in slots_base_por_hora:
            hora_str = f"{h:02d}:00"
            
            # Si es hoy, solo permitir con 2 horas de anticipación
            if dia_actual_str == hoy_str:
                if h <= now.hour + 1:  # Debe ser al menos 2 horas después de la actual
                    slots_no_disponibles.append(hora_str)
                    continue
            
            slot_ocupado = False
            slot_bloqueado = False
            coordinadora_ocupada = None
            if coordinadoras.exists():
                for coord in coordinadoras:
                    citas_coord_dia = citas_por_coordinadora_dia.get(coord.id, {}).get(dia_actual_str, set())
                    horarios_bloqueados_coord_dia = horarios_bloqueados_por_coordinadora_dia.get(coord.id, {}).get(dia_actual_str, set())
                    
                    # Si esta coordinadora tiene una cita en este horario, el slot está ocupado
                    if hora_str in citas_coord_dia:
                        slot_ocupado = True
                        coordinadora_ocupada = coord.id
                        logger.debug(f"✓ Slot {hora_str} del día {dia_actual_str} está ocupado por coordinadora {coord.id} (cita)")
                        break
                    
                    # Si esta coordinadora tiene este horario bloqueado, el slot está bloqueado
                    if hora_str in horarios_bloqueados_coord_dia:
                        slot_bloqueado = True
                        coordinadora_ocupada = coord.id
                        logger.debug(f"✓ Slot {hora_str} del día {dia_actual_str} está bloqueado por coordinadora {coord.id}")
                        break
            
            # Si ninguna coordinadora tiene el horario ocupado ni bloqueado, está disponible
            if not slot_ocupado and not slot_bloqueado and coordinadoras.exists():
                slots_libres.append(hora_str)
                logger.debug(f"  Slot {hora_str} del día {dia_actual_str} está DISPONIBLE")
            else:
                # Al menos una coordinadora tiene este horario ocupado o bloqueado (o no hay coordinadoras)
                slots_no_disponibles.append(hora_str)
                if slot_ocupado:
                    logger.debug(f"  Slot {hora_str} del día {dia_actual_str} agregado a slots_no_disponibles (ocupado por coord {coordinadora_ocupada})")
                elif slot_bloqueado:
                    logger.debug(f"  Slot {hora_str} del día {dia_actual_str} agregado a slots_no_disponibles (bloqueado por coord {coordinadora_ocupada})")
                else:
                    logger.debug(f"  Slot {hora_str} del día {dia_actual_str} agregado a slots_no_disponibles (no hay coordinadoras)")

        # Siempre agregar los slots detallados, incluso si no hay disponibles
        # Esto permite que el frontend muestre correctamente qué horarios están ocupados
        respuesta_api["slotsDetallados"][dia_actual_str] = slots_libres
        respuesta_api["slotsNoDisponibles"][dia_actual_str] = slots_no_disponibles
        
        # Debug: Log de slots para este día
        if slots_no_disponibles:
            print(f"[DEBUG] Día {dia_actual_str}: {len(slots_no_disponibles)} slots no disponibles: {sorted(slots_no_disponibles)}")
            logger.debug(f"Día {dia_actual_str}: {len(slots_no_disponibles)} slots no disponibles: {sorted(slots_no_disponibles)}")
        
        if len(slots_libres) > 0:
            respuesta_api["fechasConDisponibilidad"].append(dia_actual_str)
        else:
            respuesta_api["diasCompletos"].append(dia_actual_str)

    # Debug: Imprimir resumen final
    print(f"[DEBUG] Resumen final - Días con disponibilidad: {len(respuesta_api['fechasConDisponibilidad'])}")
    print(f"[DEBUG] Resumen final - Días completos: {len(respuesta_api['diasCompletos'])}")
    print(f"[DEBUG] Resumen final - Total días procesados: {len(respuesta_api['slotsDetallados'])}")
    for dia, slots in list(respuesta_api['slotsNoDisponibles'].items())[:5]:  # Mostrar primeros 5 días
        if slots:
            print(f"[DEBUG] Día {dia}: {len(slots)} slots no disponibles: {sorted(slots)}")
    
    return Response(respuesta_api, status=status.HTTP_200_OK)


# ----------------------------------------------
#           Vistas Privadas del Sistema
# ----------------------------------------------


# ----------- Vistas para la página ------------
def pagina_index(request):
    """
    Página principal (index) del sistema.
    Muestra descripción del sistema y botones de acceso.
    """
    return render(request, 'SIAPE/index.html')

def seguimiento_caso_estudiante(request):
    """
    Vista pública para que los estudiantes puedan ver el seguimiento de su caso.
    Requiere RUT y número de seguimiento (ID del caso) para autenticación.
    """
    solicitud = None
    error = None
    ajustes_aprobados = []
    entrevistas = []
    
    if request.method == 'POST':
        rut = request.POST.get('rut', '').strip()
        numero_seguimiento = request.POST.get('numero_seguimiento', '').strip()
        
        if rut and numero_seguimiento:
            # Validar RUT
            es_valido, mensaje_error = validar_rut_chileno(rut)
            if not es_valido:
                error = mensaje_error
                context = {
                    'solicitud': None,
                    'ajustes_aprobados': [],
                    'entrevistas': [],
                    'error': error
                }
                return render(request, 'SIAPE/seguimiento_caso.html', context)
            
            # Validar que el número de seguimiento sea un número válido
            try:
                solicitud_id = int(numero_seguimiento)
            except ValueError:
                error = 'El número de seguimiento debe ser un número válido.'
                context = {
                    'solicitud': None,
                    'ajustes_aprobados': [],
                    'entrevistas': [],
                    'error': error
                }
                return render(request, 'SIAPE/seguimiento_caso.html', context)
            
            # Buscar estudiante por RUT
            estudiante = Estudiantes.objects.filter(rut=rut).first()
            if not estudiante:
                error = 'RUT no encontrado en el sistema'
                context = {
                    'solicitud': None,
                    'ajustes_aprobados': [],
                    'entrevistas': [],
                    'error': error
                }
                return render(request, 'SIAPE/seguimiento_caso.html', context)
            
            if estudiante:
                # Buscar la solicitud por ID y verificar que pertenezca al estudiante
                try:
                    solicitud = Solicitudes.objects.filter(
                        id=solicitud_id,
                        estudiantes=estudiante
                    ).select_related(
                        'estudiantes',
                        'estudiantes__carreras',
                        'coordinadora_asignada__usuario',
                        'coordinador_tecnico_pedagogico_asignado__usuario',
                        'asesor_pedagogico_asignado__usuario'
                    ).prefetch_related(
                        'ajusteasignado_set__ajuste_razonable__categorias_ajustes',
                        'ajusteasignado_set__director_aprobador__usuario',
                        'entrevistas_set__coordinadora__usuario'
                    ).first()
                    
                    if not solicitud:
                        error = 'No se encontró un caso con ese número de seguimiento asociado a este RUT. Verifique los datos e intente nuevamente.'
                    else:
                        # Obtener solo los ajustes aprobados para mostrar al estudiante
                        ajustes_aprobados = AjusteAsignado.objects.filter(
                            solicitudes=solicitud,
                            estado_aprobacion='aprobado'
                        ).select_related('ajuste_razonable__categorias_ajustes')
                        
                        # Obtener las entrevistas relacionadas
                        entrevistas = Entrevistas.objects.filter(
                            solicitudes=solicitud
                        ).select_related('coordinadora__usuario').order_by('-fecha_entrevista')
                except Solicitudes.DoesNotExist:
                    error = 'No se encontró un caso con ese número de seguimiento. Verifique los datos e intente nuevamente.'
            else:
                error = 'RUT no encontrado en el sistema. Verifique sus datos e intente nuevamente.'
        else:
            error = 'Por favor, complete todos los campos.'
    
    context = {
        'solicitud': solicitud,
        'ajustes_aprobados': ajustes_aprobados,
        'entrevistas': entrevistas if solicitud else [],
        'error': error
    }
    
    return render(request, 'SIAPE/seguimiento_caso.html', context)

def redireccionamiento_por_rol(request):
    """
    Redirecciona al dashboard correspondiente según el rol del usuario.
    Si el usuario no está autenticado, muestra la página index.
    """
    # Si el usuario no está autenticado, mostrar la página index
    if not request.user.is_authenticated:
        return redirect('index')
    
    # Primero verificar si el usuario tiene un perfil con rol
    if hasattr(request.user, 'perfil') and request.user.perfil and request.user.perfil.rol:
        rol = request.user.perfil.rol.nombre_rol

        if rol == ROL_COORDINADORA:
            return redirect('dashboard_encargado_inclusion')
        elif rol == ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            return redirect('dashboard_coordinador_tecnico_pedagogico')
        elif rol == ROL_ASESOR:
            return redirect('dashboard_asesor')
        elif rol == ROL_DIRECTOR:
            return redirect('dashboard_director')
        elif rol == ROL_ADMIN:
            return redirect('dashboard_admin')
        elif rol == ROL_DOCENTE:
            return redirect('dashboard_docente')

    # Solo si no tiene rol o perfil, verificar si es superuser/staff para enviarlo al admin
    if request.user.is_superuser or request.user.is_staff:
        return redirect('admin:index')

    return redirect('login')
    

def vista_protegida(request):
    """    Redirecciona a login si el usuario no está autenticado """
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'vista_protegida.html')

def logout_view(request):
    logout(request)
    return redirect('index') 

@login_required
def casos_generales(request):
    """
    Vista unificada para buscar y filtrar todos los casos del sistema.
    Filtra casos según el rol del usuario:
    - Encargado de Inclusión: Casos pendientes de entrevista o asignados a él
    - Coordinador Técnico Pedagógico: Casos pendientes de formulación (que debe formular)
    - Asesor Pedagógico: Casos pendientes de preaprobación
    - Director: Casos pendientes de aprobación
    - Admin: Todos los casos
    """

    try:
        perfil = request.user.perfil
        rol_nombre = perfil.rol.nombre_rol
        ROLES_PERMITIDOS = [
            ROL_COORDINADORA,
            ROL_COORDINADOR_TECNICO_PEDAGOGICO,
            ROL_ASESOR,
            ROL_DIRECTOR,
            ROL_ADMIN
        ]
        if rol_nombre not in ROLES_PERMITIDOS:
            messages.error(request, 'No tienes permisos para acceder a esta vista.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        rol_nombre = None
        perfil = None

    # Base query con relaciones optimizadas
    solicitudes_list = Solicitudes.objects.all().select_related(
        'estudiantes', 
        'estudiantes__carreras',
        'coordinadora_asignada',
        'coordinador_tecnico_pedagogico_asignado',
        'asesor_pedagogico_asignado'
    ).prefetch_related(
        'ajusteasignado_set__ajuste_razonable__categorias_ajustes'
    ).distinct().order_by('-created_at')

    # Obtener parámetros de filtro
    q_nombre = request.GET.get('q_nombre', '').strip()
    q_fecha = request.GET.get('q_fecha', '').strip()
    q_ajuste = request.GET.get('q_ajuste', '').strip()
    q_estado = request.GET.get('q_estado', '').strip()  # Filtro por estado
    q_todos = request.GET.get('todos', '').strip()  # Permite ver todos los casos
    
    # Inicializar filtros
    filtros = Q()
    filtros_busqueda = Q()
    
    # 1. Determinar si el usuario quiere ver todos los casos (sin filtro por rol)
    tiene_todos = 'todos' in request.GET
    # Verificar si se seleccionó un estado específico (no vacío)
    tiene_estado_explicito = bool(q_estado and q_estado.strip())
    
    # 2. Aplicar filtro por defecto según el rol
    # Por defecto, SIEMPRE aplicar el filtro por rol (a menos que tenga_todos=True o sea Admin)
    # El filtro por rol se aplica a menos que el usuario explícitamente quiera ver todos los casos
    aplicar_filtro_por_rol = (not tiene_todos and rol_nombre and rol_nombre != ROL_ADMIN and perfil is not None)
    
    if aplicar_filtro_por_rol:
        if rol_nombre == ROL_COORDINADORA:
            # Encargado de Inclusión: Casos pendientes de entrevista o pendientes de formulación del caso
            filtros = Q(estado='pendiente_entrevista') | Q(estado='pendiente_formulacion_caso')
        elif rol_nombre == ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            # Coordinador Técnico Pedagógico: Casos pendientes de formulación de ajustes (que debe formular)
            filtros = Q(estado='pendiente_formulacion_ajustes')
        elif rol_nombre == ROL_ASESOR:
            # Asesor Pedagógico: Casos pendientes de preaprobación
            filtros = Q(estado='pendiente_preaprobacion')
        elif rol_nombre == ROL_DIRECTOR:
            # Director: Casos pendientes de aprobación
            filtros = Q(estado='pendiente_aprobacion')

    # 3. Filtro por estado explícito (si se proporciona)
    # Si tiene_todos=True: El estado seleccionado es el único filtro (reemplaza el filtro por rol)
    # Si tiene_todos=False y es Asesora Pedagógica: Ignorar estado explícito, mantener filtro por rol
    # Si tiene_todos=False y tiene filtro por rol: El estado se combina con el filtro por rol (si existe)
    if tiene_estado_explicito:
        if tiene_todos:
            # Si quiere ver todos los casos, el estado seleccionado es el único filtro
            filtros = Q(estado=q_estado)
        elif rol_nombre == ROL_ASESOR and aplicar_filtro_por_rol:
            # Para Asesora Pedagógica, siempre mantener el filtro por rol (pendiente_preaprobacion)
            # Ignorar el estado seleccionado si no es "Ver Todos"
            pass  # Mantener el filtro por rol aplicado anteriormente
        elif aplicar_filtro_por_rol and filtros:
            # Si tiene filtro por rol activo, combinar con el estado seleccionado
            # Nota: Esto puede resultar en 0 resultados si el estado no coincide con el rol
            filtros &= Q(estado=q_estado)
        else:
            # Si no hay filtro por rol (admin o sin rol), usar solo el estado
            filtros = Q(estado=q_estado)

    # 4. Filtros de búsqueda (se aplican además del filtro por defecto o estado seleccionado)
    if q_nombre:
        filtros_busqueda &= (
            Q(estudiantes__nombres__icontains=q_nombre) | 
            Q(estudiantes__apellidos__icontains=q_nombre) |
            Q(estudiantes__rut__icontains=q_nombre)
        )

    if q_fecha:
        try:
            fecha_obj = datetime.strptime(q_fecha, '%Y-%m-%d').date()
            filtros_busqueda &= Q(created_at__date=fecha_obj)
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")

    if q_ajuste:
        filtros_busqueda &= Q(ajusteasignado__ajuste_razonable__categorias_ajustes__id=q_ajuste)
    
    # 5. Combinar filtros base con filtros de búsqueda
    if filtros_busqueda:
        if filtros:  # Si hay filtro base (por defecto o estado), combinarlo con búsqueda
            filtros &= filtros_busqueda
        else:  # Si no hay filtro base, usar solo los de búsqueda
            filtros = filtros_busqueda
    
    # 6. Aplicar filtros
    # Aplicar filtros si existen
    if filtros:
        solicitudes_list = solicitudes_list.filter(filtros)
    # Si no hay filtros (solo para Admin con tiene_todos=True), mostrar todos los casos sin filtrar

    categorias_ajustes = CategoriasAjustes.objects.all().order_by('nombre_categoria')

    # Obtener opciones de estado para el filtro
    estados_disponibles = Solicitudes.ESTADO_CHOICES

    context = {
        'solicitudes': solicitudes_list,
        'total_casos': solicitudes_list.count(),
        'categorias_ajustes': categorias_ajustes,
        'estados_disponibles': estados_disponibles,
        'filtros_aplicados': {
            'q_nombre': q_nombre,
            'q_fecha': q_fecha,
            'q_ajuste': q_ajuste,
            'q_estado': q_estado,
        },
        'rol_usuario': rol_nombre,
        'mostrando_todos': tiene_todos
    }
    
    return render(request, 'SIAPE/casos_generales.html', context)

# ------------- Páginas del Admin ------------
@login_required
def dashboard_admin(request):
    """
    Dashboard para Administradores del Sistema.
    Muestra estadísticas globales del sistema y métricas de actividad.
    """
    import json
    from collections import Counter
    
    # Verificar permisos
    rol = None
    if hasattr(request.user, 'perfil'):
        if request.user.perfil.rol:
            rol = request.user.perfil.rol.nombre_rol

    if not request.user.is_superuser and rol != ROL_ADMIN:
        return redirect('home')

    # --- Fechas de referencia ---
    hoy = timezone.now()
    # Usar inicio del día para comparaciones más precisas
    inicio_hoy = hoy.replace(hour=0, minute=0, second=0, microsecond=0)
    hace_7_dias = inicio_hoy - timedelta(days=7)
    hace_14_dias = inicio_hoy - timedelta(days=14)
    hace_30_dias = inicio_hoy - timedelta(days=30)
    inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # --- KPIs Principales ---
    kpis = {
        'total_usuarios': Usuario.objects.filter(is_active=True).count(),
        'total_estudiantes': Estudiantes.objects.count(),
        'total_solicitudes': Solicitudes.objects.count(),
        'solicitudes_en_proceso': Solicitudes.objects.exclude(estado__in=['aprobado', 'rechazado']).count(),
        'solicitudes_aprobadas': Solicitudes.objects.filter(estado='aprobado').count(),
        'solicitudes_rechazadas': Solicitudes.objects.filter(estado='rechazado').count(),
    }
    
    # --- Usuarios Activos (últimos 7 días) ---
    usuarios_activos_7d = Usuario.objects.filter(
        last_login__gte=hace_7_dias,
        is_active=True
    ).count()
    
    # --- Estadísticas comparativas (esta semana vs anterior) ---
    solicitudes_esta_semana = Solicitudes.objects.filter(created_at__gte=hace_7_dias).count()
    solicitudes_semana_anterior = Solicitudes.objects.filter(
        created_at__gte=hace_14_dias,
        created_at__lt=hace_7_dias
    ).count()
    
    # Calcular variación porcentual
    if solicitudes_semana_anterior > 0:
        variacion_solicitudes = round(((solicitudes_esta_semana - solicitudes_semana_anterior) / solicitudes_semana_anterior) * 100, 1)
    else:
        variacion_solicitudes = 100 if solicitudes_esta_semana > 0 else 0
    
    # Usuarios nuevos esta semana vs anterior
    usuarios_nuevos_semana = Usuario.objects.filter(date_joined__gte=hace_7_dias).count()
    usuarios_nuevos_anterior = Usuario.objects.filter(
        date_joined__gte=hace_14_dias,
        date_joined__lt=hace_7_dias
    ).count()
    
    # --- Gráfico: Actividad del Sistema ---
    # Primero verificamos si hay solicitudes en los últimos 30 días
    total_solicitudes_30d = Solicitudes.objects.filter(created_at__gte=hace_30_dias).count()
    
    # Si no hay solicitudes en 30 días, buscamos la fecha de la solicitud más antigua
    if total_solicitudes_30d == 0:
        solicitud_mas_antigua = Solicitudes.objects.order_by('created_at').first()
        if solicitud_mas_antigua:
            dias_desde_primera = (hoy.date() - solicitud_mas_antigua.created_at.date()).days
            dias_a_mostrar = min(dias_desde_primera + 1, 90)
            fecha_inicio_grafico = inicio_hoy - timedelta(days=dias_a_mostrar)
        else:
            dias_a_mostrar = 30
            fecha_inicio_grafico = hace_30_dias
    else:
        dias_a_mostrar = 30
        fecha_inicio_grafico = hace_30_dias
    
    # Obtener solicitudes y procesar fechas en Python (evita problemas de TruncDate con MySQL)
    solicitudes_periodo = Solicitudes.objects.filter(
        created_at__gte=fecha_inicio_grafico
    ).values_list('created_at', flat=True)
    
    # Contar solicitudes por día usando Python
    actividad_dict = Counter()
    for fecha_creacion in solicitudes_periodo:
        if fecha_creacion:
            # Convertir a fecha local y luego a string
            fecha_local = timezone.localtime(fecha_creacion) if timezone.is_aware(fecha_creacion) else fecha_creacion
            fecha_key = fecha_local.strftime('%Y-%m-%d')
            actividad_dict[fecha_key] += 1
    
    # Crear lista de todos los días (incluyendo días sin actividad)
    actividad_labels = []
    actividad_data = []
    
    for i in range(dias_a_mostrar, -1, -1):
        dia = (hoy - timedelta(days=i)).date()
        dia_key = dia.strftime('%Y-%m-%d')
        actividad_labels.append(dia.strftime('%d/%m'))
        actividad_data.append(actividad_dict.get(dia_key, 0))
    
    # --- Gráfico: Distribución por Estado ---
    estados_count = {
        'En Proceso': Solicitudes.objects.filter(estado='en_proceso').count(),
        'Pendiente Entrevista': Solicitudes.objects.filter(estado='pendiente_entrevista').count(),
        'Pendiente Formulación': Solicitudes.objects.filter(estado='pendiente_formulacion_ajustes').count(),
        'Pendiente Preaprobación': Solicitudes.objects.filter(estado='pendiente_preaprobacion').count(),
        'Pendiente Aprobación': Solicitudes.objects.filter(estado='pendiente_aprobacion').count(),
        'Aprobado': Solicitudes.objects.filter(estado='aprobado').count(),
        'Rechazado': Solicitudes.objects.filter(estado='rechazado').count(),
    }
    
    estados_labels = list(estados_count.keys())
    estados_data = list(estados_count.values())
    
    # --- Distribución de usuarios por rol ---
    roles_count = PerfilUsuario.objects.values('rol__nombre_rol').annotate(
        total=Count('id')
    ).order_by('-total')
    
    roles_labels = [r['rol__nombre_rol'] or 'Sin Rol' for r in roles_count]
    roles_data = [r['total'] for r in roles_count]
    
    # --- Últimos accesos al sistema ---
    ultimos_accesos = Usuario.objects.filter(
        last_login__isnull=False
    ).select_related('perfil', 'perfil__rol').order_by('-last_login')[:10]
    
    # --- Solicitudes recientes (últimas 5) ---
    solicitudes_recientes = Solicitudes.objects.select_related(
        'estudiantes', 'estudiantes__carreras'
    ).order_by('-created_at')[:5]
    
    # --- Gráfico: Distribución de Solicitudes por Área ---
    distribucion_areas = Areas.objects.annotate(
        total_solicitudes=Count('carreras__estudiantes__solicitudes')
    ).filter(
        total_solicitudes__gt=0
    ).order_by('-total_solicitudes')
    
    total_solicitudes_grafico = sum(area.total_solicitudes for area in distribucion_areas)

    distribucion_con_porcentaje = []
    for area in distribucion_areas:
        porcentaje = (area.total_solicitudes / total_solicitudes_grafico * 100) if total_solicitudes_grafico > 0 else 0
        distribucion_con_porcentaje.append({
            'area': area,
            'total': area.total_solicitudes,
            'porcentaje': round(porcentaje, 1)
        })

    context = {
        'kpis': kpis,
        # Estadísticas de actividad
        'usuarios_activos_7d': usuarios_activos_7d,
        'solicitudes_esta_semana': solicitudes_esta_semana,
        'variacion_solicitudes': variacion_solicitudes,
        'usuarios_nuevos_semana': usuarios_nuevos_semana,
        'usuarios_nuevos_anterior': usuarios_nuevos_anterior,
        # Datos para gráficos (JSON)
        'actividad_labels_json': json.dumps(actividad_labels),
        'actividad_data_json': json.dumps(actividad_data),
        'dias_actividad': dias_a_mostrar,  # Para el título dinámico
        'estados_labels_json': json.dumps(estados_labels),
        'estados_data_json': json.dumps(estados_data),
        'roles_labels_json': json.dumps(roles_labels),
        'roles_data_json': json.dumps(roles_data),
        # Tablas
        'ultimos_accesos': ultimos_accesos,
        'solicitudes_recientes': solicitudes_recientes,
        'distribucion_apoyos': distribucion_con_porcentaje,
    }
    
    return render(request, 'SIAPE/dashboard_admin.html', context)

# --- Vistas de Gestión de Usuarios y Roles ---

def _check_admin_permission(request):
    """Función helper para verificar permisos de admin."""
    rol = None
    if hasattr(request.user, 'perfil') and request.user.perfil.rol:
        rol = request.user.perfil.rol.nombre_rol
    
    if not request.user.is_superuser and rol != ROL_ADMIN:
        return False
    return True

@login_required
def gestion_usuarios_admin(request):
    """
    Página para que el Admin vea y gestione usuarios y roles.
    Incluye búsqueda y filtros.
    """
    if not _check_admin_permission(request):
        return redirect('home')

    # Obtener todos los perfiles base
    perfiles = PerfilUsuario.objects.select_related(
        'usuario', 
        'rol', 
        'area'
    ).all()
    
    # Obtener parámetros de búsqueda y filtros
    q_busqueda = request.GET.get('q', '').strip()
    q_rol = request.GET.get('rol', '')
    q_area = request.GET.get('area', '')
    
    # Construir filtros con Q objects
    filtros = Q()
    
    # Búsqueda general (nombre, email, RUT)
    if q_busqueda:
        # Validar longitud para prevenir ataques
        if len(q_busqueda) > 100:
            messages.error(request, 'El término de búsqueda es demasiado largo.')
        else:
            filtros &= (
                Q(usuario__first_name__icontains=q_busqueda) |
                Q(usuario__last_name__icontains=q_busqueda) |
                Q(usuario__email__icontains=q_busqueda) |
                Q(usuario__rut__icontains=q_busqueda)
            )
    
    # Filtro por rol
    if q_rol:
        try:
            rol_id = int(q_rol)
            if rol_id > 0:
                filtros &= Q(rol_id=rol_id)
        except ValueError:
            pass  # Ignorar valores inválidos
    
    # Filtro por área
    if q_area:
        try:
            area_id = int(q_area)
            if area_id > 0:
                filtros &= Q(area_id=area_id)
        except ValueError:
            pass  # Ignorar valores inválidos
    
    # Aplicar filtros
    if filtros:
        perfiles = perfiles.filter(filtros)
    
    # Ordenar resultados
    perfiles = perfiles.order_by('usuario__first_name', 'usuario__last_name')
    
    # Paginación para perfiles (10 por página)
    page_perfiles = request.GET.get('page_usuarios', 1)
    paginator_perfiles = Paginator(perfiles, 10)
    try:
        perfiles_paginados = paginator_perfiles.page(page_perfiles)
    except PageNotAnInteger:
        perfiles_paginados = paginator_perfiles.page(1)
    except EmptyPage:
        perfiles_paginados = paginator_perfiles.page(paginator_perfiles.num_pages)
    
    # Obtener datos para los selectores
    roles_disponibles = Roles.objects.all().order_by('nombre_rol')
    areas_disponibles = Areas.objects.all().order_by('nombre')
    # Obtener roles con conteo de usuarios
    roles_list = Roles.objects.annotate(
        num_usuarios=Count('perfilusuario')
    ).order_by('nombre_rol')
    
    # Paginación para roles (10 por página)
    page_roles = request.GET.get('page_roles', 1)
    paginator_roles = Paginator(roles_list, 10)
    try:
        roles_paginados = paginator_roles.page(page_roles)
    except PageNotAnInteger:
        roles_paginados = paginator_roles.page(1)
    except EmptyPage:
        roles_paginados = paginator_roles.page(paginator_roles.num_pages)

    context = {
        'perfiles': perfiles_paginados,
        'roles_disponibles': roles_disponibles,
        'areas_disponibles': areas_disponibles,
        'roles_list': roles_paginados,
        'filtros_aplicados': {
            'q': q_busqueda,
            'rol': q_rol,
            'area': q_area,
        },
    }
    
    return render(request, 'SIAPE/gestion_usuarios_admin.html', context)

@login_required
def agregar_usuario_admin(request):
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos para realizar esta acción.', extra_tags='usuarios')
        return redirect('gestion_usuarios_admin')

    if request.method == 'POST':
        email = request.POST.get('email')
        rut = request.POST.get('rut', '').strip()
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        rol_id = request.POST.get('rol_id')
        area_id = request.POST.get('area_id')

        redirect_url = reverse('gestion_usuarios_admin') + '#seccion-usuarios'

        # Validar RUT
        es_valido, mensaje_error = validar_rut_chileno(rut)
        if not es_valido:
            messages.error(request, mensaje_error, extra_tags='usuarios')
            return redirect(redirect_url)
        
        # Validar contraseña
        es_valida_password, mensaje_error_password = validar_contraseña(password)
        if not es_valida_password:
            messages.error(request, mensaje_error_password, extra_tags='usuarios')
            return redirect(redirect_url)

        try:
            if Usuario.objects.filter(Q(email=email) | Q(rut=rut)).exists():
                messages.error(request, f'Error: Ya existe un usuario con ese Email o RUT.', extra_tags='usuarios')
                return redirect(redirect_url)
            
            rol_obj = get_object_or_404(Roles, id=rol_id)
            area_obj = None
            if area_id:
                area_obj = get_object_or_404(Areas, id=area_id)
                
            nuevo_usuario = Usuario.objects.create_user(
                email=email,
                password=password,
                rut=rut,
                first_name=first_name,
                last_name=last_name
            )
            
            PerfilUsuario.objects.create(
                usuario=nuevo_usuario,
                rol=rol_obj,
                area=area_obj
            )
            messages.success(request, f'Usuario {email} creado y asignado con el rol de {rol_obj.nombre_rol}.', extra_tags='usuarios')
            
        except Exception as e:
            messages.error(request, f'Error al crear el usuario: {str(e)}', extra_tags='usuarios')
    
    return redirect(reverse('gestion_usuarios_admin') + '#seccion-usuarios')


@login_required
def editar_usuario_admin(request, perfil_id):
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos para realizar esta acción.', extra_tags='usuarios')
        return redirect('gestion_usuarios_admin')

    redirect_url = reverse('gestion_usuarios_admin') + '#seccion-usuarios'

    if request.method == 'POST':
        try:
            perfil = get_object_or_404(PerfilUsuario.objects.select_related('usuario'), id=perfil_id)
            usuario = perfil.usuario
            
            rut = request.POST.get('rut', '').strip()
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            password = request.POST.get('password') 
            
            nuevo_rol_id = request.POST.get('rol_id')
            nuevo_area_id = request.POST.get('area_id')

            # Validar RUT
            es_valido, mensaje_error = validar_rut_chileno(rut)
            if not es_valido:
                messages.error(request, mensaje_error, extra_tags='usuarios')
                return redirect(redirect_url)

            if rut != usuario.rut and Usuario.objects.filter(rut=rut).exclude(id=usuario.id).exists():
                messages.error(request, f'Error: El RUT "{rut}" ya está en uso por otro usuario.', extra_tags='usuarios')
                return redirect(redirect_url)
            
            # Validar contraseña si se proporciona
            if password:
                es_valida_password, mensaje_error_password = validar_contraseña(password)
                if not es_valida_password:
                    messages.error(request, mensaje_error_password, extra_tags='usuarios')
                    return redirect(redirect_url)
            
            usuario.first_name = first_name
            usuario.last_name = last_name
            usuario.rut = rut
            
            if password:
                usuario.set_password(password)
            usuario.save()
            
            rol_obj = get_object_or_404(Roles, id=nuevo_rol_id)
            perfil.rol = rol_obj
            
            if nuevo_area_id:
                perfil.area = get_object_or_404(Areas, id=nuevo_area_id)
            else:
                 perfil.area = None
            perfil.save()
            
            messages.success(request, f'Se actualizó correctamente al usuario {usuario.email}.', extra_tags='usuarios')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar el usuario: {str(e)}', extra_tags='usuarios')
            
    return redirect(redirect_url)

@require_POST
@login_required
def activar_desactivar_usuario_admin(request, perfil_id):
    """
    Vista para activar o desactivar un usuario.
    """
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos para realizar esta acción.', extra_tags='usuarios')
        return redirect('gestion_usuarios_admin')
    
    redirect_url = reverse('gestion_usuarios_admin') + '#seccion-usuarios'
    
    try:
        perfil = get_object_or_404(PerfilUsuario.objects.select_related('usuario'), id=perfil_id)
        usuario = perfil.usuario
        
        # No permitir desactivar al mismo usuario que está haciendo la acción
        if usuario.id == request.user.id:
            messages.error(request, 'No puedes desactivar tu propio usuario.', extra_tags='usuarios')
            return redirect(redirect_url)
        
        # No permitir desactivar superusuarios
        if usuario.is_superuser:
            messages.error(request, 'No se puede desactivar un superusuario.', extra_tags='usuarios')
            return redirect(redirect_url)
        
        # Cambiar el estado
        usuario.is_active = not usuario.is_active
        usuario.save()
        
        estado_texto = 'activado' if usuario.is_active else 'desactivado'
        messages.success(request, f'Usuario {usuario.email} ha sido {estado_texto} correctamente.', extra_tags='usuarios')
        
    except Exception as e:
        messages.error(request, f'Error al cambiar el estado del usuario: {str(e)}', extra_tags='usuarios')
    
    return redirect(redirect_url)

@login_required
def gestion_institucional_admin(request):
    if not _check_admin_permission(request):
        return redirect('home')
    
    # Obtener querysets base
    carreras = Carreras.objects.select_related('director__usuario', 'area').annotate(
        total_estudiantes=Count('estudiantes')
    ).order_by('area__nombre', 'nombre')
    asignaturas = Asignaturas.objects.select_related('carreras', 'docente__usuario').all().order_by('nombre')
    areas = Areas.objects.annotate(
        total_carreras=Count('carreras', distinct=True),
        total_docentes=Count('perfilusuario', distinct=True)
    ).order_by('nombre')
    directores = PerfilUsuario.objects.select_related('usuario').filter(rol__nombre_rol=ROL_DIRECTOR).order_by('usuario__first_name')
    docentes = PerfilUsuario.objects.select_related('usuario').filter(rol__nombre_rol=ROL_DOCENTE).order_by('usuario__first_name')
    
    # Paginación para áreas (10 por página)
    page_areas = request.GET.get('page_areas', 1)
    paginator_areas = Paginator(areas, 10)
    try:
        areas_paginados = paginator_areas.page(page_areas)
    except PageNotAnInteger:
        areas_paginados = paginator_areas.page(1)
    except EmptyPage:
        areas_paginados = paginator_areas.page(paginator_areas.num_pages)
    
    # Paginación para carreras (10 por página)
    page_carreras = request.GET.get('page_carreras', 1)
    paginator_carreras = Paginator(carreras, 10)
    try:
        carreras_paginados = paginator_carreras.page(page_carreras)
    except PageNotAnInteger:
        carreras_paginados = paginator_carreras.page(1)
    except EmptyPage:
        carreras_paginados = paginator_carreras.page(paginator_carreras.num_pages)
    
    # Paginación para asignaturas (10 por página)
    page_asignaturas = request.GET.get('page_asignaturas', 1)
    paginator_asignaturas = Paginator(asignaturas, 10)
    try:
        asignaturas_paginados = paginator_asignaturas.page(page_asignaturas)
    except PageNotAnInteger:
        asignaturas_paginados = paginator_asignaturas.page(1)
    except EmptyPage:
        asignaturas_paginados = paginator_asignaturas.page(paginator_asignaturas.num_pages)
    
    # Para los selectores, necesitamos todas las áreas y carreras (sin paginar)
    areas_todas = Areas.objects.all().order_by('nombre')
    
    context = {
        'carreras_list': carreras_paginados,
        'asignaturas_list': asignaturas_paginados,
        'areas_list': areas_paginados,
        'areas_todas': areas_todas,  # Para los selectores de carreras
        'directores_list': directores,
        'docentes_list': docentes,
    }
    return render(request, 'SIAPE/gestion_institucional_admin.html', context)

# Vista eliminada - La funcionalidad de asignar estudiantes a asignaturas ahora la realiza el Director de Carrera
# @login_required
# def asignar_estudiantes_asignaturas_admin(request):
#     """
#     Vista para que el Administrador asigne estudiantes a asignaturas.
#     """
#     if not _check_admin_permission(request):
#         return redirect('home')
#     
#     # Obtener todos los estudiantes y asignaturas
#     estudiantes = Estudiantes.objects.select_related('carreras').all().order_by('nombres', 'apellidos')
#     asignaturas = Asignaturas.objects.select_related('carreras', 'docente__usuario').all().order_by('nombre', 'seccion')
#     carreras = Carreras.objects.all().order_by('nombre')
#     
#     # Filtrar por carrera si se proporciona (con validación)
#     carrera_id = request.GET.get('carrera', '')
#     if carrera_id:
#         # Validar que carrera_id sea un número entero
#         try:
#             carrera_id_int = int(carrera_id)
#             if carrera_id_int <= 0:
#                 raise ValueError("ID inválido")
#             carrera_seleccionada = Carreras.objects.get(id=carrera_id_int)
#             estudiantes = estudiantes.filter(carreras_id=carrera_id_int)
#             asignaturas = asignaturas.filter(carreras_id=carrera_id_int)
#         except (ValueError, Carreras.DoesNotExist):
#             carrera_seleccionada = None
#             messages.error(request, 'Carrera no válida.')
#     else:
#         carrera_seleccionada = None
#     
#     # Procesar el formulario si es POST
#     if request.method == 'POST':
#         estudiante_id = request.POST.get('estudiante')
#         asignaturas_ids = request.POST.getlist('asignaturas')
#         estado = request.POST.get('estado', 'True') == 'True'
#         
#         if estudiante_id and asignaturas_ids:
#             try:
#                 # Validar que estudiante_id sea un número entero válido
#                 estudiante_id_int = int(estudiante_id)
#                 if estudiante_id_int <= 0:
#                     raise ValueError("ID de estudiante inválido")
#                 
#                 estudiante = Estudiantes.objects.get(id=estudiante_id_int)
#                 
#                 # Validar que todos los IDs de asignaturas sean enteros válidos
#                 asignaturas_ids_int = []
#                 for asign_id in asignaturas_ids:
#                     try:
#                         asign_id_int = int(asign_id)
#                         if asign_id_int <= 0:
#                             continue
#                         asignaturas_ids_int.append(asign_id_int)
#                     except ValueError:
#                         continue
#                 
#                 if not asignaturas_ids_int:
#                     messages.error(request, 'Debe seleccionar al menos una asignatura válida.')
#                     return redirect('asignar_estudiantes_asignaturas_admin')
#                 
#                 # Limitar el número de asignaturas para prevenir abuso
#                 if len(asignaturas_ids_int) > 50:
#                     messages.error(request, 'No se pueden asignar más de 50 asignaturas a la vez.')
#                     return redirect('asignar_estudiantes_asignaturas_admin')
#                 
#                 asignaturas_seleccionadas = Asignaturas.objects.filter(id__in=asignaturas_ids_int)
#                 
#                 asignaciones_creadas = 0
#                 asignaciones_actualizadas = 0
#                 
#                 for asignatura in asignaturas_seleccionadas:
#                     # Verificar si ya existe la asignación
#                     asignacion, created = AsignaturasEnCurso.objects.get_or_create(
#                         estudiantes=estudiante,
#                         asignaturas=asignatura,
#                         defaults={'estado': estado}
#                     )
#                     
#                     if not created:
#                         # Si ya existe, actualizar el estado
#                         asignacion.estado = estado
#                         asignacion.save()
#                         asignaciones_actualizadas += 1
#                     else:
#                         asignaciones_creadas += 1
#                 
#                 if asignaciones_creadas > 0 or asignaciones_actualizadas > 0:
#                     messages.success(
#                         request, 
#                         f'Se asignaron {asignaciones_creadas} asignatura(s) nueva(s) y se actualizaron {asignaciones_actualizadas} asignación(es) existente(s) para {estudiante.nombres} {estudiante.apellidos}.'
#                     )
#                 else:
#                     messages.info(request, 'No se realizaron cambios.')
#                     
#             except Estudiantes.DoesNotExist:
#                 messages.error(request, 'El estudiante seleccionado no existe.')
#             except Exception as e:
#                 messages.error(request, f'Error al asignar: {str(e)}')
#         else:
#             messages.error(request, 'Debe seleccionar un estudiante y al menos una asignatura.')
#         
#         # Redirigir para evitar reenvío del formulario
#         return redirect('asignar_estudiantes_asignaturas_admin')
#     
#     # Obtener asignaciones existentes para mostrar en la tabla
#     asignaciones = AsignaturasEnCurso.objects.select_related(
#         'estudiantes', 
#         'estudiantes__carreras',
#         'asignaturas',
#         'asignaturas__carreras',
#         'asignaturas__docente__usuario'
#     ).all().order_by('-created_at')[:50]  # Últimas 50 asignaciones
#     
#     context = {
#         'estudiantes_list': estudiantes,
#         'asignaturas_list': asignaturas,
#         'carreras_list': carreras,
#         'carrera_seleccionada': carrera_seleccionada,
#         'asignaciones_list': asignaciones,
#     }
#     
#     return render(request, 'SIAPE/asignar_estudiantes_asignaturas_admin.html', context)

@login_required
def agregar_rol_admin(request):
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='roles')
        return redirect('gestion_usuarios_admin')
    if request.method == 'POST':
        nombre_rol = request.POST.get('nombre_rol')
        if nombre_rol and not Roles.objects.filter(nombre_rol=nombre_rol).exists():
            Roles.objects.create(nombre_rol=nombre_rol)
            messages.success(request, f'Rol "{nombre_rol}" creado exitosamente.', extra_tags='roles')
        else:
            messages.error(request, 'El nombre del rol no puede estar vacío o ya existe.', extra_tags='roles')
    return redirect(reverse('gestion_usuarios_admin') + '#seccion-roles')
@login_required
def editar_rol_admin(request, rol_id):
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='roles')
        return redirect('gestion_usuarios_admin')
    rol = get_object_or_404(Roles, id=rol_id)
    if request.method == 'POST':
        nombre_rol = request.POST.get('nombre_rol')
        if nombre_rol and not Roles.objects.filter(nombre_rol=nombre_rol).exclude(id=rol_id).exists():
            rol.nombre_rol = nombre_rol
            rol.save()
            messages.success(request, f'Rol actualizado a "{nombre_rol}".', extra_tags='roles')
        else:
            messages.error(request, 'El nombre del rol no puede estar vacío o ya existe.', extra_tags='roles')
    return redirect(reverse('gestion_usuarios_admin') + '#seccion-roles')

@require_POST
@login_required
def eliminar_rol_admin(request, rol_id):
    """
    Vista para eliminar un rol.
    Verifica que no haya usuarios con ese rol antes de eliminar.
    """
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='roles')
        return redirect('gestion_usuarios_admin')
    
    try:
        rol_id_int = int(rol_id)
        if rol_id_int <= 0:
            raise ValueError("ID inválido")
        
        rol = get_object_or_404(Roles, id=rol_id_int)
        nombre_rol = rol.nombre_rol
        
        # Verificar si hay usuarios con este rol
        usuarios_con_rol = PerfilUsuario.objects.filter(rol=rol).count()
        
        if usuarios_con_rol > 0:
            messages.error(
                request, 
                f'No se puede eliminar el rol "{nombre_rol}" porque hay {usuarios_con_rol} usuario(s) asignado(s) a este rol. '
                'Primero debe cambiar el rol de estos usuarios.',
                extra_tags='roles'
            )
        else:
            rol.delete()
            messages.success(request, f'Rol "{nombre_rol}" eliminado exitosamente.', extra_tags='roles')
    
    except ValueError:
        messages.error(request, 'ID de rol inválido.', extra_tags='roles')
    except Exception as e:
        messages.error(request, f'Error al eliminar el rol: {str(e)}', extra_tags='roles')
    
    return redirect(reverse('gestion_usuarios_admin') + '#seccion-roles')
@login_required
def agregar_area_admin(request):
    """
    Vista para agregar un nuevo área.
    """
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='areas')
        return redirect('gestion_institucional_admin')
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        
        # Validación de entrada
        if not nombre:
            messages.error(request, 'El nombre del área no puede estar vacío.', extra_tags='areas')
        elif len(nombre) > 191:
            messages.error(request, 'El nombre del área es demasiado largo (máximo 191 caracteres).', extra_tags='areas')
        elif Areas.objects.filter(nombre=nombre).exists():
            messages.error(request, f'El área "{nombre}" ya existe.', extra_tags='areas')
        else:
            try:
                Areas.objects.create(nombre=nombre)
                messages.success(request, f'Área "{nombre}" creada exitosamente.', extra_tags='areas')
            except Exception as e:
                messages.error(request, f'Error al crear el área: {str(e)}', extra_tags='areas')
    
    return redirect(reverse('gestion_institucional_admin') + '#seccion-areas')

@login_required
def editar_area_admin(request, area_id):
    """
    Vista para editar un área existente.
    """
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='areas')
        return redirect('gestion_institucional_admin')
    
    try:
        # Validar que area_id sea un entero válido
        area_id_int = int(area_id)
        if area_id_int <= 0:
            raise ValueError("ID inválido")
        
        area = get_object_or_404(Areas, id=area_id_int)
        
        if request.method == 'POST':
            nombre = request.POST.get('nombre', '').strip()
            
            # Validación de entrada
            if not nombre:
                messages.error(request, 'El nombre del área no puede estar vacío.', extra_tags='areas')
            elif len(nombre) > 191:
                messages.error(request, 'El nombre del área es demasiado largo (máximo 191 caracteres).', extra_tags='areas')
            elif Areas.objects.filter(nombre=nombre).exclude(id=area_id_int).exists():
                messages.error(request, f'El área "{nombre}" ya existe.', extra_tags='areas')
            else:
                try:
                    area.nombre = nombre
                    area.save()
                    messages.success(request, f'Área actualizada a "{nombre}".', extra_tags='areas')
                except Exception as e:
                    messages.error(request, f'Error al actualizar el área: {str(e)}', extra_tags='areas')
    except ValueError:
        messages.error(request, 'ID de área inválido.', extra_tags='areas')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}', extra_tags='areas')
    
    return redirect(reverse('gestion_institucional_admin') + '#seccion-areas')

@require_POST
@login_required
def eliminar_area_admin(request, area_id):
    """
    Vista para eliminar un área.
    Verifica que no haya carreras o usuarios asociados antes de eliminar.
    """
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='areas')
        return redirect('gestion_institucional_admin')
    
    try:
        area_id_int = int(area_id)
        if area_id_int <= 0:
            raise ValueError("ID inválido")
        
        area = get_object_or_404(Areas, id=area_id_int)
        nombre_area = area.nombre
        
        # Verificar si hay carreras asociadas
        carreras_asociadas = Carreras.objects.filter(area=area).count()
        usuarios_asociados = PerfilUsuario.objects.filter(area=area).count()
        
        if carreras_asociadas > 0 or usuarios_asociados > 0:
            mensaje = f'No se puede eliminar el área "{nombre_area}" porque está asociada a: '
            problemas = []
            if carreras_asociadas > 0:
                problemas.append(f'{carreras_asociadas} carrera(s)')
            if usuarios_asociados > 0:
                problemas.append(f'{usuarios_asociados} usuario(s)')
            messages.error(request, mensaje + ', '.join(problemas) + '.', extra_tags='areas')
        else:
            area.delete()
            messages.success(request, f'Área "{nombre_area}" eliminada exitosamente.', extra_tags='areas')
    
    except ValueError:
        messages.error(request, 'ID de área inválido.', extra_tags='areas')
    except Exception as e:
        messages.error(request, f'Error al eliminar el área: {str(e)}', extra_tags='areas')
    
    return redirect(reverse('gestion_institucional_admin') + '#seccion-areas')

@login_required
def agregar_carrera_admin(request):
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='carreras')
        return redirect('gestion_institucional_admin')
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            area_id = request.POST.get('area_id')
            director_id = request.POST.get('director_id')
            area = get_object_or_404(Areas, id=area_id)
            director = None
            if director_id:
                director = get_object_or_404(PerfilUsuario, id=director_id, rol__nombre_rol=ROL_DIRECTOR)
            Carreras.objects.create(nombre=nombre, area=area, director=director)
            messages.success(request, f'Carrera "{nombre}" creada exitosamente.', extra_tags='carreras')
        except Exception as e:
            messages.error(request, f'Error al crear la carrera: {str(e)}', extra_tags='carreras')
    return redirect(reverse('gestion_institucional_admin') + '#seccion-carreras')
@login_required
def editar_carrera_admin(request, carrera_id):
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='carreras')
        return redirect('gestion_institucional_admin')
    carrera = get_object_or_404(Carreras, id=carrera_id)
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            area_id = request.POST.get('area_id')
            director_id = request.POST.get('director_id')
            area = get_object_or_404(Areas, id=area_id)
            director = None
            if director_id:
                director = get_object_or_404(PerfilUsuario, id=director_id, rol__nombre_rol=ROL_DIRECTOR)
            carrera.nombre = nombre
            carrera.area = area
            carrera.director = director
            carrera.save()
            messages.success(request, f'Carrera "{nombre}" actualizada exitosamente.', extra_tags='carreras')
        except Exception as e:
            messages.error(request, f'Error al actualizar la carrera: {str(e)}', extra_tags='carreras')
    return redirect(reverse('gestion_institucional_admin') + '#seccion-carreras')

@require_POST
@login_required
def eliminar_carrera_admin(request, carrera_id):
    """
    Vista para eliminar una carrera.
    Verifica que no haya estudiantes o asignaturas asociadas antes de eliminar.
    """
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='carreras')
        return redirect('gestion_institucional_admin')
    
    try:
        carrera_id_int = int(carrera_id)
        if carrera_id_int <= 0:
            raise ValueError("ID inválido")
        
        carrera = get_object_or_404(Carreras, id=carrera_id_int)
        nombre_carrera = carrera.nombre
        
        # Verificar si hay estudiantes o asignaturas asociadas
        estudiantes_asociados = Estudiantes.objects.filter(carreras=carrera).count()
        asignaturas_asociadas = Asignaturas.objects.filter(carreras=carrera).count()
        
        if estudiantes_asociados > 0 or asignaturas_asociadas > 0:
            mensaje = f'No se puede eliminar la carrera "{nombre_carrera}" porque está asociada a: '
            problemas = []
            if estudiantes_asociados > 0:
                problemas.append(f'{estudiantes_asociados} estudiante(s)')
            if asignaturas_asociadas > 0:
                problemas.append(f'{asignaturas_asociadas} asignatura(s)')
            messages.error(request, mensaje + ', '.join(problemas) + '.', extra_tags='carreras')
        else:
            carrera.delete()
            messages.success(request, f'Carrera "{nombre_carrera}" eliminada exitosamente.', extra_tags='carreras')
    
    except ValueError:
        messages.error(request, 'ID de carrera inválido.', extra_tags='carreras')
    except Exception as e:
        messages.error(request, f'Error al eliminar la carrera: {str(e)}', extra_tags='carreras')
    
    return redirect(reverse('gestion_institucional_admin') + '#seccion-carreras')

@login_required
def agregar_asignatura_admin(request):
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='asignaturas')
        return redirect('gestion_institucional_admin')
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            seccion = request.POST.get('seccion')
            carrera_id = request.POST.get('carrera_id')
            docente_id = request.POST.get('docente_id')
            carrera = get_object_or_404(Carreras, id=carrera_id)
            docente = get_object_or_404(PerfilUsuario, id=docente_id, rol__nombre_rol=ROL_DOCENTE)
            
            # Determinar semestre y año actual
            hoy = timezone.localtime().date()
            anio_actual = hoy.year
            mes_actual = hoy.month
            # Otoño: Marzo-Julio (meses 3-7), Primavera: Agosto-Diciembre (meses 8-12)
            # Enero-Febrero se considera Primavera del año actual para inscripciones
            if mes_actual >= 3 and mes_actual <= 7:
                semestre_actual = 'otono'
            else:
                semestre_actual = 'primavera'
            
            Asignaturas.objects.create(
                nombre=nombre, 
                seccion=seccion, 
                carreras=carrera, 
                docente=docente,
                semestre=semestre_actual,
                anio=anio_actual,
                is_active=True  # Activa por defecto al crearse
            )
            messages.success(request, f'Asignatura "{nombre} - {seccion}" creada para {semestre_actual.capitalize()} {anio_actual}.', extra_tags='asignaturas')
        except Exception as e:
            messages.error(request, f'Error al crear la asignatura: {str(e)}', extra_tags='asignaturas')
    return redirect(reverse('gestion_institucional_admin') + '#seccion-asignaturas')
@login_required
def editar_asignatura_admin(request, asignatura_id):
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='asignaturas')
        return redirect('gestion_institucional_admin')
    asignatura = get_object_or_404(Asignaturas, id=asignatura_id)
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            seccion = request.POST.get('seccion')
            carrera_id = request.POST.get('carrera_id')
            docente_id = request.POST.get('docente_id')
            carrera = get_object_or_404(Carreras, id=carrera_id)
            docente = get_object_or_404(PerfilUsuario, id=docente_id, rol__nombre_rol=ROL_DOCENTE)
            asignatura.nombre = nombre
            asignatura.seccion = seccion
            asignatura.carreras = carrera
            asignatura.docente = docente
            asignatura.save()
            messages.success(request, f'Asignatura "{nombre} - {seccion}" actualizada.', extra_tags='asignaturas')
        except Exception as e:
            messages.error(request, f'Error al actualizar la asignatura: {str(e)}', extra_tags='asignaturas')
    return redirect(reverse('gestion_institucional_admin') + '#seccion-asignaturas')

@require_POST
@login_required
def eliminar_asignatura_admin(request, asignatura_id):
    """
    Vista para eliminar una asignatura.
    Verifica que no haya estudiantes cursando la asignatura o solicitudes asociadas antes de eliminar.
    """
    if not _check_admin_permission(request):
        messages.error(request, 'No tienes permisos.', extra_tags='asignaturas')
        return redirect('gestion_institucional_admin')
    
    try:
        asignatura_id_int = int(asignatura_id)
        if asignatura_id_int <= 0:
            raise ValueError("ID inválido")
        
        asignatura = get_object_or_404(Asignaturas, id=asignatura_id_int)
        nombre_asignatura = f"{asignatura.nombre} - {asignatura.seccion}"
        
        # Verificar si hay estudiantes cursando esta asignatura
        estudiantes_cursando = AsignaturasEnCurso.objects.filter(asignaturas=asignatura).count()
        
        # Verificar si hay solicitudes que referencien esta asignatura
        solicitudes_asociadas = Solicitudes.objects.filter(asignaturas_solicitadas=asignatura).count()
        
        if estudiantes_cursando > 0 or solicitudes_asociadas > 0:
            mensaje = f'No se puede eliminar la asignatura "{nombre_asignatura}" porque está asociada a: '
            problemas = []
            if estudiantes_cursando > 0:
                problemas.append(f'{estudiantes_cursando} estudiante(s) cursándola')
            if solicitudes_asociadas > 0:
                problemas.append(f'{solicitudes_asociadas} solicitud(es)')
            messages.error(request, mensaje + ', '.join(problemas) + '.', extra_tags='asignaturas')
        else:
            asignatura.delete()
            messages.success(request, f'Asignatura "{nombre_asignatura}" eliminada exitosamente.', extra_tags='asignaturas')
    
    except ValueError:
        messages.error(request, 'ID de asignatura inválido.', extra_tags='asignaturas')
    except Exception as e:
        messages.error(request, f'Error al eliminar la asignatura: {str(e)}', extra_tags='asignaturas')
    
    return redirect(reverse('gestion_institucional_admin') + '#seccion-asignaturas')

# --- VISTA COORDINADORA DE INCLUSIÓN ---
logger = logging.getLogger(__name__)

@login_required
def dashboard_encargado_inclusion(request):
    """
    Dashboard principal para el Encargado de Inclusión.
    Muestra KPIs y la lista de citas para el día de hoy.
    """
    
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADORA:
            # Si no es coordinadora, redirige a home
            messages.error(request, 'No tienes permisos para acceder a este panel.')
            return redirect('home')
    except AttributeError:
        # Si el usuario no tiene perfil o rol
        logger.warning(f"Usuario {request.user.email} sin perfil/rol intentó acceder al dashboard de coordinadora.")
        return redirect('home')

    # 2. --- Configuración de Fechas ---
    now = timezone.localtime(timezone.now())
    today = now.date()
    start_of_today = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    end_of_today = timezone.make_aware(datetime.combine(today, datetime.max.time()))
    
    # Cálculo de la semana (Lunes a Domingo)
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_of_week_dt = timezone.make_aware(datetime.combine(start_of_week, datetime.min.time()))
    end_of_week_dt = timezone.make_aware(datetime.combine(end_of_week, datetime.max.time()))

    # 3. --- Obtener Datos para KPIs ---
    
    # Base de entrevistas para todas las coordinadoras (filtramos por rol)
    # Como todas las coordinadoras deben ver todas las entrevistas del rol
    todas_las_coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
    entrevistas_coordinadora = Entrevistas.objects.filter(
        coordinadora__in=todas_las_coordinadoras
    ).exclude(coordinadora__isnull=True)
    
    # KPI 1: Citas del día (Query que usaremos también para la lista)
    citas_hoy_qs = entrevistas_coordinadora.filter(
        fecha_entrevista__range=(start_of_today, end_of_today)
    ).select_related(
        'solicitudes', 
        'solicitudes__estudiantes', 
        'solicitudes__estudiantes__carreras'
    ).order_by('fecha_entrevista')
    
    kpi_citas_hoy = citas_hoy_qs.count()

    # KPI 2: Citas canceladas esta semana
    kpi_citas_canceladas = entrevistas_coordinadora.filter(
        estado='cancelada',
        updated_at__range=(start_of_week_dt, end_of_week_dt)
    ).count()

    # KPI 3: Casos pendientes de Formulación del caso
    # Contamos las solicitudes que están en estado 'pendiente_formulacion_caso'
    kpi_pendientes_formulacion_caso = Solicitudes.objects.filter(
        estado='pendiente_formulacion_caso'
    ).count()
    
    # KPI 4: Casos devueltos desde Coordinador Técnico Pedagógico
    # Casos que están en 'pendiente_formulacion_caso' y que tienen ajustes asignados
    # (lo que indica que fueron formulados por la asesora técnica y luego devueltos)
    kpi_casos_devueltos_coordinador_tecnico_pedagogico = Solicitudes.objects.filter(
        estado='pendiente_formulacion_caso',
        ajusteasignado__isnull=False
    ).distinct().count()

    # 4. --- Preparar Contexto ---
    context = {
        'nombre_usuario': request.user.first_name,
        'kpis': {
            'citas_hoy': kpi_citas_hoy,
            'citas_canceladas': kpi_citas_canceladas,
            'pendientes_formulacion_caso': kpi_pendientes_formulacion_caso,
            'casos_devueltos_coordinador_tecnico_pedagogico': kpi_casos_devueltos_coordinador_tecnico_pedagogico,
        },
        'citas_del_dia_list': citas_hoy_qs, # Esta es la lista para la sección principal
    }

    # 5. --- Renderizar Template ---
    return render(request, 'SIAPE/dashboard_encargado_inclusion.html', context)

@require_POST
@login_required
def cancelar_cita_dashboard(request, entrevista_id):
    """
    Vista para que el Encargado de Inclusión cancele una cita desde el dashboard.
    """
    # 1. Verificar Permiso
    try:
        if request.user.perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('dashboard_encargado_inclusion')
    except AttributeError:
        return redirect('home')

    # 2. Lógica de la Acción
    try:
        # Cualquier coordinadora del rol puede cancelar cualquier entrevista del rol
        todas_las_coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
        entrevista = get_object_or_404(Entrevistas, id=entrevista_id, coordinadora__in=todas_las_coordinadoras)
        
        if entrevista.estado == 'pendiente':
            entrevista.estado = 'cancelada'
            entrevista.save()
            messages.success(request, 'Cita cancelada exitosamente.')
        else:
            messages.warning(request, 'Esta cita no puede ser cancelada porque ya fue realizada o cancelada anteriormente.')
    except Exception as e:
        messages.error(request, f'Error al cancelar la cita: {str(e)}')
        
    # 3. Redirigir siempre al dashboard
    return redirect('dashboard_encargado_inclusion')

@login_required
def detalle_casos_encargado_inclusion(request, solicitud_id):
    """
    Muestra el detalle de un caso específico.
    Accesible por todos los roles de asesoría.
    """
    
    # 1. --- Verificación de Permisos (Ampliado) ---
    try:
        perfil = request.user.perfil
        ROLES_PERMITIDOS = [
            ROL_COORDINADORA,
            ROL_COORDINADOR_TECNICO_PEDAGOGICO,
            ROL_ASESOR,
            ROL_DIRECTOR,
            ROL_ADMIN
        ]
        if perfil.rol.nombre_rol not in ROLES_PERMITIDOS:
            messages.error(request, 'No tienes permisos para ver esta página.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')

    # 2. --- Obtener Datos del Caso y Validar Acceso ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # Validar que el usuario tiene acceso a esta solicitud específica
    tiene_acceso = False
    if request.user.is_superuser or request.user.is_staff:
        tiene_acceso = True
    else:
        try:
            perfil = request.user.perfil
            rol = perfil.rol.nombre_rol if perfil.rol else None
            
            if rol == ROL_COORDINADORA:
                tiene_acceso = solicitud.coordinadora_asignada == perfil
            elif rol == ROL_COORDINADOR_TECNICO_PEDAGOGICO:
                # Puede ver si está asignado O si la solicitud está en estado que requiere su intervención
                tiene_acceso = (
                    solicitud.coordinador_tecnico_pedagogico_asignado == perfil or
                    solicitud.estado == 'pendiente_formulacion_ajustes'
                )
            elif rol == ROL_ASESOR:
                # Puede ver si está asignado O si la solicitud está en estado que requiere su intervención
                tiene_acceso = (
                    solicitud.asesor_pedagogico_asignado == perfil or
                    solicitud.estado == 'pendiente_preaprobacion'
                )
            elif rol == ROL_DIRECTOR:
                # Director puede ver solicitudes de estudiantes de sus carreras
                carreras_dirigidas = Carreras.objects.filter(director=perfil)
                tiene_acceso = solicitud.estudiantes.carreras in carreras_dirigidas
            elif rol == ROL_ADMIN:
                tiene_acceso = True
            elif rol == ROL_DOCENTE:
                # Docente puede ver casos de estudiantes en sus asignaturas
                mis_asignaturas = Asignaturas.objects.filter(docente=perfil)
                tiene_acceso = solicitud.asignaturas_solicitadas.filter(id__in=mis_asignaturas).exists()
        except AttributeError:
            tiene_acceso = False
    
    if not tiene_acceso:
        messages.error(request, 'No tienes permisos para ver esta solicitud.')
        return redirect('home')
    
    estudiante = solicitud.estudiantes
    
    # 3. --- Determinar acciones permitidas según el rol ---
    rol_nombre = perfil.rol.nombre_rol if perfil else None
    
    # Obtenemos los ajustes asignados
    # Si el usuario es docente, solo mostrar ajustes aprobados
    # Para otros roles, mostrar todos los ajustes
    if rol_nombre == ROL_DOCENTE:
        ajustes = AjusteAsignado.objects.filter(
            solicitudes=solicitud,
            estado_aprobacion='aprobado'
        ).select_related(
            'ajuste_razonable', 
            'ajuste_razonable__categorias_ajustes'
        )
    else:
        ajustes = AjusteAsignado.objects.filter(solicitudes=solicitud).select_related(
            'ajuste_razonable', 
            'ajuste_razonable__categorias_ajustes'
        )
    
    # Obtenemos todas las evidencias
    evidencias = Evidencias.objects.filter(solicitudes=solicitud)
    
    # Obtenemos TODAS las entrevistas relacionadas con esta solicitud
    entrevistas = Entrevistas.objects.filter(solicitudes=solicitud).order_by('-fecha_entrevista')
    
    # Obtenemos las categorías para el modal (si queremos añadir ajustes)
    categorias_ajustes = CategoriasAjustes.objects.all().order_by('nombre_categoria')
    # Permisos de edición: Solo Encargado de Inclusión, Asesor Pedagógico y Admin pueden editar la descripción del caso
    # El Coordinador Técnico Pedagógico NO puede editar el caso formulado por el Encargado de Inclusión
    # Estados editables por el Encargado de Inclusión
    ESTADOS_EDITABLES_ENCARGADO = ['pendiente_entrevista', 'pendiente_formulacion_caso']
    caso_editable_encargado = solicitud.estado in ESTADOS_EDITABLES_ENCARGADO
    
    # El Docente solo puede VER, no editar
    # El Encargado de Inclusión solo puede editar si el caso está en sus estados
    es_docente = rol_nombre == ROL_DOCENTE
    if rol_nombre == ROL_COORDINADORA:
        puede_editar_descripcion = caso_editable_encargado
        puede_agendar_cita = caso_editable_encargado
    elif rol_nombre == ROL_ASESOR:
        puede_editar_descripcion = solicitud.estado == 'pendiente_preaprobacion'
        puede_agendar_cita = False
    elif rol_nombre == ROL_ADMIN or request.user.is_superuser:
        puede_editar_descripcion = True
        puede_agendar_cita = True
    else:
        puede_editar_descripcion = False
        puede_agendar_cita = False
    
    # Acciones de Encargado de Inclusión
    puede_formular_caso = rol_nombre == ROL_COORDINADORA and solicitud.estado == 'pendiente_formulacion_caso'
    puede_enviar_coordinador_tecnico_pedagogico = rol_nombre == ROL_COORDINADORA and solicitud.estado == 'pendiente_formulacion_caso'
    
    # Estados editables por el Coordinador Técnico Pedagógico
    ESTADOS_EDITABLES_COORDINADOR_TECNICO = ['pendiente_formulacion_ajustes']
    caso_editable_coordinador_tecnico = solicitud.estado in ESTADOS_EDITABLES_COORDINADOR_TECNICO
    
    # Acciones de Coordinador Técnico Pedagógico
    puede_formular_ajustes = rol_nombre == ROL_COORDINADOR_TECNICO_PEDAGOGICO and caso_editable_coordinador_tecnico
    puede_enviar_asesor_pedagogico = rol_nombre == ROL_COORDINADOR_TECNICO_PEDAGOGICO and caso_editable_coordinador_tecnico
    puede_devolver_a_encargado_inclusion = rol_nombre == ROL_COORDINADOR_TECNICO_PEDAGOGICO and caso_editable_coordinador_tecnico
    puede_editar_ajustes_coordinador = rol_nombre == ROL_COORDINADOR_TECNICO_PEDAGOGICO and caso_editable_coordinador_tecnico
    puede_eliminar_ajustes_coordinador = rol_nombre == ROL_COORDINADOR_TECNICO_PEDAGOGICO and caso_editable_coordinador_tecnico
    
    # Estados editables por el Asesor Pedagógico
    ESTADOS_EDITABLES_ASESOR = ['pendiente_preaprobacion']
    caso_editable_asesor = solicitud.estado in ESTADOS_EDITABLES_ASESOR
    
    # Acciones de Asesor Pedagógico
    puede_enviar_a_director = rol_nombre == ROL_ASESOR and caso_editable_asesor
    puede_devolver_a_coordinador_tecnico_pedagogico = rol_nombre == ROL_ASESOR and caso_editable_asesor
    puede_editar_ajustes_asesor = rol_nombre == ROL_ASESOR and caso_editable_asesor
    puede_eliminar_ajustes_asesor = rol_nombre == ROL_ASESOR and caso_editable_asesor
    
    # Estados editables por el Director
    ESTADOS_EDITABLES_DIRECTOR = ['pendiente_aprobacion']
    caso_editable_director = solicitud.estado in ESTADOS_EDITABLES_DIRECTOR
    
    # Acciones de Director
    puede_aprobar = rol_nombre == ROL_DIRECTOR and caso_editable_director
    puede_rechazar = rol_nombre == ROL_DIRECTOR and caso_editable_director
    # El Director puede desactivar casos aprobados para enviarlos a revisión
    puede_desactivar_caso = rol_nombre == ROL_DIRECTOR and solicitud.estado == 'aprobado'

    context = {
        'solicitud': solicitud,
        'estudiante': estudiante,
        'ajustes_asignados': ajustes,
        'evidencias': evidencias,
        'entrevistas_list': entrevistas,
        'categorias_ajustes': categorias_ajustes, # Para el modal
        'rol_usuario': rol_nombre,
        'puede_editar_descripcion': puede_editar_descripcion,
        'puede_agendar_cita': puede_agendar_cita,
        'puede_formular_caso': puede_formular_caso,
        'puede_enviar_coordinador_tecnico_pedagogico': puede_enviar_coordinador_tecnico_pedagogico,
        'puede_formular_ajustes': puede_formular_ajustes,
        'puede_enviar_asesor_pedagogico': puede_enviar_asesor_pedagogico,
        'puede_devolver_a_encargado_inclusion': puede_devolver_a_encargado_inclusion,
        'puede_enviar_a_director': puede_enviar_a_director,
        'puede_devolver_a_coordinador_tecnico_pedagogico': puede_devolver_a_coordinador_tecnico_pedagogico,
        'puede_editar_ajustes_asesor': puede_editar_ajustes_asesor,
        'puede_aprobar': puede_aprobar,
        'puede_rechazar': puede_rechazar,
        'es_docente': es_docente,  # Para ocultar acciones de edición en el template
    }
    
    # Permiso para subir archivos (solo Encargado de Inclusión y solo en estados editables)
    puede_subir_archivo = rol_nombre == ROL_COORDINADORA and caso_editable_encargado
    
    context['puede_subir_archivo'] = puede_subir_archivo
    context['caso_editable_encargado'] = caso_editable_encargado
    context['caso_editable_coordinador_tecnico'] = caso_editable_coordinador_tecnico
    context['puede_editar_ajustes_coordinador'] = puede_editar_ajustes_coordinador
    context['puede_eliminar_ajustes_coordinador'] = puede_eliminar_ajustes_coordinador
    context['caso_editable_asesor'] = caso_editable_asesor
    context['puede_eliminar_ajustes_asesor'] = puede_eliminar_ajustes_asesor
    context['caso_editable_director'] = caso_editable_director
    context['puede_desactivar_caso'] = puede_desactivar_caso
    
    return render(request, 'SIAPE/detalle_casos_encargado_inclusion.html', context)

@require_POST
@login_required
def subir_archivo_caso(request, solicitud_id):
    """
    Permite al Encargado de Inclusión subir un archivo al caso.
    Solo puede hacerlo cuando el caso está en estados que le corresponden.
    """
    # Estados permitidos para edición del Encargado de Inclusión
    ESTADOS_EDITABLES_ENCARGADO = ['pendiente_entrevista', 'pendiente_formulacion_caso']
    
    # 1. Verificar Permiso
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
    except AttributeError:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
    
    # 2. Obtener la solicitud
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # 3. Verificar que el caso esté en un estado editable para el Encargado de Inclusión
    if solicitud.estado not in ESTADOS_EDITABLES_ENCARGADO and not request.user.is_superuser:
        messages.error(request, 'No puedes modificar este caso porque ya fue enviado al siguiente rol.')
        return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
    
    # 4. Validar que el encargado de inclusión tenga acceso a este caso
    if solicitud.coordinadora_asignada != perfil and not request.user.is_superuser:
        messages.error(request, 'No tienes permisos para subir archivos a este caso.')
        return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
    
    # 4. Obtener el archivo
    archivo = request.FILES.get('archivo')
    if not archivo:
        messages.error(request, 'Debe seleccionar un archivo.')
        return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
    
    # 5. Validar el archivo (usar la misma validación del serializer)
    import os
    
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
    ALLOWED_EXTENSIONS = [
        '.pdf', '.doc', '.docx', '.xls', '.xlsx', 
        '.jpg', '.jpeg', '.png', '.gif', '.txt'
    ]
    
    # Verificar tamaño
    if archivo.size > MAX_FILE_SIZE:
        messages.error(request, f'El archivo es demasiado grande. Tamaño máximo: 10 MB.')
        return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
    
    # Verificar extensión
    ext = os.path.splitext(archivo.name)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        messages.error(
            request, 
            f'Tipo de archivo no permitido. Extensiones permitidas: {", ".join(ALLOWED_EXTENSIONS)}'
        )
        return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
    
    # Verificar nombre del archivo (prevenir path traversal)
    filename = os.path.basename(archivo.name)
    if '..' in filename or '/' in filename or '\\' in filename:
        messages.error(request, 'Nombre de archivo inválido.')
        return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
    
    # 6. Crear la evidencia
    try:
        estudiante = solicitud.estudiantes
        evidencia = Evidencias.objects.create(
            archivo=archivo,
            estudiantes=estudiante,
            solicitudes=solicitud
        )
        messages.success(request, f'Archivo "{filename}" subido exitosamente.')
    except Exception as e:
        logger.error(f"Error al subir archivo: {str(e)}")
        messages.error(request, f'Error al subir el archivo: {str(e)}')
    
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)

@login_required
def detalle_casos_coordinador_tecnico_pedagogico(request, solicitud_id):
    """
    Vista para mostrar el detalle de un caso para el Coordinador Técnico Pedagógico.
    Es un wrapper que redirige a la misma vista pero con un contexto diferente.
    """
    # Verificar que el usuario es Coordinador Técnico Pedagógico
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None
    
    # Llamar a la misma función pero con el contexto específico
    return detalle_casos_encargado_inclusion(request, solicitud_id)

@require_POST
@login_required
def formular_ajuste_coordinador_tecnico_pedagogico(request, solicitud_id):
    """
    Vista para que el Coordinador Técnico Pedagógico pueda crear y asignar ajustes a un caso.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # Verificar que el caso está en el estado correcto
    if solicitud.estado != 'pendiente_formulacion_ajustes':
        messages.error(request, 'Este caso no está en estado de formulación de ajustes.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)

    # 3. --- Obtener Datos del Formulario ---
    descripcion = request.POST.get('descripcion', '').strip()
    categoria_id = request.POST.get('categoria_id', '')
    nueva_categoria = request.POST.get('nueva_categoria', '').strip()

    # 4. --- Validaciones ---
    if not descripcion:
        messages.error(request, 'La descripción del ajuste es requerida.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)

    # Verificar si se seleccionó "nueva" o si hay una categoría seleccionada
    crear_nueva_categoria = categoria_id == 'nueva' or (not categoria_id and nueva_categoria)
    
    if not categoria_id and not nueva_categoria:
        messages.error(request, 'Debe seleccionar una categoría o crear una nueva.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)

    if categoria_id and categoria_id != 'nueva' and nueva_categoria:
        messages.error(request, 'No puede seleccionar una categoría existente y crear una nueva a la vez.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)

    if crear_nueva_categoria and not nueva_categoria:
        messages.error(request, 'Debe proporcionar el nombre de la nueva categoría.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)

    try:
        # 5. --- Obtener o Crear Categoría ---
        if crear_nueva_categoria:
            if not nueva_categoria:
                messages.error(request, 'Debe proporcionar el nombre de la nueva categoría.')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            categoria, created = CategoriasAjustes.objects.get_or_create(
                nombre_categoria=nueva_categoria.strip().capitalize()
            )
            if created:
                messages.info(request, f'Categoría "{categoria.nombre_categoria}" creada exitosamente.')
        else:
            if not categoria_id or categoria_id == 'nueva':
                messages.error(request, 'Debe seleccionar una categoría válida.')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            categoria = get_object_or_404(CategoriasAjustes, id=categoria_id)

        # 6. --- Crear Ajuste Razonable ---
        ajuste_razonable = AjusteRazonable.objects.create(
            descripcion=descripcion,
            categorias_ajustes=categoria
        )

        # 7. --- Asignar Ajuste a la Solicitud ---
        AjusteAsignado.objects.create(
            ajuste_razonable=ajuste_razonable,
            solicitudes=solicitud
        )

        # 8. --- Asignar Coordinador Técnico Pedagógico al caso si no está asignado ---
        if not solicitud.coordinador_tecnico_pedagogico_asignado:
            solicitud.coordinador_tecnico_pedagogico_asignado = perfil
            solicitud.save()

        messages.success(request, 'Ajuste formulado y asignado exitosamente.')

    except Exception as e:
        logger.error(f"Error al formular ajuste: {str(e)}")
        messages.error(request, f'Error al formular el ajuste: {str(e)}')

    # 9. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud_id)

@require_POST
@login_required
def editar_ajuste_coordinador_tecnico_pedagogico(request, ajuste_asignado_id):
    """
    Vista para que el Coordinador Técnico Pedagógico pueda editar un ajuste ya asignado.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener el Ajuste Asignado ---
    ajuste_asignado = get_object_or_404(AjusteAsignado, id=ajuste_asignado_id)
    solicitud = ajuste_asignado.solicitudes
    
    # Verificar que el caso está en el estado correcto
    if solicitud.estado != 'pendiente_formulacion_ajustes':
        messages.error(request, 'Solo se pueden editar ajustes de casos en estado de formulación de ajustes.')
        return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)

    # 3. --- Obtener Datos del Formulario ---
    descripcion = request.POST.get('descripcion', '').strip()
    categoria_id = request.POST.get('categoria_id', '')
    nueva_categoria = request.POST.get('nueva_categoria', '').strip()

    # 4. --- Validaciones ---
    if not descripcion:
        messages.error(request, 'La descripción del ajuste es requerida.')
        return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)

    # Verificar si se seleccionó "nueva" o si hay una categoría seleccionada
    crear_nueva_categoria = categoria_id == 'nueva' or (not categoria_id and nueva_categoria)
    
    if not categoria_id and not nueva_categoria:
        messages.error(request, 'Debe seleccionar una categoría o crear una nueva.')
        return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)

    if categoria_id and categoria_id != 'nueva' and nueva_categoria:
        messages.error(request, 'No puede seleccionar una categoría existente y crear una nueva a la vez.')
        return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)

    if crear_nueva_categoria and not nueva_categoria:
        messages.error(request, 'Debe proporcionar el nombre de la nueva categoría.')
        return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)

    try:
        # 5. --- Obtener o Crear Categoría ---
        if crear_nueva_categoria:
            if not nueva_categoria:
                messages.error(request, 'Debe proporcionar el nombre de la nueva categoría.')
                return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)
            categoria, created = CategoriasAjustes.objects.get_or_create(
                nombre_categoria=nueva_categoria.strip().capitalize()
            )
            if created:
                messages.info(request, f'Categoría "{categoria.nombre_categoria}" creada exitosamente.')
        else:
            if not categoria_id or categoria_id == 'nueva':
                messages.error(request, 'Debe seleccionar una categoría válida.')
                return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)
            categoria = get_object_or_404(CategoriasAjustes, id=categoria_id)

        # 6. --- Actualizar Ajuste Razonable ---
        ajuste_razonable = ajuste_asignado.ajuste_razonable
        ajuste_razonable.descripcion = descripcion
        ajuste_razonable.categorias_ajustes = categoria
        ajuste_razonable.save()

        messages.success(request, 'Ajuste actualizado exitosamente.')

    except Exception as e:
        logger.error(f"Error al editar ajuste: {str(e)}")
        messages.error(request, f'Error al editar el ajuste: {str(e)}')

    # 7. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)

@require_POST
@login_required
def eliminar_ajuste_coordinador_tecnico_pedagogico(request, ajuste_asignado_id):
    """
    Vista para que el Coordinador Técnico Pedagógico pueda eliminar un ajuste asignado.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener el Ajuste Asignado ---
    ajuste_asignado = get_object_or_404(AjusteAsignado, id=ajuste_asignado_id)
    solicitud = ajuste_asignado.solicitudes
    
    # Verificar que el caso está en el estado correcto
    if solicitud.estado != 'pendiente_formulacion_ajustes':
        messages.error(request, 'Solo se pueden eliminar ajustes de casos en estado de formulación de ajustes.')
        return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud.id)

    try:
        # 3. --- Eliminar el Ajuste Asignado y el Ajuste Razonable asociado ---
        ajuste_razonable = ajuste_asignado.ajuste_razonable
        solicitud_id = solicitud.id
        ajuste_asignado.delete()
        # También eliminamos el ajuste razonable si no está siendo usado por otros ajustes asignados
        if not AjusteAsignado.objects.filter(ajuste_razonable=ajuste_razonable).exists():
            ajuste_razonable.delete()
        
        messages.success(request, 'Ajuste eliminado exitosamente.')

    except Exception as e:
        logger.error(f"Error al eliminar ajuste: {str(e)}")
        messages.error(request, f'Error al eliminar el ajuste: {str(e)}')

    # 4. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud_id)

@require_POST
@login_required
def enviar_a_coordinador_tecnico_pedagogico(request, solicitud_id):
    """
    Vista para que el Encargado de Inclusión envíe el caso al Coordinador Técnico Pedagógico.
    Cambia el estado del caso de 'pendiente_formulacion_caso' a 'pendiente_formulacion_ajustes'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # 3. --- Verificar que el caso está en el estado correcto ---
    if solicitud.estado != 'pendiente_formulacion_caso':
        messages.error(request, 'Este caso no está en estado de formulación del caso. Solo se pueden enviar casos después de formular el caso.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)
    
    try:
        # 4. --- Cambiar el estado del caso ---
        solicitud.estado = 'pendiente_formulacion_ajustes'
        # Nota: No asignamos coordinador_tecnico_pedagogico_asignado aquí porque cualquier Coordinador Técnico Pedagógico
        # puede trabajar en casos pendientes. Se asignará automáticamente cuando formulen el primer ajuste.
        solicitud.save()
        
        messages.success(request, 'Caso enviado al Coordinador Técnico Pedagógico exitosamente. El caso ahora está pendiente de formulación de ajustes.')
        
    except Exception as e:
        logger.error(f"Error al enviar caso a asesora técnica: {str(e)}")
        messages.error(request, f'Error al enviar el caso: {str(e)}')
    
    # 5. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)

@require_POST
@login_required
def enviar_a_asesor_pedagogico(request, solicitud_id):
    """
    Vista para que el Coordinador Técnico Pedagógico envíe el caso al siguiente estado
    (pendiente_preaprobacion) después de formular los ajustes.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # Verificar que el caso está en el estado correcto
    if solicitud.estado != 'pendiente_formulacion_ajustes':
        messages.error(request, 'Este caso no está en estado de formulación de ajustes.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)

    # 3. --- Verificar que hay ajustes asignados ---
    ajustes_count = AjusteAsignado.objects.filter(solicitudes=solicitud).count()
    if ajustes_count == 0:
        messages.error(request, 'Debe formular al menos un ajuste antes de enviar el caso al Asesor Pedagógico.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)

    try:
        # 4. --- Cambiar el estado del caso ---
        solicitud.estado = 'pendiente_preaprobacion'
        solicitud.save()
        
        messages.success(request, 'Caso enviado al Asesor Pedagógico exitosamente. El caso ahora está pendiente de preaprobación.')
        
    except Exception as e:
        logger.error(f"Error al enviar caso a asesor pedagógico: {str(e)}")
        messages.error(request, f'Error al enviar el caso: {str(e)}')

    # 5. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud_id)

@require_POST
@login_required
def devolver_a_encargado_inclusion(request, solicitud_id):
    """
    Vista para que el Coordinador Técnico Pedagógico devuelva el caso al Encargado de Inclusión.
    Cambia el estado del caso de 'pendiente_formulacion_ajustes' a 'pendiente_formulacion_caso'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # 3. --- Verificar que el caso está en el estado correcto ---
    if solicitud.estado != 'pendiente_formulacion_ajustes':
        messages.error(request, 'Este caso no está en estado de formulación de ajustes. Solo se pueden devolver casos pendientes de formulación de ajustes.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)
    
    try:
        # 4. --- Cambiar el estado del caso ---
        solicitud.estado = 'pendiente_formulacion_caso'
        solicitud.save()
        
        messages.success(request, 'Caso devuelto al Encargado de Inclusión exitosamente. El caso ahora está pendiente de formulación del caso.')
        
    except Exception as e:
        logger.error(f"Error al devolver caso a coordinadora: {str(e)}")
        messages.error(request, f'Error al devolver el caso: {str(e)}')
    
    # 5. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_coordinador_tecnico_pedagogico', solicitud_id=solicitud_id)

@require_POST
@login_required
def enviar_a_director(request, solicitud_id):
    """
    Vista para que el Asesor Pedagógico envíe el caso al Director.
    Cambia el estado del caso de 'pendiente_preaprobacion' a 'pendiente_aprobacion'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_ASESOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # 3. --- Verificar que el caso está en el estado correcto ---
    if solicitud.estado != 'pendiente_preaprobacion':
        messages.error(request, 'Este caso no está en estado de preaprobación. Solo se pueden enviar casos pendientes de preaprobación.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)
    
    try:
        # 4. --- Cambiar el estado del caso ---
        solicitud.estado = 'pendiente_aprobacion'
        solicitud.save()
        
        messages.success(request, 'Caso enviado al Director exitosamente. El caso ahora está pendiente de aprobación.')
        
    except Exception as e:
        logger.error(f"Error al enviar caso a director: {str(e)}")
        messages.error(request, f'Error al enviar el caso: {str(e)}')
    
    # 5. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)

@require_POST
@login_required
def devolver_a_coordinador_tecnico_pedagogico(request, solicitud_id):
    """
    Vista para que el Asesor Pedagógico devuelva el caso al Asesor Técnico Pedagógico.
    Cambia el estado del caso de 'pendiente_preaprobacion' a 'pendiente_formulacion_ajustes'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_ASESOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # 3. --- Verificar que el caso está en el estado correcto ---
    if solicitud.estado != 'pendiente_preaprobacion':
        messages.error(request, 'Este caso no está en estado de preaprobación. Solo se pueden devolver casos pendientes de preaprobación.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)
    
    try:
        # 4. --- Cambiar el estado del caso ---
        solicitud.estado = 'pendiente_formulacion_ajustes'
        solicitud.save()
        
        messages.success(request, 'Caso devuelto al Asesor Técnico Pedagógico exitosamente. El caso ahora está pendiente de formulación de ajustes.')
        
    except Exception as e:
        logger.error(f"Error al devolver caso a asesor técnico: {str(e)}")
        messages.error(request, f'Error al devolver el caso: {str(e)}')
    
    # 5. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)

@require_POST
@login_required
def editar_ajuste_asesor(request, ajuste_asignado_id):
    """
    Vista para que la Asesora Pedagógica pueda editar un ajuste ya asignado
    cuando el caso está en estado 'pendiente_preaprobacion'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_ASESOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener el Ajuste Asignado ---
    ajuste_asignado = get_object_or_404(AjusteAsignado, id=ajuste_asignado_id)
    solicitud = ajuste_asignado.solicitudes
    
    # Verificar que el caso está en el estado correcto
    if solicitud.estado != 'pendiente_preaprobacion':
        messages.error(request, 'Solo se pueden editar ajustes de casos en estado de preaprobación.')
        return redirect('detalle_caso', solicitud_id=solicitud.id)

    # 3. --- Obtener Datos del Formulario ---
    descripcion = request.POST.get('descripcion', '').strip()
    categoria_id = request.POST.get('categoria_id', '')
    nueva_categoria = request.POST.get('nueva_categoria', '').strip()

    # 4. --- Validaciones ---
    if not descripcion:
        messages.error(request, 'La descripción del ajuste es requerida.')
        return redirect('detalle_caso', solicitud_id=solicitud.id)

    # Verificar si se seleccionó "nueva" o si hay una categoría seleccionada
    crear_nueva_categoria = categoria_id == 'nueva' or (not categoria_id and nueva_categoria)
    
    if not categoria_id and not nueva_categoria:
        messages.error(request, 'Debe seleccionar una categoría o crear una nueva.')
        return redirect('detalle_caso', solicitud_id=solicitud.id)

    if categoria_id and categoria_id != 'nueva' and nueva_categoria:
        messages.error(request, 'No puede seleccionar una categoría existente y crear una nueva a la vez.')
        return redirect('detalle_caso', solicitud_id=solicitud.id)

    if crear_nueva_categoria and not nueva_categoria:
        messages.error(request, 'Debe proporcionar el nombre de la nueva categoría.')
        return redirect('detalle_caso', solicitud_id=solicitud.id)

    try:
        # 5. --- Obtener o Crear Categoría ---
        if crear_nueva_categoria:
            if not nueva_categoria:
                messages.error(request, 'Debe proporcionar el nombre de la nueva categoría.')
                return redirect('detalle_caso', solicitud_id=solicitud.id)
            categoria, created = CategoriasAjustes.objects.get_or_create(
                nombre_categoria=nueva_categoria.strip().capitalize()
            )
            if created:
                messages.info(request, f'Categoría "{categoria.nombre_categoria}" creada exitosamente.')
        else:
            if not categoria_id or categoria_id == 'nueva':
                messages.error(request, 'Debe seleccionar una categoría válida.')
                return redirect('detalle_caso', solicitud_id=solicitud.id)
            categoria = get_object_or_404(CategoriasAjustes, id=categoria_id)

        # 6. --- Actualizar Ajuste Razonable ---
        ajuste_razonable = ajuste_asignado.ajuste_razonable
        ajuste_razonable.descripcion = descripcion
        ajuste_razonable.categorias_ajustes = categoria
        ajuste_razonable.save()

        messages.success(request, 'Ajuste actualizado exitosamente.')

    except Exception as e:
        logger.error(f"Error al editar ajuste: {str(e)}")
        messages.error(request, f'Error al editar el ajuste: {str(e)}')

    # 7. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud.id)

@require_POST
@login_required
def eliminar_ajuste_asesor(request, ajuste_asignado_id):
    """
    Vista para que la Asesora Pedagógica pueda eliminar un ajuste asignado
    cuando el caso está en estado 'pendiente_preaprobacion'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_ASESOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener el Ajuste Asignado ---
    ajuste_asignado = get_object_or_404(AjusteAsignado, id=ajuste_asignado_id)
    solicitud = ajuste_asignado.solicitudes
    
    # Verificar que el caso está en el estado correcto
    if solicitud.estado != 'pendiente_preaprobacion':
        messages.error(request, 'Solo se pueden eliminar ajustes de casos en estado de preaprobación.')
        return redirect('detalle_caso', solicitud_id=solicitud.id)

    try:
        # 3. --- Eliminar el Ajuste Asignado y el Ajuste Razonable asociado ---
        ajuste_razonable = ajuste_asignado.ajuste_razonable
        solicitud_id = solicitud.id
        ajuste_asignado.delete()
        # También eliminamos el ajuste razonable si no está siendo usado por otros ajustes asignados
        if not AjusteAsignado.objects.filter(ajuste_razonable=ajuste_razonable).exists():
            ajuste_razonable.delete()
        
        messages.success(request, 'Ajuste eliminado exitosamente.')

    except Exception as e:
        logger.error(f"Error al eliminar ajuste: {str(e)}")
        messages.error(request, f'Error al eliminar el ajuste: {str(e)}')

    # 4. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)

@require_POST
@login_required
def aprobar_caso(request, solicitud_id):
    """
    Vista para que el Director apruebe el caso.
    Cambia el estado del caso de 'pendiente_aprobacion' a 'aprobado'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # 3. --- Verificar que el caso está en el estado correcto ---
    if solicitud.estado != 'pendiente_aprobacion':
        messages.error(request, 'Este caso no está en estado de aprobación. Solo se pueden aprobar casos pendientes de aprobación.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)
    
    try:
        # 4. --- Cambiar el estado del caso ---
        solicitud.estado = 'aprobado'
        solicitud.save()
        
        messages.success(request, 'Caso aprobado exitosamente. El caso ha sido aprobado e informado.')
        
    except Exception as e:
        logger.error(f"Error al aprobar caso: {str(e)}")
        messages.error(request, f'Error al aprobar el caso: {str(e)}')
    
    # 5. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)

@require_POST
@login_required
def rechazar_caso(request, solicitud_id):
    """
    Vista para que el Director rechace el caso.
    Cambia el estado del caso de 'pendiente_aprobacion' a 'pendiente_preaprobacion' 
    (vuelve a Asesoría Pedagógica para evaluación de corrección).
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # 3. --- Verificar que el caso está en el estado correcto ---
    if solicitud.estado != 'pendiente_aprobacion':
        messages.error(request, 'Este caso no está en estado de aprobación. Solo se pueden rechazar casos pendientes de aprobación.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)
    
    try:
        # 4. --- Cambiar el estado del caso (vuelve a Asesoría Pedagógica) ---
        solicitud.estado = 'pendiente_preaprobacion'
        solicitud.save()
        
        messages.warning(request, 'Caso rechazado. El caso ha sido devuelto a Asesoría Pedagógica para evaluación de corrección o archivo.')
        
    except Exception as e:
        logger.error(f"Error al rechazar caso: {str(e)}")
        messages.error(request, f'Error al rechazar el caso: {str(e)}')
    
    # 5. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)

@require_POST
@login_required
def desactivar_caso(request, solicitud_id):
    """
    Vista para que el Director desactive un caso aprobado y lo envíe a revisión.
    Cambia el estado del caso de 'aprobado' a 'pendiente_preaprobacion'.
    Esto permite que un caso ya aprobado sea reevaluado por el equipo.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener la Solicitud ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    
    # 3. --- Verificar que el caso está en estado aprobado ---
    if solicitud.estado != 'aprobado':
        messages.error(request, 'Solo se pueden desactivar casos que estén aprobados.')
        return redirect('detalle_caso', solicitud_id=solicitud_id)
    
    try:
        # 4. --- Cambiar el estado del caso (vuelve a Asesoría Pedagógica para revisión) ---
        solicitud.estado = 'pendiente_preaprobacion'
        solicitud.save()
        
        messages.warning(request, 'Caso desactivado. El caso ha sido enviado a revisión por Asesoría Pedagógica.')
        
    except Exception as e:
        logger.error(f"Error al desactivar caso: {str(e)}")
        messages.error(request, f'Error al desactivar el caso: {str(e)}')
    
    # 5. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)

@require_POST
@login_required
def aprobar_ajuste_director(request, ajuste_asignado_id):
    """
    Vista para que el Director apruebe un ajuste individual.
    Cambia el estado del ajuste de 'pendiente' a 'aprobado'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener el Ajuste Asignado ---
    ajuste_asignado = get_object_or_404(AjusteAsignado, id=ajuste_asignado_id)
    solicitud = ajuste_asignado.solicitudes
    
    # 3. --- Verificar que el caso está en el estado correcto ---
    if solicitud.estado != 'pendiente_aprobacion':
        messages.error(request, 'Este caso no está en estado de aprobación. Solo se pueden aprobar ajustes de casos pendientes de aprobación.')
        return redirect('detalle_caso', solicitud_id=solicitud.id)
    
    try:
        # 4. --- Obtener comentarios (opcional) ---
        comentarios = request.POST.get('comentarios_director', '').strip()
        
        # 5. --- Cambiar el estado del ajuste ---
        ajuste_asignado.estado_aprobacion = 'aprobado'
        ajuste_asignado.director_aprobador = perfil
        ajuste_asignado.fecha_aprobacion = timezone.now()
        if comentarios:
            ajuste_asignado.comentarios_director = comentarios
        ajuste_asignado.save()
        
        messages.success(request, f'Ajuste aprobado exitosamente: {ajuste_asignado.ajuste_razonable.descripcion[:50]}...')
        
    except Exception as e:
        logger.error(f"Error al aprobar ajuste: {str(e)}")
        messages.error(request, f'Error al aprobar el ajuste: {str(e)}')
    
    # 6. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud.id)

@require_POST
@login_required
def rechazar_ajuste_director(request, ajuste_asignado_id):
    """
    Vista para que el Director rechace un ajuste individual.
    Cambia el estado del ajuste de 'pendiente' a 'rechazado'.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
        perfil = None

    # 2. --- Obtener el Ajuste Asignado ---
    ajuste_asignado = get_object_or_404(AjusteAsignado, id=ajuste_asignado_id)
    solicitud = ajuste_asignado.solicitudes
    
    # 3. --- Verificar que el caso está en el estado correcto ---
    if solicitud.estado != 'pendiente_aprobacion':
        messages.error(request, 'Este caso no está en estado de aprobación. Solo se pueden rechazar ajustes de casos pendientes de aprobación.')
        return redirect('detalle_caso', solicitud_id=solicitud.id)
    
    try:
        # 4. --- Obtener comentarios (requerido para rechazo) ---
        comentarios = request.POST.get('comentarios_director', '').strip()
        
        if not comentarios:
            messages.error(request, 'Debe proporcionar un comentario al rechazar un ajuste.')
            return redirect('detalle_caso', solicitud_id=solicitud.id)
        
        # 5. --- Cambiar el estado del ajuste ---
        ajuste_asignado.estado_aprobacion = 'rechazado'
        ajuste_asignado.director_aprobador = perfil
        ajuste_asignado.fecha_aprobacion = timezone.now()
        ajuste_asignado.comentarios_director = comentarios
        ajuste_asignado.save()
        
        messages.warning(request, f'Ajuste rechazado: {ajuste_asignado.ajuste_razonable.descripcion[:50]}...')
        
    except Exception as e:
        logger.error(f"Error al rechazar ajuste: {str(e)}")
        messages.error(request, f'Error al rechazar el ajuste: {str(e)}')
    
    # 6. --- Redirigir de vuelta al detalle ---
    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud.id)

@require_POST # Solo permite esta vista vía POST
@login_required
def actualizar_descripcion_caso(request, solicitud_id):
    """
    Vista de acción para actualizar la descripción principal (el "caso")
    de una solicitud.
    """
    # 1. --- Verificación de Permisos (Ampliado) ---
    try:
        perfil = request.user.perfil
        # Solo Encargado de Inclusión, Asesor Pedagógico y Admin pueden editar la descripción del caso
        # El Coordinador Técnico Pedagógico NO puede editar el caso formulado por el Encargado de Inclusión
        ROLES_PERMITIDOS = [
            ROL_COORDINADORA,
            ROL_ASESOR,
            ROL_ADMIN
        ]
        if perfil.rol.nombre_rol not in ROLES_PERMITIDOS:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('detalle_caso', solicitud_id=solicitud_id)
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')

    # 2. --- Actualizar la Descripción ---
    solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
    nueva_descripcion = request.POST.get('descripcion_caso')
    
    if nueva_descripcion is not None:
        solicitud.descripcion = nueva_descripcion
        solicitud.save()
        messages.success(request, 'Descripción del caso actualizada correctamente.')
    else:
        messages.error(request, 'No se recibió la descripción.')
        
    # 3. --- Determinar la URL de redirección según el rol ---
    try:
        perfil = request.user.perfil
        rol_nombre = perfil.rol.nombre_rol if perfil else None
        if rol_nombre == ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            return redirect('detalle_caso', solicitud_id=solicitud_id)
        elif rol_nombre == ROL_COORDINADORA:
            return redirect('detalle_caso', solicitud_id=solicitud_id)
        else:
            return redirect('detalle_caso', solicitud_id=solicitud_id)
    except AttributeError:
        return redirect('detalle_caso', solicitud_id=solicitud_id)

@login_required
def panel_control_encargado_inclusion(request):
    """
    Panel de control para el Encargado de Inclusión.
    Muestra citas (hoy, semana), calendario interactivo y acciones de cita.
    """
    try:
        if request.user.perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para acceder a este panel.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    perfil_coordinadora = request.user.perfil
    
    # 1. --- Definición de Fechas ---
    now = timezone.localtime(timezone.now())
    today = now.date()
    start_of_today = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    end_of_today = timezone.make_aware(datetime.combine(today, datetime.max.time()))
    
    start_of_week = today - timedelta(days=today.weekday()) # Lunes
    end_of_week = start_of_week + timedelta(days=6) # Domingo
    start_of_week_dt = timezone.make_aware(datetime.combine(start_of_week, datetime.min.time()))
    end_of_week_dt = timezone.make_aware(datetime.combine(end_of_week, datetime.max.time()))

    # 2. --- Obtener Citas para Todos los Encargados de Inclusión (Rol Completo) ---
    
    # Verificar que el perfil de la coordinadora existe
    if not perfil_coordinadora:
        messages.error(request, 'Error: No se pudo obtener el perfil de la coordinadora.')
        return redirect('home')
    
    # Base de entrevistas para todas las coordinadoras del rol
    # Todas las coordinadoras deben ver todas las entrevistas agendadas del rol
    todas_las_coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
    entrevistas_coordinadora = Entrevistas.objects.filter(
        coordinadora__in=todas_las_coordinadoras
    ).exclude(coordinadora__isnull=True).select_related('solicitudes', 'solicitudes__estudiantes')

    # A. Citas del Día
    citas_hoy_list = entrevistas_coordinadora.filter(
        fecha_entrevista__range=(start_of_today, end_of_today)
    ).order_by('fecha_entrevista')

    # B. Citas de la Semana (Próximas de esta semana)
    citas_semana_list = entrevistas_coordinadora.filter(
        fecha_entrevista__range=(now, end_of_week_dt),
        estado='pendiente'
    ).order_by('fecha_entrevista')

    # C. Citas Pendientes de Confirmar (para la sección de "asistencia")
    citas_pendientes_confirmar = entrevistas_coordinadora.filter(
        estado='pendiente',
        fecha_entrevista__lt=now # Citas que ya pasaron
    ).order_by('fecha_entrevista')

    # D. Historial de Citas Pasadas (de la semana actual)
    citas_pasadas_semana = entrevistas_coordinadora.filter(
        fecha_entrevista__lt=now, # Citas pasadas
        fecha_entrevista__gte=start_of_week_dt, # Pero dentro de esta semana
        estado__in=['realizada', 'no_asistio']
    ).order_by('-fecha_entrevista')
    
    # 3. --- Obtener Datos para el Calendario (de TODAS las coordinadoras) ---
    todas_las_coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
    todas_las_entrevistas = Entrevistas.objects.filter(
        coordinadora__in=todas_las_coordinadoras
    ).select_related('solicitudes', 'solicitudes__estudiantes', 'coordinadora').order_by('fecha_entrevista')
    
    fechas_con_citas = set()
    citas_data = []
    
    for entrevista in todas_las_entrevistas:
        fecha_str = timezone.localtime(entrevista.fecha_entrevista).strftime('%Y-%m-%d')
        hora_str = timezone.localtime(entrevista.fecha_entrevista).strftime('%H:%M')
        fechas_con_citas.add(fecha_str)
        
        citas_data.append({
            'fecha': fecha_str,
            'hora': hora_str,
            'estudiante': f"{entrevista.solicitudes.estudiantes.nombres} {entrevista.solicitudes.estudiantes.apellidos}",
            'asunto': entrevista.solicitudes.asunto,
            'estado': entrevista.get_estado_display(),
            'estado_key': entrevista.estado,
        })
    
    fechas_citas_json = json.dumps(list(fechas_con_citas))
    citas_data_json = json.dumps(citas_data)
    
    # 4. --- Obtener feriados del año actual para el calendario ---
    from datetime import date
    import calendar as cal_module
    hoy = timezone.localtime(timezone.now()).date()
    year = hoy.year
    month = hoy.month
    
    feriados_chile = holidays.Chile(years=year)
    feriados_mes = []
    for dia in range(1, cal_module.monthrange(year, month)[1] + 1):
        fecha_dia = date(year, month, dia)
        if fecha_dia in feriados_chile:
            nombre_feriado = feriados_chile.get(fecha_dia)
            nombre_espanol = traducir_feriado_chileno(nombre_feriado)
            feriados_mes.append({
                "fecha": fecha_dia.strftime('%Y-%m-%d'),
                "nombre": nombre_espanol
            })
    
    feriados_json = json.dumps(feriados_mes)
    
    # 5. --- Datos para Modales ---
    categorias_ajustes = CategoriasAjustes.objects.all().order_by('nombre_categoria')
    
    context = {
        'citas_hoy_list': citas_hoy_list,
        'citas_semana_list': citas_semana_list,
        'fechas_citas_json': fechas_citas_json,
        'citas_data_json': citas_data_json,
        'feriados_json': feriados_json,
        'citas_pendientes_confirmar': citas_pendientes_confirmar,
        'citas_pasadas_semana_list': citas_pasadas_semana,
        'categorias_ajustes': categorias_ajustes,
        
        # Necesitaremos estos datos que antes estaban en esta vista pero
        # que ahora no estamos consultando. Los dejamos vacíos por ahora
        # para no romper los modales.
        'casos_disponibles': Solicitudes.objects.none(),
        'casos_con_ajustes': [],
        'proximas_citas': [], # Reemplazado por citas_semana_list
        'citas_realizadas': [], # Reemplazado por citas_pasadas_semana_list
        'citas_no_asistio': [], # Reemplazado por citas_pasadas_semana_list
    }
    
    return render(request, 'SIAPE/panel_control_encargado_inclusion.html', context)

@login_required
def confirmar_cita_coordinadora(request, entrevista_id):
    """
    Permite al Encargado de Inclusión confirmar la asistencia (realizada o no asistió) 
    de una entrevista que ella gestiona.
    Solo puede hacerlo cuando el caso está en estados que le corresponden.
    """
    # Estados permitidos para edición del Encargado de Inclusión
    ESTADOS_EDITABLES_ENCARGADO = ['pendiente_entrevista', 'pendiente_formulacion_caso']
    
    # 1. Verificar Permiso
    try:
        if request.user.perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('panel_control_encargado_inclusion')
    except AttributeError:
        return redirect('home')
    
    # 1.1 Verificar estado del caso
    try:
        entrevista_temp = Entrevistas.objects.select_related('solicitudes').get(id=entrevista_id)
        if entrevista_temp.solicitudes.estado not in ESTADOS_EDITABLES_ENCARGADO and not request.user.is_superuser:
            messages.error(request, 'No puedes confirmar citas para este caso porque ya fue enviado al siguiente rol.')
            return redirect('panel_control_encargado_inclusion')
    except Entrevistas.DoesNotExist:
        messages.error(request, 'La cita no existe.')
        return redirect('panel_control_encargado_inclusion')

    # 2. Lógica de la Acción
    if request.method == 'POST':
        accion = request.POST.get('accion')
        notas_adicionales = request.POST.get('notas_adicionales', '')
        try:
            # Cualquier coordinadora del rol puede confirmar cualquier entrevista del rol
            todas_las_coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
            entrevista = get_object_or_404(Entrevistas, id=entrevista_id, coordinadora__in=todas_las_coordinadoras)
            
            if accion in ['realizada', 'no_asistio']:
                entrevista.estado = accion
                if notas_adicionales:
                    if entrevista.notas:
                        entrevista.notas += f"\n\n[Confirmación - {timezone.now().strftime('%d/%m/%Y %H:%M')}]: {notas_adicionales}"
                    else:
                        entrevista.notas = f"[Confirmación - {timezone.now().strftime('%d/%m/%Y %H:%M')}]: {notas_adicionales}"
                entrevista.save()
                
                if accion == 'realizada':
                    # Cuando la entrevista se marca como realizada, el caso pasa a pendiente_formulacion_caso
                    solicitud = entrevista.solicitudes
                    if solicitud.estado == 'pendiente_entrevista':
                        solicitud.estado = 'pendiente_formulacion_caso'
                        solicitud.save()
                    messages.success(request, 'Cita marcada como realizada. El caso ahora está pendiente de formulación del caso.')
                else:
                    messages.info(request, 'Cita marcada como no asistió. Puedes reagendarla.')
            else:
                messages.error(request, 'Acción no válida.')
        except Exception as e:
            messages.error(request, f'Error al confirmar la cita: {str(e)}')
            
    # 3. Redirigir siempre al panel de control
    return redirect('panel_control_encargado_inclusion')

@login_required
def gestionar_horarios_bloqueados(request):
    """
    Vista para que el Encargado de Inclusión gestione sus horarios bloqueados.
    Permite ver, crear y eliminar horarios bloqueados.
    """
    # 1. Verificar Permiso
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    # 2. Obtener horarios bloqueados de esta coordinadora
    horarios_bloqueados = HorarioBloqueado.objects.filter(
        coordinadora=perfil
    ).order_by('fecha_hora')
    
    # Filtrar solo horarios futuros o del día actual
    now = timezone.localtime(timezone.now())
    horarios_bloqueados = horarios_bloqueados.filter(fecha_hora__gte=now)
    
    # 3. Si es POST, crear nuevo horario bloqueado
    if request.method == 'POST':
        fecha_str = request.POST.get('fecha_bloqueo')  # Formato: YYYY-MM-DD
        hora_str = request.POST.get('hora_bloqueo')    # Formato: HH:MM
        motivo = request.POST.get('motivo', '').strip()
        
        if not fecha_str or not hora_str:
            messages.error(request, 'Debe seleccionar una fecha y un horario.')
        else:
            try:
                # Parsear fecha y hora por separado
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                hora_obj = datetime.strptime(hora_str, '%H:%M').time()
                
                # Normalizar la hora a hora en punto (minutos y segundos en 0)
                hora_normalizada = hora_obj.replace(minute=0, second=0, microsecond=0)
                
                # Combinar fecha y hora en un datetime aware
                fecha_hora = timezone.make_aware(datetime.combine(fecha_obj, hora_normalizada))
                
                # Verificar que no esté en el pasado
                if fecha_hora < now:
                    messages.error(request, 'No se pueden bloquear horarios en el pasado.')
                else:
                    # Verificar que no haya una cita en ese horario
                    tiene_cita = Entrevistas.objects.filter(
                        coordinadora=perfil,
                        fecha_entrevista=fecha_hora,
                        estado='pendiente'
                    ).exists()
                    
                    if tiene_cita:
                        messages.error(request, 'No se puede bloquear un horario que ya tiene una cita programada.')
                    else:
                        # Verificar que no esté ya bloqueado
                        ya_bloqueado = HorarioBloqueado.objects.filter(
                            coordinadora=perfil,
                            fecha_hora=fecha_hora
                        ).exists()
                        
                        if ya_bloqueado:
                            messages.error(request, 'Este horario ya está bloqueado.')
                        else:
                            # Crear el horario bloqueado
                            HorarioBloqueado.objects.create(
                                coordinadora=perfil,
                                fecha_hora=fecha_hora,
                                motivo=motivo
                            )
                            messages.success(request, 'Horario bloqueado exitosamente.')
            except ValueError as e:
                messages.error(request, f'Formato de fecha u hora inválido: {str(e)}')
            except Exception as e:
                messages.error(request, f'Error al bloquear el horario: {str(e)}')
        
        return redirect('gestionar_horarios_bloqueados')
    
    # 4. Preparar contexto
    context = {
        'horarios_bloqueados': horarios_bloqueados,
    }
    
    return render(request, 'SIAPE/gestionar_horarios_bloqueados.html', context)

@require_POST
@login_required
def eliminar_horario_bloqueado(request, horario_id):
    """
    Vista para que el Encargado de Inclusión elimine un horario bloqueado.
    """
    # 1. Verificar Permiso
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    # 2. Obtener y eliminar el horario bloqueado
    try:
        horario = get_object_or_404(HorarioBloqueado, id=horario_id, coordinadora=perfil)
        horario.delete()
        messages.success(request, 'Horario desbloqueado exitosamente.')
    except Exception as e:
        messages.error(request, f'Error al desbloquear el horario: {str(e)}')
    
    return redirect('gestionar_horarios_bloqueados')

@login_required
def editar_notas_cita_coordinadora(request, entrevista_id):
    """
    Permite al Encargado de Inclusión editar las notas de una cita que él gestiona.
    Solo puede hacerlo cuando el caso está en estados que le corresponden.
    """
    # Estados permitidos para edición del Encargado de Inclusión
    ESTADOS_EDITABLES_ENCARGADO = ['pendiente_entrevista', 'pendiente_formulacion_caso']
    
    # 1. Verificar Permiso
    try:
        if request.user.perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('panel_control_encargado_inclusion')
    except AttributeError:
        return redirect('home')
    
    # 1.1 Verificar estado del caso
    try:
        entrevista_temp = Entrevistas.objects.select_related('solicitudes').get(id=entrevista_id)
        if entrevista_temp.solicitudes.estado not in ESTADOS_EDITABLES_ENCARGADO and not request.user.is_superuser:
            messages.error(request, 'No puedes editar notas de citas para este caso porque ya fue enviado al siguiente rol.')
            return redirect('panel_control_encargado_inclusion')
    except Entrevistas.DoesNotExist:
        messages.error(request, 'La cita no existe.')
        return redirect('panel_control_encargado_inclusion')

    # 2. Lógica de la Acción
    if request.method == 'POST':
        nuevas_notas = request.POST.get('notas', '')
        try:
            # Cualquier coordinadora del rol puede editar notas de cualquier entrevista del rol
            todas_las_coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
            entrevista = get_object_or_404(Entrevistas, id=entrevista_id, coordinadora__in=todas_las_coordinadoras)
            entrevista.notas = nuevas_notas
            entrevista.save()
            messages.success(request, 'Notas actualizadas correctamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar las notas: {str(e)}')
            
    # 3. Redirigir
    return redirect('panel_control_encargado_inclusion')

@login_required
def agendar_cita_coordinadora(request):
    """
    Permite al Encargado de Inclusión agendar una nueva cita para un caso.
    Solo puede hacerlo cuando el caso está en estados que le corresponden.
    """
    # Estados permitidos para edición del Encargado de Inclusión
    ESTADOS_EDITABLES_ENCARGADO = ['pendiente_entrevista', 'pendiente_formulacion_caso']
    
    # 1. Verificar Permiso
    try:
        if request.user.perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')

    # 2. Lógica de la Acción
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        
        # Verificar estado del caso antes de agendar
        if solicitud_id:
            try:
                solicitud_check = Solicitudes.objects.get(id=solicitud_id)
                if solicitud_check.estado not in ESTADOS_EDITABLES_ENCARGADO and not request.user.is_superuser:
                    messages.error(request, 'No puedes agendar citas para este caso porque ya fue enviado al siguiente rol.')
                    return redirect('detalle_casos_encargado_inclusion', solicitud_id=solicitud_id)
            except Solicitudes.DoesNotExist:
                pass  # Se manejará más adelante en el flujo
        # Obtener valores y asegurarse de que sean strings (no listas)
        fecha_raw = request.POST.get('fecha_agendar', '')
        hora_raw = request.POST.get('hora_agendar', '')
        fecha_str = fecha_raw.strip() if fecha_raw else ''
        hora_str = hora_raw.strip() if hora_raw else ''
        modalidad = request.POST.get('modalidad', '')
        notas = request.POST.get('notas', '')
        
        # Si llegaron como listas, tomar el primer elemento
        if isinstance(fecha_str, list):
            fecha_str = fecha_str[0].strip() if fecha_str else ''
        if isinstance(hora_str, list):
            hora_str = hora_str[0].strip() if hora_str else ''
        
        try:
            if not solicitud_id:
                messages.error(request, 'ID de solicitud no proporcionado.')
                return redirect('casos_generales')
            
            solicitud = get_object_or_404(Solicitudes, id=solicitud_id)
            
            # Validar que fecha_str y hora_str sean strings válidos y no vacíos
            if not fecha_str or not hora_str:
                messages.error(request, 'Debe seleccionar una fecha y un horario.')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            
            # Asegurarse de que son strings
            fecha_str = str(fecha_str)
            hora_str = str(hora_str)
            
            # Parsear fecha y hora por separado
            try:
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                hora_obj = datetime.strptime(hora_str, '%H:%M').time()
            except (ValueError, TypeError) as ve:
                messages.error(request, f'Formato de fecha u hora inválido.')
                logger.error(f'Error parseando fecha/hora: fecha_str={fecha_str}, hora_str={hora_str}, error={str(ve)}')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            
            # Normalizar la hora a hora en punto (minutos y segundos en 0)
            hora_normalizada = hora_obj.replace(minute=0, second=0, microsecond=0)
            
            # Combinar fecha y hora en un datetime naive primero
            fecha_hora_naive = datetime.combine(fecha_obj, hora_normalizada)
            
            # Verificar que el datetime naive sea válido
            if not isinstance(fecha_hora_naive, datetime):
                messages.error(request, 'Error al combinar fecha y hora.')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            
            # Convertir a datetime aware usando la zona horaria del sistema
            try:
                fecha_entrevista = timezone.make_aware(fecha_hora_naive)
            except (ValueError, TypeError) as e:
                messages.error(request, f'Error al procesar la fecha y hora seleccionadas.')
                logger.error(f'Error en make_aware: fecha_hora_naive={fecha_hora_naive}, tipo={type(fecha_hora_naive)}, error={str(e)}')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            
            # Verificar que no esté en el pasado
            now = timezone.localtime(timezone.now())
            if fecha_entrevista < now:
                messages.error(request, 'No se pueden agendar citas en el pasado.')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            
            # Buscar coordinadora disponible para el horario seleccionado
            todas_las_coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
            coordinadora_asignada = None
            
            from .models import HorarioBloqueado
            for coord in todas_las_coordinadoras:
                tiene_cita = Entrevistas.objects.filter(
                    coordinadora=coord,
                    fecha_entrevista=fecha_entrevista
                ).exists()
                tiene_horario_bloqueado = HorarioBloqueado.objects.filter(
                    coordinadora=coord,
                    fecha_hora=fecha_entrevista
                ).exists()
                if not tiene_cita and not tiene_horario_bloqueado:
                    coordinadora_asignada = coord
                    break
            
            # Si ninguna coordinadora está disponible, usar la primera (fallback)
            if not coordinadora_asignada:
                coordinadora_asignada = todas_las_coordinadoras.first()
            
            if not coordinadora_asignada:
                messages.error(request, 'No hay coordinadoras disponibles para agendar la cita.')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            
            # Verificar que no haya una cita ya agendada para esta solicitud en este horario
            cita_existente = Entrevistas.objects.filter(
                solicitudes=solicitud,
                fecha_entrevista=fecha_entrevista
            ).exists()
            
            if cita_existente:
                messages.error(request, 'Ya existe una cita agendada para este caso en ese horario.')
                return redirect('detalle_caso', solicitud_id=solicitud_id)
            
            # Crear la nueva entrevista
            nueva_entrevista = Entrevistas.objects.create(
                solicitudes=solicitud,
                coordinadora=coordinadora_asignada,
                fecha_entrevista=fecha_entrevista,
                modalidad=modalidad,
                notas=notas,
                estado='pendiente'
            )
            
            # Si el caso está en estado 'pendiente_entrevista', mantenerlo así
            # (no cambiar el estado automáticamente al agendar)
            
            messages.success(request, 'Cita agendada correctamente.')
        except ValueError as e:
            messages.error(request, f'Formato de fecha inválido: {str(e)}')
        except Exception as e:
            messages.error(request, f'Error al agendar la cita: {str(e)}')
    
    # 3. Redirigir al detalle del caso
    solicitud_id = request.POST.get('solicitud_id') if request.method == 'POST' else request.GET.get('solicitud_id')
    if solicitud_id:
        return redirect('detalle_caso', solicitud_id=solicitud_id)
    return redirect('casos_generales')

@login_required
def reagendar_cita_coordinadora(request, entrevista_id):
    """
    Permite al Encargado de Inclusión reagendar una cita (usualmente una que 'no asistió').
    Crea una nueva entrevista y actualiza la antigua.
    Solo puede hacerlo cuando el caso está en estados que le corresponden.
    """
    # Estados permitidos para edición del Encargado de Inclusión
    ESTADOS_EDITABLES_ENCARGADO = ['pendiente_entrevista', 'pendiente_formulacion_caso']
    
    # 1. Verificar Permiso
    try:
        if request.user.perfil.rol.nombre_rol != ROL_COORDINADORA:
            messages.error(request, 'No tienes permisos para realizar esta acción.')
            return redirect('panel_control_encargado_inclusion')
    except AttributeError:
        return redirect('home')
    
    # 1.1 Verificar estado del caso
    try:
        entrevista_temp = Entrevistas.objects.select_related('solicitudes').get(id=entrevista_id)
        if entrevista_temp.solicitudes.estado not in ESTADOS_EDITABLES_ENCARGADO and not request.user.is_superuser:
            messages.error(request, 'No puedes reagendar citas para este caso porque ya fue enviado al siguiente rol.')
            return redirect('detalle_casos_encargado_inclusion', solicitud_id=entrevista_temp.solicitudes.id)
    except Entrevistas.DoesNotExist:
        messages.error(request, 'La cita no existe.')
        return redirect('panel_control_encargado_inclusion')

    # 2. Lógica de la Acción
    if request.method == 'POST':
        # Obtener valores y asegurarse de que sean strings (no listas)
        fecha_raw = request.POST.get('fecha_reagendar', '')
        hora_raw = request.POST.get('hora_reagendar', '')
        fecha_str = fecha_raw.strip() if fecha_raw else ''
        hora_str = hora_raw.strip() if hora_raw else ''
        nueva_modalidad = request.POST.get('nueva_modalidad', '')
        notas_reagendamiento = request.POST.get('notas_reagendamiento', '')
        
        # Si llegaron como listas, tomar el primer elemento
        if isinstance(fecha_str, list):
            fecha_str = fecha_str[0].strip() if fecha_str else ''
        if isinstance(hora_str, list):
            hora_str = hora_str[0].strip() if hora_str else ''
        try:
            # Cualquier coordinadora del rol puede reagendar cualquier entrevista del rol
            todas_las_coordinadoras = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADORA)
            entrevista_original = get_object_or_404(Entrevistas, id=entrevista_id, coordinadora__in=todas_las_coordinadoras)
            
            # Validar que fecha_str y hora_str sean strings válidos y no vacíos
            if not fecha_str or not hora_str:
                messages.error(request, 'Debe seleccionar una fecha y un horario.')
                return redirect('detalle_caso', solicitud_id=entrevista_original.solicitudes.id)
            
            # Asegurarse de que son strings
            fecha_str = str(fecha_str)
            hora_str = str(hora_str)
            
            # Parsear fecha y hora por separado
            try:
                fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                hora_obj = datetime.strptime(hora_str, '%H:%M').time()
            except (ValueError, TypeError) as ve:
                messages.error(request, f'Formato de fecha u hora inválido.')
                logger.error(f'Error parseando fecha/hora en reagendar: fecha_str={fecha_str}, hora_str={hora_str}, error={str(ve)}')
                return redirect('detalle_caso', solicitud_id=entrevista_original.solicitudes.id)
            
            # Normalizar la hora a hora en punto (minutos y segundos en 0)
            hora_normalizada = hora_obj.replace(minute=0, second=0, microsecond=0)
            
            # Combinar fecha y hora en un datetime naive primero
            fecha_hora_naive = datetime.combine(fecha_obj, hora_normalizada)
            
            # Verificar que el datetime naive sea válido
            if not isinstance(fecha_hora_naive, datetime):
                messages.error(request, 'Error al combinar fecha y hora.')
                return redirect('detalle_caso', solicitud_id=entrevista_original.solicitudes.id)
            
            # Convertir a datetime aware usando la zona horaria del sistema
            try:
                nueva_fecha = timezone.make_aware(fecha_hora_naive)
            except (ValueError, TypeError) as e:
                messages.error(request, f'Error al procesar la fecha y hora seleccionadas.')
                logger.error(f'Error en make_aware (reagendar): fecha_hora_naive={fecha_hora_naive}, tipo={type(fecha_hora_naive)}, error={str(e)}')
                return redirect('detalle_caso', solicitud_id=entrevista_original.solicitudes.id)
            
            # Crear la nueva cita (mantenemos la misma coordinadora asignada originalmente)
            nueva_entrevista = Entrevistas.objects.create(
                solicitudes=entrevista_original.solicitudes, 
                coordinadora=entrevista_original.coordinadora, # Mantiene la coordinadora original
                fecha_entrevista=nueva_fecha, 
                modalidad=nueva_modalidad or entrevista_original.modalidad,
                notas=f"Reagendada desde cita del {entrevista_original.fecha_entrevista.strftime('%d/%m/%Y %H:%M')}. {notas_reagendamiento}" if notas_reagendamiento else f"Reagendada desde cita del {entrevista_original.fecha_entrevista.strftime('%d/%m/%Y %H:%M')}.",
                estado='pendiente' # La nueva cita está pendiente
            )
            
            # Actualizar la cita original a 'no asistió' si estaba 'pendiente'
            if entrevista_original.estado == 'pendiente':
                entrevista_original.estado = 'no_asistio' 
            # Si ya era 'no_asistio', se mantiene así.
            entrevista_original.save()
            
            messages.success(request, 'Cita reagendada correctamente.')
            return redirect('detalle_caso', solicitud_id=entrevista_original.solicitudes.id)
        except Exception as e:
            messages.error(request, f'Error al reagendar la cita: {str(e)}')
            # Intentar redirigir al caso si es posible, sino al panel
            try:
                entrevista_original = get_object_or_404(Entrevistas, id=entrevista_id)
                return redirect('detalle_caso', solicitud_id=entrevista_original.solicitudes.id)
            except:
                return redirect('panel_control_encargado_inclusion')
            
    # 3. Redirigir (si no es POST)
    return redirect('panel_control_encargado_inclusion')


# --- VISTA ASESOR PEDAGÓGICO ---
@login_required
def dashboard_asesor(request):
    """
    Dashboard principal para la Asesora Pedagógica.
    La Asesora Pedagógica es la Jefa del área de asesoría pedagógica,
    por ende debe monitorear la información de todos los casos.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_ASESOR:
            messages.error(request, 'No tienes permisos para acceder a este panel.')
            return redirect('home')
    except AttributeError:
        logger.warning(f"Usuario {request.user.email} sin perfil/rol intentó acceder al dashboard de asesora pedagógica.")
        return redirect('home')
    
    # 2. --- Configuración de Fechas ---
    now = timezone.localtime(timezone.now())
    today = now.date()
    
    # Cálculo de la semana (Lunes a Domingo)
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_of_week_dt = timezone.make_aware(datetime.combine(start_of_week, datetime.min.time()))
    end_of_week_dt = timezone.make_aware(datetime.combine(end_of_week, datetime.max.time()))
    
    # 3. --- Obtener Datos para KPIs ---
    
    # KPI 1: Casos nuevos (creados esta semana)
    casos_nuevos_semana = Solicitudes.objects.filter(
        created_at__range=(start_of_week_dt, end_of_week_dt)
    ).count()
    
    # KPI 2: Casos devueltos desde Director de Carrera
    # Casos que están en 'pendiente_preaprobacion' y que tienen ajustes asignados
    # (lo que indica que fueron preaprobados y enviados al Director, pero fueron rechazados/devueltos)
    # Esto es una aproximación: casos con ajustes que están en preaprobación
    casos_devueltos_director = Solicitudes.objects.filter(
        estado='pendiente_preaprobacion',
        ajusteasignado__isnull=False
    ).distinct().count()
    
    # KPI 3-9: Un KPI por cada estado de los casos
    # Crear lista de tuplas (estado_valor, estado_nombre, cantidad) para facilitar el acceso en el template
    estados_con_cantidad = []
    for estado_valor, estado_nombre in Solicitudes.ESTADO_CHOICES:
        cantidad = Solicitudes.objects.filter(estado=estado_valor).count()
        estados_con_cantidad.append({
            'valor': estado_valor,
            'nombre': estado_nombre,
            'cantidad': cantidad
        })
    
    # 4. --- Obtener Lista de Casos Pendientes de Preaprobación ---
    # Casos que están en estado 'pendiente_preaprobacion' y que requieren revisión por parte de la Asesora Pedagógica
    casos_pendientes_preaprobacion = Solicitudes.objects.filter(
        estado='pendiente_preaprobacion'
    ).select_related(
        'estudiantes',
        'estudiantes__carreras'
    ).order_by('-updated_at')[:10]  # Los más recientes primero, limitar a 10
    
    # 5. --- Preparar Contexto ---
    context = {
        'nombre_usuario': request.user.first_name,
        'kpis': {
            'casos_nuevos_semana': casos_nuevos_semana,
            'casos_devueltos_director': casos_devueltos_director,
        },
        'estados_con_cantidad': estados_con_cantidad,
        'casos_pendientes_preaprobacion': casos_pendientes_preaprobacion,
    }
    
    return render(request, 'SIAPE/dashboard_asesor.html', context)


@login_required
def estadisticas_asesor_pedagogico(request):
    """
    Vista completa de estadísticas para el Asesor Pedagógico.
    Incluye estadísticas por roles, ajustes, fechas, carreras y rendimiento del sistema.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_ASESOR:
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
    
    # 2. --- Obtener Rango de Tiempo Seleccionado ---
    rango_seleccionado = request.GET.get('rango', 'mes')  # mes, semestre, año, historico
    
    # 3. --- Configuración de Fechas según Rango ---
    now = timezone.localtime(timezone.now())
    today = now.date()
    
    fecha_inicio = None
    fecha_fin_dt = timezone.make_aware(datetime.combine(today, datetime.max.time()))
    
    if rango_seleccionado == 'mes':
        # Último mes (30 días)
        fecha_inicio = today - timedelta(days=30)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Mes'
    elif rango_seleccionado == 'semestre':
        # Último semestre (6 meses)
        fecha_inicio = today - timedelta(days=180)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Semestre'
    elif rango_seleccionado == 'año':
        # Último año
        fecha_inicio = today.replace(month=1, day=1)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Año'
    else:  # historico
        # Todo el histórico (sin filtro de fecha)
        fecha_inicio_dt = None
        rango_nombre = 'Histórico Completo'
    
    # Rango de fechas para análisis
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_of_month = today.replace(day=1)
    start_of_year = today.replace(month=1, day=1)
    
    # 3. --- ESTADÍSTICAS POR ESTADO ---
    total_casos_sistema = Solicitudes.objects.count()
    estados_stats = []
    for estado_valor, estado_nombre in Solicitudes.ESTADO_CHOICES:
        cantidad = Solicitudes.objects.filter(estado=estado_valor).count()
        porcentaje = round((cantidad / total_casos_sistema * 100) if total_casos_sistema > 0 else 0, 1)
        estados_stats.append({
            'valor': estado_valor,
            'nombre': estado_nombre,
            'cantidad': cantidad,
            'porcentaje': porcentaje
        })
    
    # 4. --- ESTADÍSTICAS POR ROLES (casos asignados) ---
    # Encargado de Inclusión
    casos_encargado_inclusion = Solicitudes.objects.filter(
        coordinadora_asignada__isnull=False
    ).count()
    
    # Coordinador Técnico Pedagógico
    casos_coordinador_tecnico = Solicitudes.objects.filter(
        coordinador_tecnico_pedagogico_asignado__isnull=False
    ).count()
    
    # Asesor Pedagógico
    casos_asesor_pedagogico = Solicitudes.objects.filter(
        asesor_pedagogico_asignado__isnull=False
    ).count()
    
    # Sin asignar
    casos_sin_asignar = Solicitudes.objects.filter(
        coordinadora_asignada__isnull=True,
        coordinador_tecnico_pedagogico_asignado__isnull=True,
        asesor_pedagogico_asignado__isnull=True
    ).count()
    
    roles_stats = {
        'encargado_inclusion': casos_encargado_inclusion,
        'coordinador_tecnico': casos_coordinador_tecnico,
        'asesor_pedagogico': casos_asesor_pedagogico,
        'sin_asignar': casos_sin_asignar,
    }
    
    # 5. --- ESTADÍSTICAS POR AJUSTES ---
    total_ajustes = AjusteAsignado.objects.count()
    ajustes_aprobados = AjusteAsignado.objects.filter(estado_aprobacion='aprobado').count()
    ajustes_rechazados = AjusteAsignado.objects.filter(estado_aprobacion='rechazado').count()
    ajustes_pendientes = AjusteAsignado.objects.filter(estado_aprobacion='pendiente').count()
    
    # Por categoría
    ajustes_por_categoria = {}
    for categoria in CategoriasAjustes.objects.all():
        total_cat = AjusteRazonable.objects.filter(categorias_ajustes=categoria).count()
        aprobados_cat = AjusteAsignado.objects.filter(
            ajuste_razonable__categorias_ajustes=categoria,
            estado_aprobacion='aprobado'
        ).count()
        ajustes_por_categoria[categoria.nombre_categoria] = {
            'total': total_cat,
            'aprobados': aprobados_cat,
            'tasa_aprobacion': round((aprobados_cat / total_cat * 100) if total_cat > 0 else 0, 1)
        }
    
    # 6. --- ESTADÍSTICAS POR FECHA (Tendencias) según Rango ---
    casos_por_mes = []
    casos_por_dia = []
    
    if rango_seleccionado == 'mes':
        # Último mes: mostrar por día
        for i in range(29, -1, -1):
            fecha_dia = today - timedelta(days=i)
            dia_inicio_dt = timezone.make_aware(datetime.combine(fecha_dia, datetime.min.time()))
            dia_fin_dt = timezone.make_aware(datetime.combine(fecha_dia, datetime.max.time()))
            
            cantidad = Solicitudes.objects.filter(
                created_at__range=(dia_inicio_dt, dia_fin_dt)
            ).count()
            
            casos_por_dia.append({
                'fecha': fecha_dia.strftime('%d/%m'),
                'cantidad': cantidad
            })
        
        # También por semana para el gráfico de mes
        semanas_en_mes = 4
        for i in range(semanas_en_mes - 1, -1, -1):
            semana_inicio = today - timedelta(days=(i * 7) + today.weekday())
            semana_fin = semana_inicio + timedelta(days=6)
            if semana_fin > today:
                semana_fin = today
            
            semana_inicio_dt = timezone.make_aware(datetime.combine(semana_inicio, datetime.min.time()))
            semana_fin_dt = timezone.make_aware(datetime.combine(semana_fin, datetime.max.time()))
            
            cantidad = Solicitudes.objects.filter(
                created_at__range=(semana_inicio_dt, semana_fin_dt)
            ).count()
            
            casos_por_mes.append({
                'mes': f'Sem {i+1}',
                'cantidad': cantidad
            })
    
    elif rango_seleccionado == 'semestre':
        # Último semestre: mostrar por mes
        for i in range(5, -1, -1):
            fecha_mes = today - timedelta(days=30 * i)
            mes_inicio = fecha_mes.replace(day=1)
            if i == 0:
                mes_fin = today
            else:
                siguiente_mes = mes_inicio + timedelta(days=32)
                mes_fin = siguiente_mes.replace(day=1) - timedelta(days=1)
            
            mes_inicio_dt = timezone.make_aware(datetime.combine(mes_inicio, datetime.min.time()))
            mes_fin_dt = timezone.make_aware(datetime.combine(mes_fin, datetime.max.time()))
            
            cantidad = Solicitudes.objects.filter(
                created_at__range=(mes_inicio_dt, mes_fin_dt)
            ).count()
            
            casos_por_mes.append({
                'mes': mes_inicio.strftime('%b %Y'),
                'cantidad': cantidad
            })
        
        # Por semana para el gráfico diario
        for i in range(25, -1, -1):  # Últimas 26 semanas
            semana_inicio = today - timedelta(days=(i * 7) + today.weekday())
            semana_fin = semana_inicio + timedelta(days=6)
            if semana_fin > today:
                semana_fin = today
            
            semana_inicio_dt = timezone.make_aware(datetime.combine(semana_inicio, datetime.min.time()))
            semana_fin_dt = timezone.make_aware(datetime.combine(semana_fin, datetime.max.time()))
            
            cantidad = Solicitudes.objects.filter(
                created_at__range=(semana_inicio_dt, semana_fin_dt)
            ).count()
            
            casos_por_dia.append({
                'fecha': semana_inicio.strftime('%d/%m'),
                'cantidad': cantidad
            })
    
    elif rango_seleccionado == 'año':
        # Último año: mostrar por mes
        for i in range(11, -1, -1):
            fecha_mes = today - timedelta(days=30 * i)
            mes_inicio = fecha_mes.replace(day=1)
            if i == 0:
                mes_fin = today
            else:
                siguiente_mes = mes_inicio + timedelta(days=32)
                mes_fin = siguiente_mes.replace(day=1) - timedelta(days=1)
            
            mes_inicio_dt = timezone.make_aware(datetime.combine(mes_inicio, datetime.min.time()))
            mes_fin_dt = timezone.make_aware(datetime.combine(mes_fin, datetime.max.time()))
            
            cantidad = Solicitudes.objects.filter(
                created_at__range=(mes_inicio_dt, mes_fin_dt)
            ).count()
            
            casos_por_mes.append({
                'mes': mes_inicio.strftime('%b %Y'),
                'cantidad': cantidad
            })
        
        # Por mes para el gráfico diario también
        casos_por_dia = casos_por_mes.copy()
    
    else:  # historico
        # Histórico completo: mostrar por año
        # Obtener el año más antiguo
        primer_caso = Solicitudes.objects.order_by('created_at').first()
        if primer_caso:
            año_inicio = primer_caso.created_at.year
            año_actual = today.year
            
            # Por año
            for año in range(año_inicio, año_actual + 1):
                año_inicio_dt = timezone.make_aware(datetime(año, 1, 1))
                if año == año_actual:
                    año_fin_dt = timezone.make_aware(datetime.combine(today, datetime.max.time()))
                else:
                    año_fin_dt = timezone.make_aware(datetime(año + 1, 1, 1)) - timedelta(seconds=1)
                
                cantidad = Solicitudes.objects.filter(
                    created_at__range=(año_inicio_dt, año_fin_dt)
                ).count()
                
                casos_por_mes.append({
                    'mes': str(año),
                    'cantidad': cantidad
                })
            
            # Por mes (últimos 24 meses)
            for i in range(23, -1, -1):
                fecha_mes = today - timedelta(days=30 * i)
                mes_inicio = fecha_mes.replace(day=1)
                if i == 0:
                    mes_fin = today
                else:
                    siguiente_mes = mes_inicio + timedelta(days=32)
                    mes_fin = siguiente_mes.replace(day=1) - timedelta(days=1)
                
                mes_inicio_dt = timezone.make_aware(datetime.combine(mes_inicio, datetime.min.time()))
                mes_fin_dt = timezone.make_aware(datetime.combine(mes_fin, datetime.max.time()))
                
                cantidad = Solicitudes.objects.filter(
                    created_at__range=(mes_inicio_dt, mes_fin_dt)
                ).count()
                
                casos_por_dia.append({
                    'fecha': mes_inicio.strftime('%b %Y'),
                    'cantidad': cantidad
                })
    
    # 7. --- ESTADÍSTICAS POR CARRERA ---
    carreras_stats = []
    for carrera in Carreras.objects.all():
        total_casos = Solicitudes.objects.filter(estudiantes__carreras=carrera).count()
        casos_aprobados = Solicitudes.objects.filter(
            estudiantes__carreras=carrera,
            estado='aprobado'
        ).count()
        casos_pendientes = Solicitudes.objects.filter(
            estudiantes__carreras=carrera
        ).exclude(estado__in=['aprobado', 'rechazado']).count()
        
        if total_casos > 0:
            carreras_stats.append({
                'nombre': carrera.nombre,
                'total': total_casos,
                'aprobados': casos_aprobados,
                'pendientes': casos_pendientes,
                'tasa_aprobacion': round((casos_aprobados / total_casos * 100), 1)
            })
    
    # Ordenar por total de casos
    carreras_stats.sort(key=lambda x: x['total'], reverse=True)
    
    # 8. --- ESTADÍSTICAS DE RENDIMIENTO ---
    # Tiempo promedio de resolución (casos aprobados)
    casos_aprobados_completos = Solicitudes.objects.filter(estado='aprobado')
    tiempos_resolucion = []
    for caso in casos_aprobados_completos:
        if caso.created_at and caso.updated_at:
            tiempo = (caso.updated_at - caso.created_at).days
            tiempos_resolucion.append(tiempo)
    
    tiempo_promedio = round(sum(tiempos_resolucion) / len(tiempos_resolucion), 1) if tiempos_resolucion else 0
    
    # Casos resueltos esta semana
    start_of_week_dt = timezone.make_aware(datetime.combine(start_of_week, datetime.min.time()))
    end_of_week_dt = timezone.make_aware(datetime.combine(end_of_week, datetime.max.time()))
    
    casos_resueltos_semana = Solicitudes.objects.filter(
        estado__in=['aprobado', 'rechazado'],
        updated_at__range=(start_of_week_dt, end_of_week_dt)
    ).count()
    
    # Tasa de aprobación general (valores iniciales, se filtrarán después si aplica)
    total_resueltos = Solicitudes.objects.filter(estado__in=['aprobado', 'rechazado']).count()
    total_aprobados = Solicitudes.objects.filter(estado='aprobado').count()
    tasa_aprobacion_general = round((total_aprobados / total_resueltos * 100) if total_resueltos > 0 else 0, 1)
    
    # Valores iniciales para totales
    total_casos = Solicitudes.objects.count()
    
    # 9. --- ESTADÍSTICAS DE USUARIOS ACTIVOS ---
    # Usuarios activos por rol (últimos 30 días)
    usuarios_activos_por_rol = {}
    for rol_nombre in [ROL_COORDINADORA, ROL_COORDINADOR_TECNICO_PEDAGOGICO, ROL_ASESOR, ROL_DIRECTOR]:
        usuarios_rol = PerfilUsuario.objects.filter(rol__nombre_rol=rol_nombre).count()
        usuarios_activos_por_rol[rol_nombre] = usuarios_rol
    
    # 10. --- ESTADÍSTICAS DE AJUSTES POR ESTADO ---
    ajustes_stats = {
        'total': total_ajustes,
        'aprobados': ajustes_aprobados,
        'rechazados': ajustes_rechazados,
        'pendientes': ajustes_pendientes,
        'tasa_aprobacion': round((ajustes_aprobados / total_ajustes * 100) if total_ajustes > 0 else 0, 1)
    }
    
    # 11. --- Filtrar Estadísticas por Rango de Tiempo (si aplica) ---
    if rango_seleccionado != 'historico' and fecha_inicio_dt:
        # Filtrar estados por rango
        estados_stats_filtrados = []
        total_filtrado = Solicitudes.objects.filter(created_at__gte=fecha_inicio_dt).count()
        for estado_valor, estado_nombre in Solicitudes.ESTADO_CHOICES:
            cantidad = Solicitudes.objects.filter(
                estado=estado_valor,
                created_at__gte=fecha_inicio_dt
            ).count()
            porcentaje = round((cantidad / total_filtrado * 100) if total_filtrado > 0 else 0, 1)
            estados_stats_filtrados.append({
                'valor': estado_valor,
                'nombre': estado_nombre,
                'cantidad': cantidad,
                'porcentaje': porcentaje
            })
        estados_stats = estados_stats_filtrados
        
        # Filtrar roles por rango
        casos_encargado_inclusion_filtrado = Solicitudes.objects.filter(
            coordinadora_asignada__isnull=False,
            created_at__gte=fecha_inicio_dt
        ).count()
        
        casos_coordinador_tecnico_filtrado = Solicitudes.objects.filter(
            coordinador_tecnico_pedagogico_asignado__isnull=False,
            created_at__gte=fecha_inicio_dt
        ).count()
        
        casos_asesor_pedagogico_filtrado = Solicitudes.objects.filter(
            asesor_pedagogico_asignado__isnull=False,
            created_at__gte=fecha_inicio_dt
        ).count()
        
        casos_sin_asignar_filtrado = Solicitudes.objects.filter(
            coordinadora_asignada__isnull=True,
            coordinador_tecnico_pedagogico_asignado__isnull=True,
            asesor_pedagogico_asignado__isnull=True,
            created_at__gte=fecha_inicio_dt
        ).count()
        
        roles_stats = {
            'encargado_inclusion': casos_encargado_inclusion_filtrado,
            'coordinador_tecnico': casos_coordinador_tecnico_filtrado,
            'asesor_pedagogico': casos_asesor_pedagogico_filtrado,
            'sin_asignar': casos_sin_asignar_filtrado,
        }
    
    # 12. --- Filtrar KPIs por Rango de Tiempo ---
    if rango_seleccionado != 'historico' and fecha_inicio_dt:
        total_casos_filtrado = Solicitudes.objects.filter(created_at__gte=fecha_inicio_dt).count()
        total_ajustes_filtrado = AjusteAsignado.objects.filter(
            created_at__gte=fecha_inicio_dt
        ).count() if fecha_inicio_dt else total_ajustes
        ajustes_aprobados_filtrado = AjusteAsignado.objects.filter(
            estado_aprobacion='aprobado',
            created_at__gte=fecha_inicio_dt
        ).count() if fecha_inicio_dt else ajustes_aprobados
        ajustes_rechazados_filtrado = AjusteAsignado.objects.filter(
            estado_aprobacion='rechazado',
            created_at__gte=fecha_inicio_dt
        ).count() if fecha_inicio_dt else ajustes_rechazados
        ajustes_pendientes_filtrado = AjusteAsignado.objects.filter(
            estado_aprobacion='pendiente',
            created_at__gte=fecha_inicio_dt
        ).count() if fecha_inicio_dt else ajustes_pendientes
        
        ajustes_stats = {
            'total': total_ajustes_filtrado,
            'aprobados': ajustes_aprobados_filtrado,
            'rechazados': ajustes_rechazados_filtrado,
            'pendientes': ajustes_pendientes_filtrado,
            'tasa_aprobacion': round((ajustes_aprobados_filtrado / total_ajustes_filtrado * 100) if total_ajustes_filtrado > 0 else 0, 1)
        }
        
        # Tiempo promedio y casos resueltos filtrados
        casos_aprobados_filtrados = Solicitudes.objects.filter(
            estado='aprobado',
            created_at__gte=fecha_inicio_dt
        ) if fecha_inicio_dt else Solicitudes.objects.filter(estado='aprobado')
        
        tiempos_resolucion_filtrados = []
        for caso in casos_aprobados_filtrados:
            if caso.created_at and caso.updated_at:
                tiempo = (caso.updated_at - caso.created_at).days
                tiempos_resolucion_filtrados.append(tiempo)
        
        tiempo_promedio = round(sum(tiempos_resolucion_filtrados) / len(tiempos_resolucion_filtrados), 1) if tiempos_resolucion_filtrados else 0
        
        # Casos resueltos en el rango
        casos_resueltos_rango = Solicitudes.objects.filter(
            estado__in=['aprobado', 'rechazado'],
            updated_at__gte=fecha_inicio_dt
        ).count() if fecha_inicio_dt else casos_resueltos_semana
        
        total_resueltos_filtrado = Solicitudes.objects.filter(
            estado__in=['aprobado', 'rechazado'],
            created_at__gte=fecha_inicio_dt
        ).count() if fecha_inicio_dt else total_resueltos
        
        total_aprobados_filtrado = Solicitudes.objects.filter(
            estado='aprobado',
            created_at__gte=fecha_inicio_dt
        ).count() if fecha_inicio_dt else total_aprobados
        
        tasa_aprobacion_general = round((total_aprobados_filtrado / total_resueltos_filtrado * 100) if total_resueltos_filtrado > 0 else 0, 1)
        
        total_casos = total_casos_filtrado
        total_ajustes = total_ajustes_filtrado
        casos_resueltos_semana = casos_resueltos_rango
    
    # 13. --- Preparar Contexto ---
    context = {
        'rango_seleccionado': rango_seleccionado,
        'rango_nombre': rango_nombre,
        'estados_stats': estados_stats,
        'roles_stats': roles_stats,
        'ajustes_stats': ajustes_stats,
        'ajustes_por_categoria': ajustes_por_categoria,
        'casos_por_mes': casos_por_mes,
        'casos_por_dia': casos_por_dia,
        'carreras_stats': carreras_stats[:10],  # Top 10
        'tiempo_promedio': tiempo_promedio,
        'casos_resueltos_semana': casos_resueltos_semana,
        'tasa_aprobacion_general': tasa_aprobacion_general,
        'usuarios_activos_por_rol': usuarios_activos_por_rol,
        'total_casos': total_casos,
        'total_ajustes': total_ajustes,
    }
    
    return render(request, 'SIAPE/estadisticas_asesor_pedagogico.html', context)


def obtener_datos_estadisticas_por_rango(rango_seleccionado):
    """
    Función auxiliar para obtener datos de estadísticas según el rango de tiempo.
    Reutiliza la lógica de estadisticas_asesor_pedagogico.
    """
    now = timezone.localtime(timezone.now())
    today = now.date()
    
    fecha_inicio_dt = None
    rango_nombre = ''
    
    if rango_seleccionado == 'mes':
        fecha_inicio = today - timedelta(days=30)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Mes'
    elif rango_seleccionado == 'semestre':
        fecha_inicio = today - timedelta(days=180)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Semestre'
    elif rango_seleccionado == 'año':
        fecha_inicio = today.replace(month=1, day=1)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Año'
    else:  # historico
        fecha_inicio_dt = None
        rango_nombre = 'Histórico Completo'
    
    # Obtener datos filtrados
    if rango_seleccionado != 'historico' and fecha_inicio_dt:
        queryset_casos = Solicitudes.objects.filter(created_at__gte=fecha_inicio_dt)
        queryset_ajustes = AjusteAsignado.objects.filter(created_at__gte=fecha_inicio_dt)
    else:
        queryset_casos = Solicitudes.objects.all()
        queryset_ajustes = AjusteAsignado.objects.all()
    
    total_casos = queryset_casos.count()
    total_ajustes = queryset_ajustes.count()
    
    # Casos por estado
    casos_por_estado = {}
    for estado_valor, estado_nombre in Solicitudes.ESTADO_CHOICES:
        cantidad = queryset_casos.filter(estado=estado_valor).count()
        casos_por_estado[estado_nombre] = cantidad
    
    # Roles stats
    roles_stats = {
        'encargado_inclusion': queryset_casos.filter(coordinadora_asignada__isnull=False).count(),
        'coordinador_tecnico': queryset_casos.filter(coordinador_tecnico_pedagogico_asignado__isnull=False).count(),
        'asesor_pedagogico': queryset_casos.filter(asesor_pedagogico_asignado__isnull=False).count(),
        'sin_asignar': queryset_casos.filter(
            coordinadora_asignada__isnull=True,
            coordinador_tecnico_pedagogico_asignado__isnull=True,
            asesor_pedagogico_asignado__isnull=True
        ).count(),
    }
    
    # Ajustes stats
    ajustes_stats = {
        'total': total_ajustes,
        'aprobados': queryset_ajustes.filter(estado_aprobacion='aprobado').count(),
        'rechazados': queryset_ajustes.filter(estado_aprobacion='rechazado').count(),
        'pendientes': queryset_ajustes.filter(estado_aprobacion='pendiente').count(),
    }
    
    # Tiempo promedio de resolución
    casos_aprobados = queryset_casos.filter(estado='aprobado')
    tiempos_resolucion = []
    for caso in casos_aprobados:
        if caso.created_at and caso.updated_at:
            tiempo = (caso.updated_at - caso.created_at).days
            tiempos_resolucion.append(tiempo)
    tiempo_promedio = round(sum(tiempos_resolucion) / len(tiempos_resolucion), 1) if tiempos_resolucion else 0
    
    # Tasa de aprobación
    total_resueltos = queryset_casos.filter(estado__in=['aprobado', 'rechazado']).count()
    total_aprobados = queryset_casos.filter(estado='aprobado').count()
    tasa_aprobacion = round((total_aprobados / total_resueltos * 100) if total_resueltos > 0 else 0, 1)
    
    # Casos resueltos en el rango
    casos_resueltos = queryset_casos.filter(estado__in=['aprobado', 'rechazado']).count()
    
    # Ajustes por categoría
    ajustes_por_categoria = {}
    for categoria in CategoriasAjustes.objects.all():
        total_cat = queryset_ajustes.filter(ajuste_razonable__categorias_ajustes=categoria).count()
        aprobados_cat = queryset_ajustes.filter(
            ajuste_razonable__categorias_ajustes=categoria,
            estado_aprobacion='aprobado'
        ).count()
        ajustes_por_categoria[categoria.nombre_categoria] = {
            'total': total_cat,
            'aprobados': aprobados_cat,
            'tasa_aprobacion': round((aprobados_cat / total_cat * 100) if total_cat > 0 else 0, 1)
        }
    
    # Estadísticas por carrera (top 10)
    carreras_stats = []
    for carrera in Carreras.objects.all():
        total_casos_carrera = queryset_casos.filter(estudiantes__carreras=carrera).count()
        casos_aprobados_carrera = queryset_casos.filter(
            estudiantes__carreras=carrera,
            estado='aprobado'
        ).count()
        if total_casos_carrera > 0:
            carreras_stats.append({
                'nombre': carrera.nombre,
                'total': total_casos_carrera,
                'aprobados': casos_aprobados_carrera,
                'tasa_aprobacion': round((casos_aprobados_carrera / total_casos_carrera * 100), 1)
            })
    carreras_stats.sort(key=lambda x: x['total'], reverse=True)
    carreras_stats = carreras_stats[:10]
    
    # Usuarios activos por rol
    usuarios_activos_por_rol = {}
    for rol_nombre in [ROL_COORDINADORA, ROL_COORDINADOR_TECNICO_PEDAGOGICO, ROL_ASESOR, ROL_DIRECTOR]:
        usuarios_rol = PerfilUsuario.objects.filter(rol__nombre_rol=rol_nombre).count()
        usuarios_activos_por_rol[rol_nombre] = usuarios_rol
    
    # Casos por mes (últimos 12 meses o según rango)
    casos_por_mes = []
    if rango_seleccionado == 'mes':
        # Último mes: por semana
        for i in range(3, -1, -1):
            semana_inicio = today - timedelta(days=(i * 7) + today.weekday())
            semana_fin = semana_inicio + timedelta(days=6)
            if semana_fin > today:
                semana_fin = today
            semana_inicio_dt = timezone.make_aware(datetime.combine(semana_inicio, datetime.min.time()))
            semana_fin_dt = timezone.make_aware(datetime.combine(semana_fin, datetime.max.time()))
            cantidad = queryset_casos.filter(created_at__range=(semana_inicio_dt, semana_fin_dt)).count()
            casos_por_mes.append({
                'periodo': f'Sem {i+1}',
                'cantidad': cantidad
            })
    elif rango_seleccionado == 'semestre':
        # Último semestre: por mes
        for i in range(5, -1, -1):
            fecha_mes = today - timedelta(days=30 * i)
            mes_inicio = fecha_mes.replace(day=1)
            if i == 0:
                mes_fin = today
            else:
                siguiente_mes = mes_inicio + timedelta(days=32)
                mes_fin = siguiente_mes.replace(day=1) - timedelta(days=1)
            mes_inicio_dt = timezone.make_aware(datetime.combine(mes_inicio, datetime.min.time()))
            mes_fin_dt = timezone.make_aware(datetime.combine(mes_fin, datetime.max.time()))
            cantidad = queryset_casos.filter(created_at__range=(mes_inicio_dt, mes_fin_dt)).count()
            casos_por_mes.append({
                'periodo': mes_inicio.strftime('%b %Y'),
                'cantidad': cantidad
            })
    else:
        # Año o histórico: por mes (últimos 12 meses)
        for i in range(11, -1, -1):
            fecha_mes = today - timedelta(days=30 * i)
            mes_inicio = fecha_mes.replace(day=1)
            if i == 0:
                mes_fin = today
            else:
                siguiente_mes = mes_inicio + timedelta(days=32)
                mes_fin = siguiente_mes.replace(day=1) - timedelta(days=1)
            mes_inicio_dt = timezone.make_aware(datetime.combine(mes_inicio, datetime.min.time()))
            mes_fin_dt = timezone.make_aware(datetime.combine(mes_fin, datetime.max.time()))
            cantidad = queryset_casos.filter(created_at__range=(mes_inicio_dt, mes_fin_dt)).count()
            casos_por_mes.append({
                'periodo': mes_inicio.strftime('%b %Y'),
                'cantidad': cantidad
            })
    
    return {
        'rango_nombre': rango_nombre,
        'total_casos': total_casos,
        'total_ajustes': total_ajustes,
        'casos_por_estado': casos_por_estado,
        'roles_stats': roles_stats,
        'ajustes_stats': ajustes_stats,
        'tiempo_promedio': tiempo_promedio,
        'tasa_aprobacion': tasa_aprobacion,
        'casos_resueltos': casos_resueltos,
        'ajustes_por_categoria': ajustes_por_categoria,
        'carreras_stats': carreras_stats,
        'usuarios_activos_por_rol': usuarios_activos_por_rol,
        'casos_por_mes': casos_por_mes,
        'fecha_inicio_dt': fecha_inicio_dt
    }


@login_required
def generar_reporte_pdf_asesor(request):
    """
    Genera un reporte PDF con las estadísticas del Asesor Pedagógico según el rango de tiempo seleccionado.
    """
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_ASESOR:
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
    
    rango_seleccionado = request.GET.get('rango', 'mes')
    datos = obtener_datos_estadisticas_por_rango(rango_seleccionado)
    
    # Crear el objeto HttpResponse con el tipo de contenido PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_estadisticas_{rango_seleccionado}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    # Crear el objeto PDF usando BytesIO
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Contenedor para los elementos del PDF
    elements = []
    
    # Estilos con colores rojo, blanco y negro
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#dc3545'),  # Rojo
        spaceAfter=20,
        alignment=1,  # Centrado
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#000000'),  # Negro
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    # Colores del esquema
    color_rojo = colors.HexColor('#dc3545')
    color_negro = colors.HexColor('#000000')
    color_blanco = colors.white
    color_gris_claro = colors.HexColor('#f5f5f5')
    
    # Título
    elements.append(Paragraph('Reporte de Estadísticas SIAPE', title_style))
    elements.append(Paragraph(f'Rango de Tiempo: {datos["rango_nombre"]}', heading_style))
    elements.append(Paragraph(f'Fecha de Generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 0.4*inch))
    
    # KPIs Principales
    elements.append(Paragraph('Indicadores Principales', heading_style))
    kpi_data = [
        ['Indicador', 'Valor'],
        ['Total Casos', str(datos['total_casos'])],
        ['Total Ajustes', str(datos['total_ajustes'])],
        ['Ajustes Aprobados', str(datos['ajustes_stats']['aprobados'])],
        ['Ajustes Rechazados', str(datos['ajustes_stats']['rechazados'])],
        ['Ajustes Pendientes', str(datos['ajustes_stats']['pendientes'])],
        ['Casos Resueltos', str(datos['casos_resueltos'])],
        ['Tasa de Aprobación', f"{datos['tasa_aprobacion']}%"],
        ['Tiempo Promedio Resolución', f"{datos['tiempo_promedio']} días"],
    ]
    kpi_table = Table(kpi_data, colWidths=[4*inch, 2*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Casos por Estado
    elements.append(Paragraph('Casos por Estado', heading_style))
    estado_data = [['Estado', 'Cantidad']]
    for estado, cantidad in datos['casos_por_estado'].items():
        estado_data.append([estado, str(cantidad)])
    
    estado_table = Table(estado_data, colWidths=[4*inch, 2*inch])
    estado_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(estado_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Casos por Rol
    elements.append(Paragraph('Casos Asignados por Rol', heading_style))
    rol_data = [
        ['Rol', 'Cantidad'],
        ['Encargado de Inclusión', str(datos['roles_stats']['encargado_inclusion'])],
        ['Coordinador Técnico Pedagógico', str(datos['roles_stats']['coordinador_tecnico'])],
        ['Asesor Pedagógico', str(datos['roles_stats']['asesor_pedagogico'])],
        ['Sin Asignar', str(datos['roles_stats']['sin_asignar'])],
    ]
    rol_table = Table(rol_data, colWidths=[4*inch, 2*inch])
    rol_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(rol_table)
    elements.append(PageBreak())
    
    # Ajustes por Categoría
    if datos['ajustes_por_categoria']:
        elements.append(Paragraph('Ajustes por Categoría', heading_style))
        categoria_data = [['Categoría', 'Total', 'Aprobados', 'Tasa Aprobación (%)']]
        for categoria, stats in datos['ajustes_por_categoria'].items():
            categoria_data.append([
                categoria,
                str(stats['total']),
                str(stats['aprobados']),
                f"{stats['tasa_aprobacion']}%"
            ])
        
        categoria_table = Table(categoria_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.5*inch])
        categoria_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
            ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
            ('GRID', (0, 0), (-1, -1), 1, color_negro),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
        ]))
        elements.append(categoria_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Estadísticas por Carrera (Top 10)
    if datos['carreras_stats']:
        elements.append(Paragraph('Estadísticas por Carrera (Top 10)', heading_style))
        carrera_data = [['Carrera', 'Total Casos', 'Aprobados', 'Tasa Aprobación (%)']]
        for carrera in datos['carreras_stats']:
            carrera_data.append([
                carrera['nombre'][:40],  # Limitar longitud
                str(carrera['total']),
                str(carrera['aprobados']),
                f"{carrera['tasa_aprobacion']}%"
            ])
        
        carrera_table = Table(carrera_data, colWidths=[3*inch, 1.2*inch, 1.2*inch, 1.5*inch])
        carrera_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
            ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (0, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
            ('GRID', (0, 0), (-1, -1), 1, color_negro),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
        ]))
        elements.append(carrera_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Usuarios Activos por Rol
    elements.append(Paragraph('Usuarios Activos por Rol', heading_style))
    usuarios_data = [['Rol', 'Cantidad de Usuarios']]
    for rol, cantidad in datos['usuarios_activos_por_rol'].items():
        usuarios_data.append([rol, str(cantidad)])
    
    usuarios_table = Table(usuarios_data, colWidths=[4*inch, 2*inch])
    usuarios_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(usuarios_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Casos por Período
    if datos['casos_por_mes']:
        elements.append(Paragraph('Casos por Período', heading_style))
        periodo_data = [['Período', 'Cantidad de Casos']]
        for periodo in datos['casos_por_mes']:
            periodo_data.append([periodo['periodo'], str(periodo['cantidad'])])
        
        periodo_table = Table(periodo_data, colWidths=[3*inch, 3*inch])
        periodo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
            ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
            ('GRID', (0, 0), (-1, -1), 1, color_negro),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
        ]))
        elements.append(periodo_table)
    
    # Construir el PDF
    doc.build(elements)
    
    # Obtener el valor del BytesIO buffer y escribirlo en la respuesta
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


@login_required
def generar_reporte_excel_asesor(request):
    """
    Genera un archivo Excel con los datos según el rango de tiempo seleccionado.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_ASESOR:
            # Si no tiene permisos, devolver un error HTTP en lugar de redirect
            return HttpResponse('No tienes permisos para acceder a esta página.', status=403)
    except AttributeError:
        if not request.user.is_superuser:
            return HttpResponse('No tienes permisos para acceder a esta página.', status=403)
    
    try:
        rango_seleccionado = request.GET.get('rango', 'mes')
        datos = obtener_datos_estadisticas_por_rango(rango_seleccionado)
    except Exception as e:
        # Si hay un error al obtener los datos, devolver un error HTTP
        return HttpResponse(f'Error al generar el reporte: {str(e)}', status=500)
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Estadísticas"
    
    # Estilos
    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    row = 1
    
    # Título
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = f"Reporte de Estadísticas - {datos['rango_nombre']}"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # KPIs
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = "Indicadores Principales (KPIs)"
    ws[f'A{row}'].font = title_font
    row += 1
    
    # Encabezados de KPIs
    headers = ['KPI', 'Valor', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Datos de KPIs
    kpis_data = [
        ['Total Casos', datos['total_casos']],
        ['Total Ajustes', datos['total_ajustes']],
        ['Ajustes Aprobados', datos['ajustes_stats']['aprobados']],
        ['Ajustes Rechazados', datos['ajustes_stats']['rechazados']],
        ['Ajustes Pendientes', datos['ajustes_stats']['pendientes']],
    ]
    
    for kpi, valor in kpis_data:
        ws.cell(row=row, column=1).value = kpi
        ws.cell(row=row, column=2).value = valor
        ws.cell(row=row, column=3).value = datos['rango_nombre']
        row += 1
    
    row += 1
    
    # Casos por Estado
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "Casos por Estado"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['Estado', 'Cantidad', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    for estado, cantidad in datos['casos_por_estado'].items():
        ws.cell(row=row, column=1).value = estado
        ws.cell(row=row, column=2).value = cantidad
        ws.cell(row=row, column=3).value = datos['rango_nombre']
        row += 1
    
    row += 1
    
    # Casos por Rol
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "Casos por Rol"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['Rol', 'Cantidad', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    roles_data = [
        ['Encargado de Inclusión', datos['roles_stats']['encargado_inclusion']],
        ['Coordinador Técnico Pedagógico', datos['roles_stats']['coordinador_tecnico']],
        ['Asesor Pedagógico', datos['roles_stats']['asesor_pedagogico']],
    ]
    
    for rol, cantidad in roles_data:
        ws.cell(row=row, column=1).value = rol
        ws.cell(row=row, column=2).value = cantidad
        ws.cell(row=row, column=3).value = datos['rango_nombre']
        row += 1
    
    row += 1
    
    # Detalle de Casos
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].value = "Detalle de Casos"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['ID', 'Estudiante', 'Carrera', 'Estado', 'Fecha Creación', 'Asunto']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Agregar datos detallados de casos
    if datos['fecha_inicio_dt']:
        casos = Solicitudes.objects.filter(created_at__gte=datos['fecha_inicio_dt']).select_related('estudiantes', 'estudiantes__carreras')[:1000]
    else:
        casos = Solicitudes.objects.all().select_related('estudiantes', 'estudiantes__carreras')[:1000]
    
    for caso in casos:
        estudiante_nombre = f"{caso.estudiantes.nombres} {caso.estudiantes.apellidos}" if caso.estudiantes else "N/A"
        carrera_nombre = caso.estudiantes.carreras.nombre if caso.estudiantes and caso.estudiantes.carreras else "N/A"
        fecha_creacion = timezone.localtime(caso.created_at).strftime('%Y-%m-%d %H:%M:%S') if caso.created_at else "N/A"
        
        ws.cell(row=row, column=1).value = caso.id
        ws.cell(row=row, column=2).value = estudiante_nombre
        ws.cell(row=row, column=3).value = carrera_nombre
        ws.cell(row=row, column=4).value = caso.get_estado_display()
        ws.cell(row=row, column=5).value = fecha_creacion
        ws.cell(row=row, column=6).value = caso.asunto[:50] if caso.asunto else "N/A"
        row += 1
    
    # Ajustar ancho de columnas
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                # Saltar celdas fusionadas que no tienen valor directamente
                if hasattr(cell, 'value') and cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP usando BytesIO para evitar problemas
    try:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_excel_{rango_seleccionado}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        output.close()
        return response
    except Exception as e:
        return HttpResponse(f'Error al generar el archivo Excel: {str(e)}', status=500)


# ----------------------------------------------------
#                VISTA DIRECTOR DE CARRERA
# ----------------------------------------------------

# @login_required
# def dashboard_director(request):
#     """
#     Dashboard para el Director de Carrera.
#     Muestra casos pendientes de su aprobación y un historial
#     de casos aprobados de sus carreras.
#     """
#     try:
#         perfil_director = request.user.perfil
#         if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
#             messages.error(request, 'No tienes permisos para esta acción.')
#             return redirect('home')
#     except AttributeError:
#         return redirect('home')

#     # 1. Encontrar las carreras que este director gestiona
#     carreras_del_director = Carreras.objects.filter(director=perfil_director)
    
#     # 2. Base de solicitudes de sus carreras
#     solicitudes_base = Solicitudes.objects.filter(
#         estudiantes__carreras__in=carreras_del_director
#     ).select_related(
#         'estudiantes', 
#         'estudiantes__carreras'
#     )

#     # 3. Filtrar solicitudes PENDIENTES (estado 'pendiente_aprobacion')
#     solicitudes_pendientes = solicitudes_base.filter(
#         estado='pendiente_aprobacion'
#     ).order_by('updated_at') # Más antiguas (recién llegadas) primero

#     # 4. Filtrar el HISTORIAL (solo 'aprobado')
#     solicitudes_historial = solicitudes_base.filter(
#         estado='aprobado'
#     ).order_by('-updated_at') # Más recientes primero

#     # 5. KPIs (Específicos del Director)
#     kpis = {
#         'total_pendientes': solicitudes_pendientes.count(),
#         'total_aprobados': solicitudes_historial.count(),
#     }

#     context = {
#         'nombre_usuario': request.user.first_name,
#         'solicitudes_pendientes': solicitudes_pendientes,
#         'solicitudes_historial': solicitudes_historial,
#         'kpis': kpis,
#     }
    
#     return render(request, 'SIAPE/dashboard_director.html', context)
@login_required
def dashboard_director(request):
    """
    Dashboard para el Director de Carrera.
    Muestra casos pendientes de su aprobación y un historial
    de casos aprobados de sus carreras.
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')

    # 1. Encontrar las carreras que este director gestiona
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    
    # Si no tiene carreras asignadas, mostrar mensaje y retornar lista vacía
    if not carreras_del_director.exists():
        # Determinar semestre actual para mostrar en los KPIs
        mes_actual = timezone.localtime(timezone.now()).month
        if mes_actual >= 3 and mes_actual <= 8:
            semestre_actual = 1
        else:
            semestre_actual = 2
        
        messages.warning(request, 'No tienes carreras asignadas. Contacta a un administrador para que te asigne carreras.')
        context = {
            'nombre_usuario': request.user.first_name,
            'solicitudes_pendientes': Solicitudes.objects.none(),
            'solicitudes_historial': Solicitudes.objects.none(),
            'kpis': {
                'total_pendientes': 0,
                'total_aprobados': 0,
                'total_rechazados': 0,
                'semestre_actual': semestre_actual,
            },
        }
        return render(request, 'SIAPE/dashboard_director.html', context)
    
    # Obtener IDs de las carreras para hacer el filtro más eficiente
    carreras_ids = carreras_del_director.values_list('id', flat=True)
    
    # 1.5. Determinar el semestre actual basado en la fecha
    # Semestre 1: Marzo - Agosto (meses 3-8)
    # Semestre 2: Septiembre - Febrero (meses 9-12, 1-2)
    mes_actual = timezone.localtime(timezone.now()).month
    if mes_actual >= 3 and mes_actual <= 8:
        semestre_actual = 1
    else:
        semestre_actual = 2
    
    # 2. Base de solicitudes de sus carreras - usando IDs para mejor rendimiento
    solicitudes_base = Solicitudes.objects.filter(
        estudiantes__carreras__id__in=carreras_ids
    ).select_related(
        'estudiantes', 
        'estudiantes__carreras'
    ).distinct()
    
    # 2.5. Filtrar solicitudes por semestre actual del estudiante
    solicitudes_base_semestre = solicitudes_base.filter(
        estudiantes__semestre_actual=semestre_actual
    )

    # 3. Filtrar solicitudes PENDIENTES (estado 'pendiente_aprobacion')
    # Estos son los casos que el Asesor Pedagógico le envió.
    solicitudes_pendientes = solicitudes_base.filter(
        estado='pendiente_aprobacion'
    ).order_by('updated_at') # Más antiguas (recién llegadas) primero

    # 4. Filtrar el HISTORIAL (casos 'aprobados' o 'rechazados')
    solicitudes_historial_base = solicitudes_base.filter(
        estado__in=['aprobado', 'rechazado']
    ).order_by('-updated_at') # Más recientes primero

    # 5. KPIs (Específicos del Director) - filtrar por semestre actual
    kpis = {
        'total_pendientes': solicitudes_base_semestre.filter(estado='pendiente_aprobacion').count(),
        'total_aprobados': solicitudes_base_semestre.filter(estado='aprobado').count(),
        'total_rechazados': solicitudes_base_semestre.filter(estado='rechazado').count(),
        'semestre_actual': semestre_actual,
    }

    # 6. Paginación del historial (10 por página)
    page_historial = request.GET.get('page_historial', 1)
    paginator_historial = Paginator(solicitudes_historial_base, 10)
    try:
        solicitudes_historial = paginator_historial.page(page_historial)
    except PageNotAnInteger:
        solicitudes_historial = paginator_historial.page(1)
    except EmptyPage:
        solicitudes_historial = paginator_historial.page(paginator_historial.num_pages)

    context = {
        'nombre_usuario': request.user.first_name,
        'solicitudes_pendientes': solicitudes_pendientes,
        'solicitudes_historial': solicitudes_historial,
        'kpis': kpis,
    }
    
    return render(request, 'SIAPE/dashboard_director.html', context)

@login_required
def carreras_director(request):
    """
    Muestra las carreras asignadas al Director de Carrera logueado.
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')

    # Buscamos las carreras donde el director sea el usuario actual
    carreras_list = Carreras.objects.filter(
        director=perfil_director
    ).select_related(
        'area'
    ).annotate(
        total_estudiantes=Count('estudiantes') # Contamos los estudiantes via FK
    ).order_by('nombre')

    context = {
        'carreras_list': carreras_list,
        'total_carreras': carreras_list.count()
    }
    return render(request, 'SIAPE/carreras_director.html', context)

@login_required
def estudiantes_por_carrera_director(request, carrera_id):
    """
    Muestra la lista de estudiantes de una carrera específica
    gestionada por el Director de Carrera.
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')

    # 1. Obtener la carrera y verificar que el director sea el correcto
    # Esta es la comprobación de seguridad:
    carrera = get_object_or_404(Carreras, id=carrera_id, director=perfil_director)

    # 2. Obtener los estudiantes de esa carrera
    estudiantes_list = Estudiantes.objects.filter(
        carreras=carrera
    ).order_by('apellidos', 'nombres')

    context = {
        'carrera': carrera,
        'estudiantes_list': estudiantes_list,
        'total_estudiantes': estudiantes_list.count(),
        'nombre_usuario': request.user.first_name, # Para la plantilla
    }
    # 3. Renderizar un nuevo template que crearemos a continuación
    return render(request, 'SIAPE/estudiantes_carrera_director.html', context)

@login_required
def perfil_estudiante_director(request, estudiante_id):
    """
    Muestra el perfil completo de un estudiante con sus solicitudes
    para el Director de Carrera.
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')

    # 1. Obtener el estudiante
    estudiante = get_object_or_404(Estudiantes, id=estudiante_id)
    
    # 2. Verificar que el director tenga acceso a la carrera del estudiante
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    if estudiante.carreras not in carreras_del_director:
        messages.error(request, 'No tienes permisos para ver este estudiante.')
        return redirect('carreras_director')
    
    # 3. Obtener todas las solicitudes del estudiante ordenadas por fecha (más recientes primero)
    solicitudes = Solicitudes.objects.filter(
        estudiantes=estudiante
    ).select_related(
        'estudiantes',
        'estudiantes__carreras',
        'coordinadora_asignada',
        'coordinador_tecnico_pedagogico_asignado',
        'asesor_pedagogico_asignado'
    ).prefetch_related(
        'ajusteasignado_set',
        'ajusteasignado_set__ajuste_razonable',
        'ajusteasignado_set__ajuste_razonable__categorias_ajustes'
    ).order_by('-created_at')
    
    # 4. Estadísticas del estudiante
    total_solicitudes = solicitudes.count()
    solicitudes_aprobadas = solicitudes.filter(estado='aprobado').count()
    solicitudes_rechazadas = solicitudes.filter(estado='rechazado').count()
    solicitudes_pendientes = solicitudes.exclude(estado__in=['aprobado', 'rechazado']).count()
    
    # 5. Obtener asignaturas en curso del estudiante
    asignaturas_en_curso = estudiante.asignaturasencurso_set.filter(estado=True).select_related('asignaturas')
    
    context = {
        'estudiante': estudiante,
        'solicitudes': solicitudes,
        'total_solicitudes': total_solicitudes,
        'solicitudes_aprobadas': solicitudes_aprobadas,
        'solicitudes_rechazadas': solicitudes_rechazadas,
        'solicitudes_pendientes': solicitudes_pendientes,
        'asignaturas_en_curso': asignaturas_en_curso,
        'nombre_usuario': request.user.first_name,
        'carrera': estudiante.carreras,
    }
    
    return render(request, 'SIAPE/perfil_estudiante_director.html', context)

@login_required
def estadisticas_director(request):
    """
    Panel de Estadísticas para el Director de Carrera.
    Muestra gráficos sobre el estado y tipo de ajustes con filtros de tiempo.
    """
    
    # 1. --- Verificación de Permisos ---
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')

    # 2. --- Obtener Rango de Tiempo Seleccionado ---
    rango_seleccionado = request.GET.get('rango', 'mes')  # mes, semestre, año, historico
    
    # 3. --- Configuración de Fechas según Rango ---
    now = timezone.localtime(timezone.now())
    today = now.date()
    
    fecha_inicio_dt = None
    fecha_fin_dt = timezone.make_aware(datetime.combine(today, datetime.max.time()))
    
    if rango_seleccionado == 'mes':
        fecha_inicio = today - timedelta(days=30)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Mes'
    elif rango_seleccionado == 'semestre':
        fecha_inicio = today - timedelta(days=180)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Semestre'
    elif rango_seleccionado == 'año':
        fecha_inicio = today.replace(month=1, day=1)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Año'
    else:  # historico
        fecha_inicio_dt = None
        rango_nombre = 'Histórico Completo'

    # 4. --- Base Query ---
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    
    # Si no tiene carreras asignadas, retornar con datos vacíos
    if not carreras_del_director.exists():
        context = {
            'nombre_usuario': request.user.first_name,
            'rango_seleccionado': rango_seleccionado,
            'rango_nombre': rango_nombre,
            'total_casos': 0,
            'casos_aprobados': 0,
            'casos_pendientes': 0,
            'total_ajustes': 0,
            'ajustes_aprobados': 0,
            'ajustes_rechazados': 0,
            'ajustes_pendientes': 0,
            'tasa_aprobacion': 0,
            'tasa_aprobacion_ajustes': 0,
            'casos_nuevos_semana': 0,
            'pie_chart_data_json': json.dumps({'labels': [], 'datasets': [{'data': []}]}),
            'bar_chart_tipos_json': json.dumps({'labels': [], 'datasets': [{'data': []}]}),
            'bar_chart_secciones_json': json.dumps({'labels': [], 'datasets': [{'data': []}]}),
            'chart_tendencia_json': json.dumps({'labels': [], 'datasets': [{'data': []}]}),
            'carreras_stats': [],
            'estados_stats': [],
        }
        return render(request, 'SIAPE/estadisticas_director.html', context)
    
    # Obtener IDs de las carreras para hacer el filtro más eficiente
    carreras_ids = carreras_del_director.values_list('id', flat=True)
    
    # Base de solicitudes de sus carreras - usando IDs para mejor rendimiento
    solicitudes_base = Solicitudes.objects.filter(
        estudiantes__carreras__id__in=carreras_ids
    ).select_related(
        'estudiantes',
        'estudiantes__carreras'
    ).distinct()
    
    # Base de ajustes de sus carreras
    # No usar distinct() ya que cada AjusteAsignado es único y no debería haber duplicados
    ajustes_base = AjusteAsignado.objects.filter(
        solicitudes__estudiantes__carreras__id__in=carreras_ids
    ).select_related(
        'solicitudes',
        'solicitudes__estudiantes',
        'solicitudes__estudiantes__carreras'
    )
    
    # Aplicar filtro de tiempo (si está definido)
    if fecha_inicio_dt:
        solicitudes_base = solicitudes_base.filter(created_at__gte=fecha_inicio_dt, created_at__lte=fecha_fin_dt)
        ajustes_base = ajustes_base.filter(solicitudes__created_at__gte=fecha_inicio_dt, solicitudes__created_at__lte=fecha_fin_dt)

    # 5. --- KPIs ---
    total_casos = solicitudes_base.count()
    casos_aprobados = solicitudes_base.filter(estado='aprobado').count()
    casos_pendientes = solicitudes_base.exclude(estado__in=['aprobado', 'rechazado']).count()
    total_ajustes = ajustes_base.count()
    ajustes_aprobados = ajustes_base.filter(estado_aprobacion='aprobado').count()
    ajustes_rechazados = ajustes_base.filter(estado_aprobacion='rechazado').count()
    ajustes_pendientes = ajustes_base.filter(estado_aprobacion='pendiente').count()
    
    tasa_aprobacion = round((casos_aprobados / total_casos * 100) if total_casos > 0 else 0, 1)
    tasa_aprobacion_ajustes = round((ajustes_aprobados / total_ajustes * 100) if total_ajustes > 0 else 0, 1)
    
    # Casos nuevos esta semana (sin filtro de tiempo, siempre de los últimos 7 días)
    semana_pasada = timezone.now() - timedelta(days=7)
    casos_nuevos_semana = Solicitudes.objects.filter(
        estudiantes__carreras__id__in=carreras_ids,
        created_at__gte=semana_pasada
    ).distinct().count()

    # 6. --- Gráfico 1: Estado de Ajustes (Doughnut Chart) ---
    estado_ajustes = ajustes_base.values('estado_aprobacion').annotate(total=Count('id')).order_by('estado_aprobacion')
    
    pie_labels = []
    pie_data = []
    pie_colors = []
    
    for d in estado_ajustes:
        estado = d['estado_aprobacion']
        pie_labels.append(estado.capitalize())
        pie_data.append(d['total'])
        if estado == 'aprobado':
            pie_colors.append('rgba(40, 167, 69, 0.7)')  # Verde
        elif estado == 'rechazado':
            pie_colors.append('rgba(220, 53, 69, 0.7)')  # Rojo
        else:
            pie_colors.append('rgba(253, 126, 20, 0.7)')  # Naranja
    
    pie_chart_data = {
        'labels': pie_labels,
        'datasets': [{
            'data': pie_data,
            'backgroundColor': pie_colors,
            'borderColor': '#ffffff',
        }]
    }

    # 7. --- Gráfico 2: Ajustes Aprobados y Rechazados por Categoría ---
    # Obtener todos los ajustes (aprobados y rechazados) con sus categorías
    ajustes_por_categoria_qs = ajustes_base.filter(
        estado_aprobacion__in=['aprobado', 'rechazado']
    ).select_related(
        'ajuste_razonable',
        'ajuste_razonable__categorias_ajustes'
    )
    
    # Convertir a lista para procesar
    ajustes_por_categoria_lista = list(ajustes_por_categoria_qs)
    
    # Procesar: agrupar por categoría y estado
    categorias_estados = {}  # {categoria: {'aprobado': count, 'rechazado': count}}
    
    for ajuste in ajustes_por_categoria_lista:
        if ajuste.ajuste_razonable and ajuste.ajuste_razonable.categorias_ajustes:
            categoria_nombre = ajuste.ajuste_razonable.categorias_ajustes.nombre_categoria
        else:
            categoria_nombre = 'Sin categoría'
        
        estado = ajuste.estado_aprobacion
        
        if categoria_nombre not in categorias_estados:
            categorias_estados[categoria_nombre] = {'aprobado': 0, 'rechazado': 0}
        
        if estado == 'aprobado':
            categorias_estados[categoria_nombre]['aprobado'] += 1
        elif estado == 'rechazado':
            categorias_estados[categoria_nombre]['rechazado'] += 1
    
    # Ordenar por total de ajustes (aprobados + rechazados) descendente
    categorias_ordenadas = sorted(
        categorias_estados.items(),
        key=lambda x: x[1]['aprobado'] + x[1]['rechazado'],
        reverse=True
    )
    
    # Limitar a las primeras 15 categorías si hay muchas
    if len(categorias_ordenadas) > 15:
        categorias_ordenadas = categorias_ordenadas[:15]
    
    # Preparar datos para el gráfico de barras agrupadas
    bar_labels_categorias = [item[0] for item in categorias_ordenadas]
    bar_data_aprobados = [item[1]['aprobado'] for item in categorias_ordenadas]
    bar_data_rechazados = [item[1]['rechazado'] for item in categorias_ordenadas]
    
    bar_chart_tipos = {
        'labels': bar_labels_categorias,
        'datasets': [
            {
                'label': 'Aprobados',
                'data': bar_data_aprobados,
                'backgroundColor': 'rgba(40, 167, 69, 0.7)',  # Verde
            },
            {
                'label': 'Rechazados',
                'data': bar_data_rechazados,
                'backgroundColor': 'rgba(220, 53, 69, 0.7)',  # Rojo
            }
        ]
    }

    # 8. --- Gráfico 3: Secciones con Más Ajustes Aprobados ---
    # Obtener solo ajustes aprobados para las secciones
    # Usar asignaturas_en_curso a través del estudiante en lugar del ManyToMany
    ajustes_aprobados_secciones_qs = ajustes_base.filter(
        estado_aprobacion='aprobado'
    ).select_related(
        'solicitudes',
        'solicitudes__estudiantes',
        'solicitudes__estudiantes__carreras'
    )

    # Convertir a lista para procesar
    ajustes_aprobados_secciones_lista = list(ajustes_aprobados_secciones_qs)

    # Pre-cargar todas las asignaturas_en_curso relacionadas para optimizar
    from SIAPE.models import AsignaturasEnCurso

    estudiantes_ids = [ajuste.solicitudes.estudiantes_id for ajuste in ajustes_aprobados_secciones_lista]
    asignaturas_en_curso_dict = {}

    if estudiantes_ids:
        asignaturas_en_curso_lista = AsignaturasEnCurso.objects.filter(
            estudiantes_id__in=estudiantes_ids,
            estado=True  # Solo asignaturas activas
        ).select_related('asignaturas')
        
        # Organizar por estudiante_id para acceso rápido
        for aec in asignaturas_en_curso_lista:
            estudiante_id = aec.estudiantes_id
            if estudiante_id not in asignaturas_en_curso_dict:
                asignaturas_en_curso_dict[estudiante_id] = []
            asignaturas_en_curso_dict[estudiante_id].append(aec.asignaturas)

    # Contar ajustes por sección (asignatura + sección)
    secciones_counter = Counter()

    for ajuste in ajustes_aprobados_secciones_lista:
        solicitud = ajuste.solicitudes
        estudiante_id = solicitud.estudiantes_id
        
        # Obtener asignaturas a través de asignaturas_en_curso del estudiante
        asignaturas = asignaturas_en_curso_dict.get(estudiante_id, [])
        
        if asignaturas:
            for asignatura in asignaturas:
                # Crear clave única: "Nombre Asignatura (Sección)"
                if asignatura.seccion:
                    clave = f"{asignatura.nombre} ({asignatura.seccion})"
                else:
                    clave = f"{asignatura.nombre} (Sin sección)"
                secciones_counter[clave] += 1
        else:
            # Si no hay asignaturas, contar como "Sin asignatura"
            secciones_counter["Sin asignatura"] += 1

    # Obtener las secciones más frecuentes (top 15)
    top_secciones = secciones_counter.most_common(15)

    # Ordenar por cantidad descendente (most_common ya lo hace, pero asegurémonos)
    top_secciones.sort(key=lambda x: x[1], reverse=True)

    bar_labels_secciones = [item[0] for item in top_secciones]
    bar_data_secciones = [item[1] for item in top_secciones]

    bar_chart_secciones = {
        'labels': bar_labels_secciones,
        'datasets': [{
            'label': 'Ajustes Aprobados',
            'data': bar_data_secciones,
            'backgroundColor': 'rgba(40, 167, 69, 0.7)',  # Verde para aprobados
        }]
    }

    # 9. --- Gráfico 4: Tendencia de Casos por Tiempo ---
    # IMPORTANTE: Usar solicitudes_base que ya tiene el filtro de tiempo aplicado
    casos_por_tiempo = []
    if rango_seleccionado == 'mes':
        # Por día - solo mostrar días dentro del rango seleccionado
        if fecha_inicio_dt:
            fecha_inicio_calc = fecha_inicio_dt.date()
        else:
            # Si es histórico, usar la fecha del primer caso
            primer_caso = solicitudes_base.order_by('created_at').first()
            if primer_caso:
                fecha_inicio_calc = primer_caso.created_at.date()
            else:
                fecha_inicio_calc = today - timedelta(days=30)
        
        for i in range(29, -1, -1):
            fecha_dia = today - timedelta(days=i)
            # Solo incluir días dentro del rango seleccionado
            if fecha_inicio_dt and fecha_dia < fecha_inicio_calc:
                continue
            dia_inicio_dt = timezone.make_aware(datetime.combine(fecha_dia, datetime.min.time()))
            dia_fin_dt = timezone.make_aware(datetime.combine(fecha_dia, datetime.max.time()))
            # Usar solicitudes_base filtrado y aplicar filtro adicional por día
            cantidad = solicitudes_base.filter(
                created_at__range=(dia_inicio_dt, dia_fin_dt)
            ).distinct().count()
            casos_por_tiempo.append({'fecha': fecha_dia.strftime('%d/%m'), 'cantidad': cantidad})
    elif rango_seleccionado == 'semestre':
        # Por mes - usar solicitudes_base que ya tiene el filtro de tiempo aplicado
        if fecha_inicio_dt:
            fecha_inicio_calc = fecha_inicio_dt.date()
        else:
            primer_caso = solicitudes_base.order_by('created_at').first()
            if primer_caso:
                fecha_inicio_calc = primer_caso.created_at.date()
            else:
                fecha_inicio_calc = today - timedelta(days=180)
        
        for i in range(5, -1, -1):
            fecha_mes = today - timedelta(days=30 * i)
            mes_inicio = fecha_mes.replace(day=1)
            # Solo incluir meses dentro del rango seleccionado
            if fecha_inicio_dt and mes_inicio < fecha_inicio_calc:
                continue
            if i == 0:
                mes_fin = today
            else:
                siguiente_mes = mes_inicio + timedelta(days=32)
                mes_fin = siguiente_mes.replace(day=1) - timedelta(days=1)
            mes_inicio_dt = timezone.make_aware(datetime.combine(mes_inicio, datetime.min.time()))
            mes_fin_dt = timezone.make_aware(datetime.combine(mes_fin, datetime.max.time()))
            # Usar solicitudes_base filtrado y aplicar filtro adicional por mes
            cantidad = solicitudes_base.filter(
                created_at__range=(mes_inicio_dt, mes_fin_dt)
            ).distinct().count()
            casos_por_tiempo.append({'fecha': mes_inicio.strftime('%b %Y'), 'cantidad': cantidad})
    elif rango_seleccionado == 'año':
        # Por mes - usar solicitudes_base que ya tiene el filtro de tiempo aplicado
        if fecha_inicio_dt:
            fecha_inicio_calc = fecha_inicio_dt.date()
        else:
            primer_caso = solicitudes_base.order_by('created_at').first()
            if primer_caso:
                fecha_inicio_calc = primer_caso.created_at.date()
            else:
                fecha_inicio_calc = today.replace(month=1, day=1)
        
        for i in range(11, -1, -1):
            fecha_mes = today - timedelta(days=30 * i)
            mes_inicio = fecha_mes.replace(day=1)
            # Solo incluir meses dentro del rango seleccionado
            if fecha_inicio_dt and mes_inicio < fecha_inicio_calc:
                continue
            if i == 0:
                mes_fin = today
            else:
                siguiente_mes = mes_inicio + timedelta(days=32)
                mes_fin = siguiente_mes.replace(day=1) - timedelta(days=1)
            mes_inicio_dt = timezone.make_aware(datetime.combine(mes_inicio, datetime.min.time()))
            mes_fin_dt = timezone.make_aware(datetime.combine(mes_fin, datetime.max.time()))
            # Usar solicitudes_base filtrado y aplicar filtro adicional por mes
            cantidad = solicitudes_base.filter(
                created_at__range=(mes_inicio_dt, mes_fin_dt)
            ).distinct().count()
            casos_por_tiempo.append({'fecha': mes_inicio.strftime('%b %Y'), 'cantidad': cantidad})
    else:  # historico
        # Por año
        primer_caso = solicitudes_base.order_by('created_at').first()
        if primer_caso:
            año_inicio = primer_caso.created_at.year
            año_actual = today.year
            for año in range(año_inicio, año_actual + 1):
                año_inicio_dt = timezone.make_aware(datetime(año, 1, 1))
                if año == año_actual:
                    año_fin_dt = timezone.make_aware(datetime.combine(today, datetime.max.time()))
                else:
                    año_fin_dt = timezone.make_aware(datetime(año + 1, 1, 1)) - timedelta(seconds=1)
                # Usar solicitudes_base (histórico completo, sin filtro de tiempo adicional)
                cantidad = solicitudes_base.filter(
                    created_at__range=(año_inicio_dt, año_fin_dt)
                ).distinct().count()
                casos_por_tiempo.append({'fecha': str(año), 'cantidad': cantidad})

    chart_tendencia_data = {
        'labels': [d['fecha'] for d in casos_por_tiempo],
        'datasets': [{
            'label': 'Casos',
            'data': [d['cantidad'] for d in casos_por_tiempo],
            'borderColor': 'rgba(211, 47, 47, 1)',
            'backgroundColor': 'rgba(211, 47, 47, 0.1)',
            'tension': 0.4
        }]
    }

    # 10. --- Estadísticas por Carrera ---
    carreras_stats = []
    for carrera in carreras_del_director:
        casos_carrera = Solicitudes.objects.filter(
            estudiantes__carreras=carrera
        )
        # Aplicar filtro de tiempo si existe
        if fecha_inicio_dt:
            casos_carrera = casos_carrera.filter(created_at__gte=fecha_inicio_dt, created_at__lte=fecha_fin_dt)
        total_carrera = casos_carrera.distinct().count()
        aprobados_carrera = casos_carrera.filter(estado='aprobado').distinct().count()
        if total_carrera > 0:
            carreras_stats.append({
                'nombre': carrera.nombre,
                'total': total_carrera,
                'aprobados': aprobados_carrera,
                'tasa_aprobacion': round((aprobados_carrera / total_carrera * 100), 1)
            })
    
    carreras_stats.sort(key=lambda x: x['total'], reverse=True)

    # 11. --- Estadísticas por Estado ---
    estados_stats = []
    total_casos_estados = solicitudes_base.count()
    for estado_valor, estado_nombre in Solicitudes.ESTADO_CHOICES:
        cantidad = solicitudes_base.filter(estado=estado_valor).count()
        porcentaje = round((cantidad / total_casos_estados * 100) if total_casos_estados > 0 else 0, 1)
        estados_stats.append({
            'valor': estado_valor,
            'nombre': estado_nombre,
            'cantidad': cantidad,
            'porcentaje': porcentaje
        })

    context = {
        'nombre_usuario': request.user.first_name,
        'rango_seleccionado': rango_seleccionado,
        'rango_nombre': rango_nombre,
        # KPIs
        'total_casos': total_casos,
        'casos_aprobados': casos_aprobados,
        'casos_pendientes': casos_pendientes,
        'total_ajustes': total_ajustes,
        'ajustes_aprobados': ajustes_aprobados,
        'ajustes_rechazados': ajustes_rechazados,
        'ajustes_pendientes': ajustes_pendientes,
        'tasa_aprobacion': tasa_aprobacion,
        'tasa_aprobacion_ajustes': tasa_aprobacion_ajustes,
        'casos_nuevos_semana': casos_nuevos_semana,
        # Gráficos
        'pie_chart_data_json': json.dumps(pie_chart_data),
        'bar_chart_tipos_json': json.dumps(bar_chart_tipos),
        'bar_chart_secciones_json': json.dumps(bar_chart_secciones),
        'chart_tendencia_json': json.dumps(chart_tendencia_data),
        # Estadísticas
        'carreras_stats': carreras_stats,
        'estados_stats': estados_stats,
    }
    
    return render(request, 'SIAPE/estadisticas_director.html', context)


@login_required
def generar_reporte_pdf_director(request):
    """
    Genera un reporte PDF con las estadísticas del Director de Carrera según el rango de tiempo seleccionado.
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    rango_seleccionado = request.GET.get('rango', 'mes')
    
    # Obtener datos usando la misma lógica que estadisticas_director
    now = timezone.localtime(timezone.now())
    today = now.date()
    
    fecha_inicio_dt = None
    fecha_fin_dt = timezone.make_aware(datetime.combine(today, datetime.max.time()))
    
    if rango_seleccionado == 'mes':
        fecha_inicio = today - timedelta(days=30)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Mes'
    elif rango_seleccionado == 'semestre':
        fecha_inicio = today - timedelta(days=180)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Semestre'
    elif rango_seleccionado == 'año':
        fecha_inicio = today.replace(month=1, day=1)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Año'
    else:
        fecha_inicio_dt = None
        rango_nombre = 'Histórico Completo'
    
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    carreras_ids = carreras_del_director.values_list('id', flat=True)
    
    solicitudes_base = Solicitudes.objects.filter(estudiantes__carreras__id__in=carreras_ids).distinct()
    ajustes_base = AjusteAsignado.objects.filter(solicitudes__estudiantes__carreras__id__in=carreras_ids).distinct()
    
    if fecha_inicio_dt:
        solicitudes_base = solicitudes_base.filter(created_at__gte=fecha_inicio_dt, created_at__lte=fecha_fin_dt)
        ajustes_base = ajustes_base.filter(solicitudes__created_at__gte=fecha_inicio_dt, solicitudes__created_at__lte=fecha_fin_dt)
    
    # Calcular estadísticas básicas
    total_casos = solicitudes_base.count()
    casos_aprobados = solicitudes_base.filter(estado='aprobado').count()
    total_ajustes = ajustes_base.count()
    ajustes_aprobados = ajustes_base.filter(estado_aprobacion='aprobado').count()
    ajustes_rechazados = ajustes_base.filter(estado_aprobacion='rechazado').count()
    ajustes_pendientes = ajustes_base.filter(estado_aprobacion='pendiente').count()
    tasa_aprobacion = round((casos_aprobados / total_casos * 100) if total_casos > 0 else 0, 1)
    
    # Estadísticas de Asignaturas
    asignaturas_base = Asignaturas.objects.filter(carreras__id__in=carreras_ids)
    total_asignaturas = asignaturas_base.count()
    asignaturas_activas = asignaturas_base.filter(is_active=True).count()
    asignaturas_inactivas = asignaturas_base.filter(is_active=False).count()
    
    # Asignaturas por semestre
    asignaturas_por_semestre = asignaturas_base.values('semestre', 'anio').annotate(
        total=Count('id')
    ).order_by('-anio', 'semestre')
    
    # Estadísticas de Estudiantes
    estudiantes_base = Estudiantes.objects.filter(carreras__id__in=carreras_ids)
    total_estudiantes = estudiantes_base.count()
    estudiantes_con_ajustes = estudiantes_base.filter(
        solicitudes__in=solicitudes_base
    ).distinct().count()
    
    # Estudiantes por semestre
    estudiantes_por_semestre = estudiantes_base.values('semestre_actual').annotate(
        total=Count('id')
    ).order_by('semestre_actual')
    
    # Estadísticas de Docentes
    # Obtener docentes que tienen asignaturas en las carreras del director
    docentes_ids = asignaturas_base.values_list('docente_id', flat=True).distinct()
    docentes_base = PerfilUsuario.objects.filter(id__in=docentes_ids, rol__nombre_rol='Docente')
    total_docentes = docentes_base.count()
    
    # Docentes que comentaron ajustes (más confiable)
    docentes_que_comentaron = ajustes_base.filter(
        docente_comentador__isnull=False,
        docente_comentador__in=docentes_base
    ).values('docente_comentador__usuario__first_name',
             'docente_comentador__usuario__last_name').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Docentes por asignatura
    docentes_por_asignatura = asignaturas_base.values(
        'docente__usuario__first_name',
        'docente__usuario__last_name'
    ).annotate(
        total_asignaturas=Count('id')
    ).order_by('-total_asignaturas')
    
    # Docentes con ajustes aprobados/rechazados (a través de asignaturas relacionadas)
    # Obtener ajustes que tienen asignaturas relacionadas
    docentes_ajustes_aprobados = {}
    docentes_ajustes_rechazados = {}
    
    # Obtener ajustes aprobados/rechazados con sus solicitudes y asignaturas
    ajustes_aprobados_con_asignaturas = ajustes_base.filter(
        estado_aprobacion='aprobado',
        solicitudes__asignaturas_solicitadas__docente__in=docentes_base
    ).select_related('solicitudes').prefetch_related('solicitudes__asignaturas_solicitadas__docente__usuario')
    
    ajustes_rechazados_con_asignaturas = ajustes_base.filter(
        estado_aprobacion='rechazado',
        solicitudes__asignaturas_solicitadas__docente__in=docentes_base
    ).select_related('solicitudes').prefetch_related('solicitudes__asignaturas_solicitadas__docente__usuario')
    
    # Contar aprobados por docente
    for ajuste in ajustes_aprobados_con_asignaturas:
        asignaturas_solicitud = ajuste.solicitudes.asignaturas_solicitadas.filter(
            docente__in=docentes_base
        ).select_related('docente__usuario')
        for asignatura in asignaturas_solicitud:
            if asignatura.docente:
                docente_nombre = f"{asignatura.docente.usuario.first_name} {asignatura.docente.usuario.last_name}"
                docentes_ajustes_aprobados[docente_nombre] = docentes_ajustes_aprobados.get(docente_nombre, 0) + 1
    
    # Contar rechazados por docente
    for ajuste in ajustes_rechazados_con_asignaturas:
        asignaturas_solicitud = ajuste.solicitudes.asignaturas_solicitadas.filter(
            docente__in=docentes_base
        ).select_related('docente__usuario')
        for asignatura in asignaturas_solicitud:
            if asignatura.docente:
                docente_nombre = f"{asignatura.docente.usuario.first_name} {asignatura.docente.usuario.last_name}"
                docentes_ajustes_rechazados[docente_nombre] = docentes_ajustes_rechazados.get(docente_nombre, 0) + 1
    
    # Estadísticas de Inscripciones (AsignaturasEnCurso)
    inscripciones_base = AsignaturasEnCurso.objects.filter(
        asignaturas__carreras__id__in=carreras_ids
    )
    total_inscripciones = inscripciones_base.count()
    inscripciones_activas = inscripciones_base.filter(estado=True).count()
    
    # Inscripciones por asignatura
    inscripciones_por_asignatura = inscripciones_base.values(
        'asignaturas__nombre',
        'asignaturas__seccion'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:20]  # Top 20
    
    # Crear el objeto HttpResponse con el tipo de contenido PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_director_{rango_seleccionado}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    # Crear el objeto PDF usando BytesIO
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Contenedor para los elementos del PDF
    elements = []
    
    # Estilos con colores rojo, blanco y negro
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#D32F2F'),  # Rojo INACAP
        spaceAfter=20,
        alignment=1,  # Centrado
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#000000'),  # Negro
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    # Colores del esquema
    color_rojo = colors.HexColor('#D32F2F')
    color_negro = colors.HexColor('#000000')
    color_blanco = colors.white
    color_gris_claro = colors.HexColor('#f5f5f5')
    
    # Título
    elements.append(Paragraph('Reporte de Estadísticas - Director de Carrera', title_style))
    elements.append(Paragraph(f'Rango de Tiempo: {rango_nombre}', heading_style))
    elements.append(Paragraph(f'Fecha de Generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elements.append(Paragraph(f'Generado por: {request.user.get_full_name()}', styles['Normal']))
    elements.append(Spacer(1, 0.4*inch))
    
    # Introducción
    intro_text = f"""
    Este reporte presenta un análisis completo de las estadísticas de las carreras bajo su dirección 
    para el período: <b>{rango_nombre}</b>. El documento incluye indicadores clave de rendimiento (KPIs), 
    análisis de casos y ajustes razonables, estadísticas por carrera, asignaturas, estudiantes y docentes.
    """
    intro_style = ParagraphStyle(
        'IntroStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        spaceAfter=15,
        alignment=4,  # Justificado
    )
    elements.append(Paragraph(intro_text, intro_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # KPIs Principales
    elements.append(Paragraph('Indicadores Principales', heading_style))
    kpi_text = """
    Los siguientes indicadores proporcionan una visión general del estado de las solicitudes de ajustes 
    razonables en sus carreras. Estos datos reflejan el total de casos gestionados, su estado de aprobación 
    y la distribución de ajustes asignados.
    """
    elements.append(Paragraph(kpi_text, intro_style))
    elements.append(Spacer(1, 0.1*inch))
    kpi_data = [
        ['Indicador', 'Valor'],
        ['Total Casos', str(total_casos)],
        ['Casos Aprobados', str(casos_aprobados)],
        ['Total Ajustes', str(total_ajustes)],
        ['Ajustes Aprobados', str(ajustes_aprobados)],
        ['Ajustes Rechazados', str(ajustes_rechazados)],
        ['Ajustes Pendientes', str(ajustes_pendientes)],
        ['Tasa de Aprobación', f"{tasa_aprobacion}%"],
        ['Total Asignaturas', str(total_asignaturas)],
        ['Asignaturas Activas', str(asignaturas_activas)],
        ['Asignaturas Inactivas', str(asignaturas_inactivas)],
        ['Total Estudiantes', str(total_estudiantes)],
        ['Estudiantes con Ajustes', str(estudiantes_con_ajustes)],
        ['Total Docentes', str(total_docentes)],
        ['Total Inscripciones', str(total_inscripciones)],
        ['Inscripciones Activas', str(inscripciones_activas)],
    ]
    kpi_table = Table(kpi_data, colWidths=[4*inch, 2*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Casos por Estado
    elements.append(Paragraph('Casos por Estado', heading_style))
    estado_data = [['Estado', 'Cantidad', 'Porcentaje']]
    for estado_valor, estado_nombre in Solicitudes.ESTADO_CHOICES:
        cantidad = solicitudes_base.filter(estado=estado_valor).count()
        porcentaje = round((cantidad / total_casos * 100) if total_casos > 0 else 0, 1)
        estado_data.append([estado_nombre, str(cantidad), f"{porcentaje}%"])
    
    estado_table = Table(estado_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    estado_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(estado_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Gráfico de Casos por Estado (Gráfico de pastel mejorado)
    try:
        estado_counts = {}
        estado_labels_short = {}
        for estado_valor, estado_nombre in Solicitudes.ESTADO_CHOICES:
            cantidad = solicitudes_base.filter(estado=estado_valor).count()
            if cantidad > 0:
                estado_counts[estado_nombre] = cantidad
                # Acortar etiquetas largas para mejor visualización
                if len(estado_nombre) > 30:
                    estado_labels_short[estado_nombre] = estado_nombre[:27] + '...'
                else:
                    estado_labels_short[estado_nombre] = estado_nombre
        
        if estado_counts:
            fig, ax = plt.subplots(figsize=(8, 6))
            colors_pie = ['#4CAF50', '#FF9800', '#f44336', '#2196F3', '#9E9E9E', '#FFC107', '#00BCD4']
            
            # Ordenar por cantidad descendente para mejor visualización
            sorted_estados = sorted(estado_counts.items(), key=lambda x: x[1], reverse=True)
            valores = [v for _, v in sorted_estados]
            etiquetas = [estado_labels_short[k] for k, _ in sorted_estados]
            
            wedges, texts, autotexts = ax.pie(
                valores,
                labels=etiquetas,
                autopct=lambda pct: f'{pct:.1f}%\n({int(pct/100*sum(valores))})' if pct > 3 else '',
                colors=colors_pie[:len(valores)],
                startangle=90,
                textprops={'fontsize': 8, 'fontweight': 'bold'},
                pctdistance=0.85,
                labeldistance=1.1
            )
            
            # Mejorar la legibilidad de los textos
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(9)
            
            for text in texts:
                text.set_fontsize(8)
            
            ax.set_title('Distribución de Casos por Estado', fontsize=12, fontweight='bold', pad=20)
            
            # Agregar leyenda fuera del gráfico
            ax.legend(wedges, [f'{k}: {v}' for k, v in sorted_estados], 
                     loc='center left', bbox_to_anchor=(1, 0, 0.5, 1), fontsize=8)
            
            plt.tight_layout()
            # Usar BytesIO en lugar de archivo temporal para evitar problemas de permisos
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight', facecolor='white')
            plt.close()
            img_buffer.seek(0)
            
            img = Image(img_buffer, width=7*inch, height=5.25*inch)
            elements.append(img)
            elements.append(Spacer(1, 0.2*inch))
            img_buffer.close()
    except Exception as e:
        pass
    
    # Estadísticas por Carrera
    elements.append(Paragraph('Estadísticas por Carrera', heading_style))
    carrera_text = """
    El análisis por carrera permite identificar qué programas académicos presentan mayor demanda de ajustes 
    razonables y su tasa de aprobación. Esta información es valiosa para la planificación académica y la 
    asignación de recursos.
    """
    elements.append(Paragraph(carrera_text, intro_style))
    elements.append(Spacer(1, 0.1*inch))
    carrera_data = [['Carrera', 'Total Casos', 'Aprobados', 'Tasa Aprobación']]
    for carrera in carreras_del_director:
        casos_carrera = solicitudes_base.filter(estudiantes__carreras=carrera)
        total_carrera = casos_carrera.count()
        aprobados_carrera = casos_carrera.filter(estado='aprobado').count()
        tasa_carrera = round((aprobados_carrera / total_carrera * 100) if total_carrera > 0 else 0, 1)
        carrera_data.append([carrera.nombre, str(total_carrera), str(aprobados_carrera), f"{tasa_carrera}%"])
    
    carrera_table = Table(carrera_data, colWidths=[3*inch, 1.5*inch, 1.5*inch, 1*inch])
    carrera_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(carrera_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Gráfico de Tasa de Aprobación por Carrera
    try:
        carrera_names = []
        tasa_aprobaciones = []
        for carrera in carreras_del_director:
            casos_carrera = solicitudes_base.filter(estudiantes__carreras=carrera)
            total_carrera = casos_carrera.count()
            if total_carrera > 0:
                aprobados_carrera = casos_carrera.filter(estado='aprobado').count()
                tasa_carrera = round((aprobados_carrera / total_carrera * 100), 1)
                carrera_names.append(carrera.nombre[:20])
                tasa_aprobaciones.append(tasa_carrera)
        
        if carrera_names:
            fig, ax = plt.subplots(figsize=(8, 5))
            bars = ax.barh(carrera_names, tasa_aprobaciones, color='#D32F2F', edgecolor='black', linewidth=1)
            ax.set_xlabel('Tasa de Aprobación (%)', fontsize=10, fontweight='bold')
            ax.set_title('Tasa de Aprobación por Carrera', fontsize=11, fontweight='bold', pad=15)
            ax.set_xlim(0, 100)
            ax.grid(axis='x', alpha=0.3, linestyle='--')
            
            # Agregar valores en las barras
            for i, bar in enumerate(bars):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2.,
                       f'{width}%',
                       ha='left', va='center', fontsize=9, fontweight='bold', pad=5)
            
            plt.tight_layout()
            # Usar BytesIO en lugar de archivo temporal para evitar problemas de permisos
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer.seek(0)
            
            img = Image(img_buffer, width=6*inch, height=3.75*inch)
            elements.append(img)
            elements.append(Spacer(1, 0.2*inch))
            img_buffer.close()
    except Exception as e:
        pass
    
    # Estadísticas de Asignaturas
    elements.append(Paragraph('Estadísticas de Asignaturas', heading_style))
    asignaturas_text = """
    Esta sección detalla las asignaturas ofrecidas en sus carreras, incluyendo información sobre docentes 
    asignados, estado de las asignaturas y distribución por semestre. Los datos ayudan a comprender la 
    estructura académica y la carga docente.
    """
    elements.append(Paragraph(asignaturas_text, intro_style))
    elements.append(Spacer(1, 0.1*inch))
    asignaturas_data = [['Asignatura', 'Sección', 'Carrera', 'Docente', 'Estado', 'Semestre']]
    for asignatura in asignaturas_base.select_related('carreras', 'docente__usuario')[:50]:  # Top 50
        docente_nombre = f"{asignatura.docente.usuario.first_name} {asignatura.docente.usuario.last_name}" if asignatura.docente else "Sin docente"
        estado = "Activa" if asignatura.is_active else "Inactiva"
        semestre_str = asignatura.periodo_completo if asignatura.semestre else "Sin periodo"
        asignaturas_data.append([
            asignatura.nombre[:30],
            asignatura.seccion,
            asignatura.carreras.nombre[:25],
            docente_nombre[:30],
            estado,
            semestre_str[:20]
        ])
    
    asignaturas_table = Table(asignaturas_data, colWidths=[1.2*inch, 0.8*inch, 1.2*inch, 1.2*inch, 0.8*inch, 1*inch])
    asignaturas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(asignaturas_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Asignaturas por Semestre
    if asignaturas_por_semestre:
        elements.append(Paragraph('Asignaturas por Semestre', heading_style))
        semestre_data = [['Semestre', 'Año', 'Total']]
        for item in asignaturas_por_semestre:
            semestre_nombre = dict(SEMESTRE_CHOICES).get(item['semestre'], item['semestre']) if item['semestre'] else "Sin semestre"
            semestre_data.append([semestre_nombre, str(item['anio']) if item['anio'] else "N/A", str(item['total'])])
        
        semestre_table = Table(semestre_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        semestre_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
            ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
            ('GRID', (0, 0), (-1, -1), 1, color_negro),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
        ]))
        elements.append(semestre_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Estadísticas de Estudiantes
    elements.append(Paragraph('Estadísticas de Estudiantes', heading_style))
    estudiantes_data = [['Carrera', 'Total Estudiantes', 'Con Ajustes', 'Porcentaje']]
    for carrera in carreras_del_director:
        estudiantes_carrera = estudiantes_base.filter(carreras=carrera)
        total_est_carrera = estudiantes_carrera.count()
        con_ajustes_carrera = estudiantes_carrera.filter(
            solicitudes__in=solicitudes_base
        ).distinct().count()
        porcentaje_ajustes = round((con_ajustes_carrera / total_est_carrera * 100) if total_est_carrera > 0 else 0, 1)
        estudiantes_data.append([
            carrera.nombre[:40],
            str(total_est_carrera),
            str(con_ajustes_carrera),
            f"{porcentaje_ajustes}%"
        ])
    
    estudiantes_table = Table(estudiantes_data, colWidths=[3*inch, 1.5*inch, 1.5*inch, 1*inch])
    estudiantes_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(estudiantes_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Gráfico de Estudiantes con Ajustes por Carrera
    try:
        carrera_est_names = []
        porcentajes_ajustes = []
        for carrera in carreras_del_director:
            estudiantes_carrera = estudiantes_base.filter(carreras=carrera)
            total_est_carrera = estudiantes_carrera.count()
            if total_est_carrera > 0:
                con_ajustes_carrera = estudiantes_carrera.filter(
                    solicitudes__in=solicitudes_base
                ).distinct().count()
                porcentaje_ajustes = round((con_ajustes_carrera / total_est_carrera * 100), 1)
                carrera_est_names.append(carrera.nombre[:20])
                porcentajes_ajustes.append(porcentaje_ajustes)
        
        if carrera_est_names:
            fig, ax = plt.subplots(figsize=(8, 5))
            bars = ax.bar(carrera_est_names, porcentajes_ajustes, color='#2196F3', edgecolor='black', linewidth=1)
            ax.set_ylabel('Porcentaje de Estudiantes (%)', fontsize=10, fontweight='bold')
            ax.set_title('Porcentaje de Estudiantes con Ajustes por Carrera', fontsize=11, fontweight='bold', pad=15)
            ax.set_ylim(0, max(porcentajes_ajustes) * 1.2 if porcentajes_ajustes else 100)
            ax.grid(axis='y', alpha=0.3, linestyle='--')
            plt.xticks(rotation=45, ha='right')
            
            # Agregar valores en las barras
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{height}%',
                       ha='center', va='bottom', fontsize=9, fontweight='bold')
            
            plt.tight_layout()
            # Usar BytesIO en lugar de archivo temporal para evitar problemas de permisos
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer.seek(0)
            
            img = Image(img_buffer, width=6*inch, height=3.75*inch)
            elements.append(img)
            elements.append(Spacer(1, 0.2*inch))
            img_buffer.close()
    except Exception as e:
        pass
    
    # Estudiantes por Semestre
    if estudiantes_por_semestre:
        elements.append(Paragraph('Estudiantes por Semestre Actual', heading_style))
        est_semestre_data = [['Semestre', 'Total Estudiantes']]
        for item in estudiantes_por_semestre:
            semestre_num = str(item['semestre_actual']) if item['semestre_actual'] else "Sin semestre"
            est_semestre_data.append([f"Semestre {semestre_num}", str(item['total'])])
        
        est_semestre_table = Table(est_semestre_data, colWidths=[2*inch, 2*inch])
        est_semestre_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
            ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
            ('GRID', (0, 0), (-1, -1), 1, color_negro),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
        ]))
        elements.append(est_semestre_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Estadísticas de Docentes
    elements.append(Paragraph('Estadísticas de Docentes', heading_style))
    docentes_text = """
    El análisis de docentes muestra la participación del cuerpo académico en el proceso de ajustes razonables. 
    Se incluye información sobre asignaturas asignadas, ajustes aprobados y rechazados, así como comentarios 
    realizados por los docentes durante el proceso de evaluación.
    """
    elements.append(Paragraph(docentes_text, intro_style))
    elements.append(Spacer(1, 0.1*inch))
    docentes_data = [['Docente', 'Total Asignaturas', 'Ajustes Aprobados', 'Ajustes Rechazados', 'Comentarios']]
    
    # Combinar datos de docentes
    docentes_dict = {}
    for item in docentes_por_asignatura:
        nombre = f"{item['docente__usuario__first_name']} {item['docente__usuario__last_name']}"
        docentes_dict[nombre] = {
            'asignaturas': item['total_asignaturas'],
            'aprobados': 0,
            'rechazados': 0,
            'comentarios': 0
        }
    
    # Agregar datos de aprobados y rechazados
    for nombre, cantidad in docentes_ajustes_aprobados.items():
        if nombre not in docentes_dict:
            docentes_dict[nombre] = {'asignaturas': 0, 'aprobados': 0, 'rechazados': 0, 'comentarios': 0}
        docentes_dict[nombre]['aprobados'] = cantidad
    
    for nombre, cantidad in docentes_ajustes_rechazados.items():
        if nombre not in docentes_dict:
            docentes_dict[nombre] = {'asignaturas': 0, 'aprobados': 0, 'rechazados': 0, 'comentarios': 0}
        docentes_dict[nombre]['rechazados'] = cantidad
    
    for item in docentes_que_comentaron:
        nombre = f"{item['docente_comentador__usuario__first_name']} {item['docente_comentador__usuario__last_name']}"
        if nombre not in docentes_dict:
            docentes_dict[nombre] = {'asignaturas': 0, 'aprobados': 0, 'rechazados': 0, 'comentarios': 0}
        docentes_dict[nombre]['comentarios'] = item['total']
    
    for nombre, datos in sorted(docentes_dict.items(), key=lambda x: x[1]['asignaturas'], reverse=True)[:30]:  # Top 30
        docentes_data.append([
            nombre[:35],
            str(datos['asignaturas']),
            str(datos['aprobados']),
            str(datos['rechazados']),
            str(datos['comentarios'])
        ])
    
    docentes_table = Table(docentes_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    docentes_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
        ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
        ('GRID', (0, 0), (-1, -1), 1, color_negro),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
    ]))
    elements.append(docentes_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Gráfico de Ajustes por Docente (Top 10)
    try:
        docentes_top = sorted(docentes_dict.items(), key=lambda x: x[1]['aprobados'] + x[1]['rechazados'], reverse=True)[:10]
        if docentes_top:
            docentes_nombres = [d[0][:15] for d in docentes_top]
            aprobados_data = [d[1]['aprobados'] for d in docentes_top]
            rechazados_data = [d[1]['rechazados'] for d in docentes_top]
            
            fig, ax = plt.subplots(figsize=(10, 6))
            x = range(len(docentes_nombres))
            width = 0.35
            bars1 = ax.bar([i - width/2 for i in x], aprobados_data, width, label='Aprobados', color='#4CAF50', edgecolor='black')
            bars2 = ax.bar([i + width/2 for i in x], rechazados_data, width, label='Rechazados', color='#f44336', edgecolor='black')
            ax.set_ylabel('Cantidad de Ajustes', fontsize=10, fontweight='bold')
            ax.set_title('Top 10 Docentes: Ajustes Aprobados vs Rechazados', fontsize=11, fontweight='bold', pad=15)
            ax.set_xticks(x)
            ax.set_xticklabels(docentes_nombres, rotation=45, ha='right')
            ax.legend(fontsize=9)
            ax.grid(axis='y', alpha=0.3, linestyle='--')
            plt.tight_layout()
            # Usar BytesIO en lugar de archivo temporal para evitar problemas de permisos
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer.seek(0)
            
            img = Image(img_buffer, width=7*inch, height=4.2*inch)
            elements.append(img)
            elements.append(Spacer(1, 0.2*inch))
            img_buffer.close()
    except Exception as e:
        pass
    
    # Inscripciones por Asignatura (Top 20)
    if inscripciones_por_asignatura:
        elements.append(Paragraph('Top 20 Asignaturas con Más Inscripciones', heading_style))
        inscripciones_data = [['Asignatura', 'Sección', 'Total Inscripciones']]
        for item in inscripciones_por_asignatura:
            inscripciones_data.append([
                item['asignaturas__nombre'][:40],
                item['asignaturas__seccion'],
                str(item['total'])
            ])
        
        inscripciones_table = Table(inscripciones_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        inscripciones_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), color_rojo),
            ('TEXTCOLOR', (0, 0), (-1, 0), color_blanco),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), color_gris_claro),
            ('GRID', (0, 0), (-1, -1), 1, color_negro),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [color_blanco, color_gris_claro]),
        ]))
        elements.append(inscripciones_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Conclusión
    elements.append(PageBreak())
    elements.append(Paragraph('Conclusiones y Recomendaciones', heading_style))
    conclusion_text = f"""
    <b>Resumen Ejecutivo:</b><br/><br/>
    
    Durante el período analizado ({rango_nombre}), se registraron <b>{total_casos}</b> casos de solicitudes de ajustes 
    razonables en las carreras bajo su dirección. De estos, <b>{casos_aprobados}</b> fueron aprobados, lo que representa 
    una tasa de aprobación del <b>{tasa_aprobacion}%</b>.<br/><br/>
    
    Se asignaron un total de <b>{total_ajustes}</b> ajustes razonables, de los cuales <b>{ajustes_aprobados}</b> fueron 
    aprobados y <b>{ajustes_rechazados}</b> fueron rechazados. Actualmente hay <b>{ajustes_pendientes}</b> ajustes en 
    estado pendiente de evaluación.<br/><br/>
    
    <b>Recomendaciones:</b><br/>
    • Continuar monitoreando la tasa de aprobación para identificar tendencias.<br/>
    • Revisar los casos pendientes para agilizar el proceso de evaluación.<br/>
    • Analizar las carreras con mayor demanda de ajustes para identificar necesidades específicas.<br/>
    • Mantener comunicación fluida con docentes para la implementación efectiva de los ajustes aprobados.<br/><br/>
    
    Este reporte fue generado el {timezone.now().strftime("%d de %B de %Y a las %H:%M")} horas.
    """
    conclusion_style = ParagraphStyle(
        'ConclusionStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        spaceAfter=15,
        alignment=4,  # Justificado
    )
    elements.append(Paragraph(conclusion_text, conclusion_style))
    
    # Construir el PDF
    doc.build(elements)
    
    # Obtener el valor del buffer y escribirlo en la respuesta
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


@login_required
def generar_reporte_excel_director(request):
    """
    Genera un archivo Excel con los datos según el rango de tiempo seleccionado.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    rango_seleccionado = request.GET.get('rango', 'mes')
    
    # Obtener datos usando la misma lógica que estadisticas_director
    now = timezone.localtime(timezone.now())
    today = now.date()
    
    fecha_inicio_dt = None
    fecha_fin_dt = timezone.make_aware(datetime.combine(today, datetime.max.time()))
    
    if rango_seleccionado == 'mes':
        fecha_inicio = today - timedelta(days=30)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Mes'
    elif rango_seleccionado == 'semestre':
        fecha_inicio = today - timedelta(days=180)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Semestre'
    elif rango_seleccionado == 'año':
        fecha_inicio = today.replace(month=1, day=1)
        fecha_inicio_dt = timezone.make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
        rango_nombre = 'Último Año'
    else:
        fecha_inicio_dt = None
        rango_nombre = 'Histórico Completo'
    
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    carreras_ids = carreras_del_director.values_list('id', flat=True)
    
    solicitudes_base = Solicitudes.objects.filter(estudiantes__carreras__id__in=carreras_ids).distinct()
    ajustes_base = AjusteAsignado.objects.filter(solicitudes__estudiantes__carreras__id__in=carreras_ids).distinct()
    
    if fecha_inicio_dt:
        solicitudes_base = solicitudes_base.filter(created_at__gte=fecha_inicio_dt, created_at__lte=fecha_fin_dt)
        ajustes_base = ajustes_base.filter(solicitudes__created_at__gte=fecha_inicio_dt, solicitudes__created_at__lte=fecha_fin_dt)
    
    # Estadísticas adicionales (misma lógica que PDF)
    asignaturas_base = Asignaturas.objects.filter(carreras__id__in=carreras_ids)
    estudiantes_base = Estudiantes.objects.filter(carreras__id__in=carreras_ids)
    docentes_ids = asignaturas_base.values_list('docente_id', flat=True).distinct()
    docentes_base = PerfilUsuario.objects.filter(id__in=docentes_ids, rol__nombre_rol='Docente')
    inscripciones_base = AsignaturasEnCurso.objects.filter(asignaturas__carreras__id__in=carreras_ids)
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Estadísticas"
    
    # Estilos
    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    row = 1
    
    # Título
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = f"Reporte de Estadísticas Director - {rango_nombre}"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # KPIs
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "Indicadores Principales (KPIs)"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['KPI', 'Valor', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Calcular KPIs
    total_casos = solicitudes_base.count()
    casos_aprobados = solicitudes_base.filter(estado='aprobado').count()
    total_ajustes = ajustes_base.count()
    ajustes_aprobados = ajustes_base.filter(estado_aprobacion='aprobado').count()
    ajustes_rechazados = ajustes_base.filter(estado_aprobacion='rechazado').count()
    ajustes_pendientes = ajustes_base.filter(estado_aprobacion='pendiente').count()
    total_asignaturas = asignaturas_base.count()
    asignaturas_activas = asignaturas_base.filter(is_active=True).count()
    asignaturas_inactivas = asignaturas_base.filter(is_active=False).count()
    total_estudiantes = estudiantes_base.count()
    estudiantes_con_ajustes = estudiantes_base.filter(solicitudes__in=solicitudes_base).distinct().count()
    total_docentes = docentes_base.count()
    total_inscripciones = inscripciones_base.count()
    inscripciones_activas = inscripciones_base.filter(estado=True).count()
    
    kpis_data = [
        ['Total Casos', total_casos],
        ['Casos Aprobados', casos_aprobados],
        ['Total Ajustes', total_ajustes],
        ['Ajustes Aprobados', ajustes_aprobados],
        ['Ajustes Rechazados', ajustes_rechazados],
        ['Ajustes Pendientes', ajustes_pendientes],
        ['Total Asignaturas', total_asignaturas],
        ['Asignaturas Activas', asignaturas_activas],
        ['Asignaturas Inactivas', asignaturas_inactivas],
        ['Total Estudiantes', total_estudiantes],
        ['Estudiantes con Ajustes', estudiantes_con_ajustes],
        ['Total Docentes', total_docentes],
        ['Total Inscripciones', total_inscripciones],
        ['Inscripciones Activas', inscripciones_activas],
    ]
    
    for kpi, valor in kpis_data:
        ws.cell(row=row, column=1).value = kpi
        ws.cell(row=row, column=2).value = valor
        ws.cell(row=row, column=3).value = rango_nombre
        row += 1
    
    row += 1
    
    # Casos por Estado
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "Casos por Estado"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['Estado', 'Cantidad', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    for estado_valor, estado_nombre in Solicitudes.ESTADO_CHOICES:
        cantidad = solicitudes_base.filter(estado=estado_valor).count()
        ws.cell(row=row, column=1).value = estado_nombre
        ws.cell(row=row, column=2).value = cantidad
        ws.cell(row=row, column=3).value = rango_nombre
        row += 1
    
    row += 1
    
    # Estadísticas por Carrera
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'].value = "Estadísticas por Carrera"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['Carrera', 'Total Casos', 'Aprobados', 'Tasa Aprobación', 'Total Estudiantes', 'Estudiantes con Ajustes', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    for carrera in carreras_del_director:
        casos_carrera = solicitudes_base.filter(estudiantes__carreras=carrera)
        total_carrera = casos_carrera.count()
        aprobados_carrera = casos_carrera.filter(estado='aprobado').count()
        tasa_carrera = round((aprobados_carrera / total_carrera * 100) if total_carrera > 0 else 0, 1)
        estudiantes_carrera = estudiantes_base.filter(carreras=carrera)
        total_est_carrera = estudiantes_carrera.count()
        est_con_ajustes = estudiantes_carrera.filter(solicitudes__in=solicitudes_base).distinct().count()
        
        ws.cell(row=row, column=1).value = carrera.nombre
        ws.cell(row=row, column=2).value = total_carrera
        ws.cell(row=row, column=3).value = aprobados_carrera
        ws.cell(row=row, column=4).value = f"{tasa_carrera}%"
        ws.cell(row=row, column=5).value = total_est_carrera
        ws.cell(row=row, column=6).value = est_con_ajustes
        ws.cell(row=row, column=7).value = rango_nombre
        row += 1
    
    row += 1
    
    # Asignaturas por Semestre
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = "Asignaturas por Semestre"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['Semestre', 'Año', 'Total', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    asignaturas_por_semestre = asignaturas_base.values('semestre', 'anio').annotate(total=Count('id')).order_by('-anio', 'semestre')
    for item in asignaturas_por_semestre:
        semestre_nombre = dict(Asignaturas.SEMESTRE_CHOICES).get(item['semestre'], item['semestre']) if item['semestre'] else "Sin semestre"
        ws.cell(row=row, column=1).value = semestre_nombre
        ws.cell(row=row, column=2).value = str(item['anio']) if item['anio'] else "N/A"
        ws.cell(row=row, column=3).value = item['total']
        ws.cell(row=row, column=4).value = rango_nombre
        row += 1
    
    row += 1
    
    # Estudiantes por Semestre
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = "Estudiantes por Semestre Actual"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['Semestre', 'Total Estudiantes', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    estudiantes_por_semestre = estudiantes_base.values('semestre_actual').annotate(total=Count('id')).order_by('semestre_actual')
    for item in estudiantes_por_semestre:
        semestre_num = f"Semestre {item['semestre_actual']}" if item['semestre_actual'] else "Sin semestre"
        ws.cell(row=row, column=1).value = semestre_num
        ws.cell(row=row, column=2).value = item['total']
        ws.cell(row=row, column=3).value = rango_nombre
        row += 1
    
    row += 1
    
    # Estadísticas de Docentes
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].value = "Estadísticas de Docentes"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['Docente', 'Total Asignaturas', 'Ajustes Aprobados', 'Ajustes Rechazados', 'Comentarios', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Obtener datos de docentes (similar a PDF)
    docentes_por_asignatura = asignaturas_base.values(
        'docente__usuario__first_name',
        'docente__usuario__last_name'
    ).annotate(total_asignaturas=Count('id')).order_by('-total_asignaturas')
    
    docentes_que_comentaron = ajustes_base.filter(
        docente_comentador__isnull=False,
        docente_comentador__in=docentes_base
    ).values('docente_comentador__usuario__first_name',
             'docente_comentador__usuario__last_name').annotate(total=Count('id')).order_by('-total')
    
    # Agrupar datos de docentes
    docentes_dict = {}
    for item in docentes_por_asignatura:
        nombre = f"{item['docente__usuario__first_name']} {item['docente__usuario__last_name']}"
        docentes_dict[nombre] = {
            'asignaturas': item['total_asignaturas'],
            'aprobados': 0,
            'rechazados': 0,
            'comentarios': 0
        }
    
    # Agregar comentarios
    for item in docentes_que_comentaron:
        nombre = f"{item['docente_comentador__usuario__first_name']} {item['docente_comentador__usuario__last_name']}"
        if nombre not in docentes_dict:
            docentes_dict[nombre] = {'asignaturas': 0, 'aprobados': 0, 'rechazados': 0, 'comentarios': 0}
        docentes_dict[nombre]['comentarios'] = item['total']
    
    # Agregar aprobados/rechazados (a través de asignaturas)
    ajustes_con_asignaturas = ajustes_base.filter(solicitudes__asignaturas_solicitadas__docente__in=docentes_base)
    for ajuste in ajustes_con_asignaturas.select_related('solicitudes'):
        asignaturas_solicitud = ajuste.solicitudes.asignaturas_solicitadas.filter(docente__in=docentes_base)
        for asignatura in asignaturas_solicitud:
            docente_nombre = f"{asignatura.docente.usuario.first_name} {asignatura.docente.usuario.last_name}"
            if docente_nombre not in docentes_dict:
                docentes_dict[docente_nombre] = {'asignaturas': 0, 'aprobados': 0, 'rechazados': 0, 'comentarios': 0}
            if ajuste.estado_aprobacion == 'aprobado':
                docentes_dict[docente_nombre]['aprobados'] = docentes_dict[docente_nombre].get('aprobados', 0) + 1
            elif ajuste.estado_aprobacion == 'rechazado':
                docentes_dict[docente_nombre]['rechazados'] = docentes_dict[docente_nombre].get('rechazados', 0) + 1
    
    for nombre, datos in sorted(docentes_dict.items(), key=lambda x: x[1]['asignaturas'], reverse=True):
        ws.cell(row=row, column=1).value = nombre
        ws.cell(row=row, column=2).value = datos['asignaturas']
        ws.cell(row=row, column=3).value = datos['aprobados']
        ws.cell(row=row, column=4).value = datos['rechazados']
        ws.cell(row=row, column=5).value = datos['comentarios']
        ws.cell(row=row, column=6).value = rango_nombre
        row += 1
    
    row += 1
    
    # Inscripciones por Asignatura (Top 20)
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = "Top 20 Asignaturas con Más Inscripciones"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['Asignatura', 'Sección', 'Total Inscripciones', 'Rango de Tiempo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    inscripciones_por_asignatura = inscripciones_base.values(
        'asignaturas__nombre',
        'asignaturas__seccion'
    ).annotate(total=Count('id')).order_by('-total')[:20]
    for item in inscripciones_por_asignatura:
        ws.cell(row=row, column=1).value = item['asignaturas__nombre']
        ws.cell(row=row, column=2).value = item['asignaturas__seccion']
        ws.cell(row=row, column=3).value = item['total']
        ws.cell(row=row, column=4).value = rango_nombre
        row += 1
    
    row += 1
    
    # Detalle de Casos
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].value = "Detalle de Casos"
    ws[f'A{row}'].font = title_font
    row += 1
    
    headers = ['ID', 'Estudiante', 'Carrera', 'Estado', 'Fecha Creación', 'Asunto']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    casos = solicitudes_base.select_related('estudiantes', 'estudiantes__carreras')[:1000]
    for caso in casos:
        estudiante_nombre = f"{caso.estudiantes.nombres} {caso.estudiantes.apellidos}" if caso.estudiantes else "N/A"
        carrera_nombre = caso.estudiantes.carreras.nombre if caso.estudiantes and caso.estudiantes.carreras else "N/A"
        fecha_creacion = timezone.localtime(caso.created_at).strftime('%Y-%m-%d %H:%M:%S') if caso.created_at else "N/A"
        
        ws.cell(row=row, column=1).value = caso.id
        ws.cell(row=row, column=2).value = estudiante_nombre
        ws.cell(row=row, column=3).value = carrera_nombre
        ws.cell(row=row, column=4).value = caso.get_estado_display()
        ws.cell(row=row, column=5).value = fecha_creacion
        ws.cell(row=row, column=6).value = caso.asunto[:50] if caso.asunto else "N/A"
        row += 1
    
    # Ajustar ancho de columnas
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                # Saltar celdas fusionadas que no tienen valor directamente
                if hasattr(cell, 'value') and cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear respuesta HTTP usando BytesIO para evitar problemas
    try:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_excel_director_{rango_seleccionado}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        output.close()
        return response
    except Exception as e:
        return HttpResponse(f'Error al generar el archivo Excel: {str(e)}', status=500)


# ----------------------------------------------------
#           GESTIÓN DE ASIGNATURAS (DIRECTOR)
# ----------------------------------------------------

@login_required
def gestion_asignaturas_director(request):
    """
    Vista para que el Director gestione las asignaturas de sus carreras.
    Permite ver, activar/desactivar y filtrar asignaturas.
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    # Desactivar automáticamente asignaturas de semestres vencidos
    asignaturas_desactivadas = desactivar_asignaturas_semestre_vencido()
    if asignaturas_desactivadas > 0:
        messages.info(request, f'Se desactivaron automáticamente {asignaturas_desactivadas} asignatura(s) de semestres anteriores.')
    
    # Obtener carreras del director
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    
    # Filtros
    filtro_estado = request.GET.get('estado', 'todas')  # todas, activas, inactivas
    filtro_carrera = request.GET.get('carrera', '')
    filtro_semestre = request.GET.get('semestre', '')
    
    # Obtener asignaturas
    asignaturas = Asignaturas.objects.filter(
        carreras__in=carreras_del_director
    ).select_related('carreras', 'docente__usuario').order_by('-is_active', 'nombre', 'seccion')
    
    # Aplicar filtros
    if filtro_estado == 'activas':
        asignaturas = asignaturas.filter(is_active=True)
    elif filtro_estado == 'inactivas':
        asignaturas = asignaturas.filter(is_active=False)
    
    if filtro_carrera:
        asignaturas = asignaturas.filter(carreras_id=filtro_carrera)
    
    if filtro_semestre:
        asignaturas = asignaturas.filter(semestre=filtro_semestre)
    
    # Estadísticas
    total_asignaturas = Asignaturas.objects.filter(carreras__in=carreras_del_director).count()
    total_activas = Asignaturas.objects.filter(carreras__in=carreras_del_director, is_active=True).count()
    total_inactivas = Asignaturas.objects.filter(carreras__in=carreras_del_director, is_active=False).count()
    
    context = {
        'asignaturas': asignaturas,
        'carreras': carreras_del_director,
        'total_asignaturas': total_asignaturas,
        'total_activas': total_activas,
        'total_inactivas': total_inactivas,
        'filtro_estado': filtro_estado,
        'filtro_carrera': filtro_carrera,
        'filtro_semestre': filtro_semestre,
        'semestres': [('otono', 'Otoño (Marzo-Julio)'), ('primavera', 'Primavera (Agosto-Diciembre)')],
    }
    
    return render(request, 'SIAPE/gestion_asignaturas_director.html', context)


@login_required
@require_POST
def toggle_asignatura_estado(request, asignatura_id):
    """
    Activa o desactiva una asignatura.
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    asignatura = get_object_or_404(Asignaturas, id=asignatura_id, carreras__in=carreras_del_director)
    
    # Toggle estado
    asignatura.is_active = not asignatura.is_active
    asignatura.save()
    
    estado_texto = "activada" if asignatura.is_active else "desactivada"
    messages.success(request, f'Asignatura "{asignatura.nombre} - {asignatura.seccion}" {estado_texto} correctamente.')
    
    return redirect('gestion_asignaturas_director')


@login_required
@require_POST
def bulk_toggle_asignaturas(request):
    """
    Activa o desactiva múltiples asignaturas a la vez.
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    accion = request.POST.get('accion')  # 'activar' o 'desactivar'
    asignaturas_ids = request.POST.getlist('asignaturas_ids')
    
    if not asignaturas_ids:
        messages.warning(request, 'No se seleccionaron asignaturas.')
        return redirect('gestion_asignaturas_director')
    
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    
    # Filtrar solo asignaturas del director
    asignaturas = Asignaturas.objects.filter(
        id__in=asignaturas_ids,
        carreras__in=carreras_del_director
    )
    
    nuevo_estado = accion == 'activar'
    count = asignaturas.update(is_active=nuevo_estado)
    
    estado_texto = "activadas" if nuevo_estado else "desactivadas"
    messages.success(request, f'{count} asignatura(s) {estado_texto} correctamente.')
    
    return redirect('gestion_asignaturas_director')


# ----------------------------------------------------
#           CARGA MASIVA DE DATOS (DIRECTOR)
# ----------------------------------------------------

@login_required
def gestion_carga_masiva_director(request):
    """
    Panel principal para la gestión de carga masiva de datos.
    Permite al Director de Carrera subir archivos Excel con:
    - Estudiantes
    - Asignaturas
    - Docentes
    - Asignaciones estudiante-asignatura
    - Asignaciones docente-asignatura
    """
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    # Obtener carreras del director
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    
    context = {
        'nombre_usuario': request.user.first_name,
        'carreras': carreras_del_director,
    }
    
    return render(request, 'SIAPE/gestion_carga_masiva_director.html', context)


@login_required
@require_POST
def cargar_estudiantes_excel(request):
    """
    Procesa un archivo Excel con datos de estudiantes.
    Columnas esperadas: RUT, Nombres, Apellidos, Email, Telefono (opcional), Carrera_ID
    """
    import openpyxl
    from django.db import transaction
    
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            return Response({'error': 'No tienes permisos'}, status=403)
    except AttributeError:
        return Response({'error': 'Perfil no encontrado'}, status=403)
    
    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Debe seleccionar un archivo Excel.')
        return redirect('gestion_carga_masiva_director')
    
    # Validar extensión
    if not archivo.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'El archivo debe ser un Excel (.xlsx o .xls).')
        return redirect('gestion_carga_masiva_director')
    
    # Obtener carreras del director para validación
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    carreras_ids = list(carreras_del_director.values_list('id', flat=True))
    
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Validar encabezados
        headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
        required_headers = ['rut', 'nombres', 'apellidos', 'email', 'semestre_actual']
        
        for h in required_headers:
            if h not in headers:
                messages.error(request, f'El archivo debe contener la columna: {h.upper()}')
                return redirect('gestion_carga_masiva_director')
        
        # Obtener índices de columnas
        col_idx = {h: headers.index(h) for h in headers if h}
        
        creados = 0
        actualizados = 0
        errores = []
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    rut = str(row[col_idx.get('rut', 0)] or '').strip()
                    nombres = str(row[col_idx.get('nombres', 1)] or '').strip()
                    apellidos = str(row[col_idx.get('apellidos', 2)] or '').strip()
                    email = str(row[col_idx.get('email', 3)] or '').strip()
                    telefono = row[col_idx.get('telefono', -1)] if 'telefono' in col_idx else None
                    carrera_id = row[col_idx.get('carrera_id', -1)] if 'carrera_id' in col_idx else None
                    semestre_actual = row[col_idx.get('semestre_actual', -1)] if 'semestre_actual' in col_idx else None
                    
                    # Validaciones básicas
                    if not rut or not nombres or not apellidos or not email:
                        errores.append(f'Fila {row_num}: Datos incompletos')
                        continue
                    
                    # Validar RUT
                    es_valido, mensaje_error = validar_rut_chileno(rut)
                    if not es_valido:
                        errores.append(f'Fila {row_num}: {mensaje_error}')
                        continue
                    
                    # Validar semestre
                    semestre_valido = None
                    if semestre_actual is not None:
                        try:
                            semestre_int = int(semestre_actual)
                            if 1 <= semestre_int <= 8:
                                semestre_valido = semestre_int
                            else:
                                errores.append(f'Fila {row_num}: Semestre debe estar entre 1 y 8')
                                continue
                        except (ValueError, TypeError):
                            errores.append(f'Fila {row_num}: Semestre inválido (debe ser un número entre 1 y 8)')
                            continue
                    else:
                        errores.append(f'Fila {row_num}: Semestre actual es requerido')
                        continue
                    
                    # Validar carrera (si se proporciona)
                    carrera = None
                    if carrera_id:
                        try:
                            carrera_id_int = int(carrera_id)
                            if carrera_id_int not in carreras_ids:
                                errores.append(f'Fila {row_num}: Carrera {carrera_id} no pertenece a tus carreras asignadas')
                                continue
                            carrera = Carreras.objects.get(id=carrera_id_int)
                        except (ValueError, Carreras.DoesNotExist):
                            errores.append(f'Fila {row_num}: Carrera ID inválido')
                            continue
                    else:
                        # Si no se especifica carrera, usar la primera del director
                        carrera = carreras_del_director.first()
                    
                    if not carrera:
                        errores.append(f'Fila {row_num}: No se pudo determinar la carrera')
                        continue
                    
                    # Crear o actualizar estudiante
                    estudiante, created = Estudiantes.objects.update_or_create(
                        rut=rut,
                        defaults={
                            'nombres': nombres,
                            'apellidos': apellidos,
                            'email': email,
                            'numero': int(telefono) if telefono and str(telefono).isdigit() else None,
                            'carreras': carrera,
                            'semestre_actual': semestre_valido
                        }
                    )
                    
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {row_num}: {str(e)}')
        
        # Mensaje de resultado
        msg = f'Proceso completado: {creados} estudiantes creados, {actualizados} actualizados.'
        if errores:
            msg += f' {len(errores)} errores encontrados.'
            for error in errores[:5]:  # Mostrar solo los primeros 5 errores
                messages.warning(request, error)
        messages.success(request, msg)
        
    except Exception as e:
        messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return redirect('gestion_carga_masiva_director')


@login_required
@require_POST
def cargar_docentes_excel(request):
    """
    Procesa un archivo Excel con datos de docentes.
    Columnas esperadas: RUT, Nombres, Apellidos, Email, Password (opcional)
    Crea Usuario + PerfilUsuario con rol Docente.
    """
    import openpyxl
    from django.db import transaction
    
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Debe seleccionar un archivo Excel.')
        return redirect('gestion_carga_masiva_director')
    
    if not archivo.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'El archivo debe ser un Excel (.xlsx o .xls).')
        return redirect('gestion_carga_masiva_director')
    
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
        required_headers = ['rut', 'nombres', 'apellidos', 'email']
        
        for h in required_headers:
            if h not in headers:
                messages.error(request, f'El archivo debe contener la columna: {h.upper()}')
                return redirect('gestion_carga_masiva_director')
        
        col_idx = {h: headers.index(h) for h in headers if h}
        
        # Obtener rol Docente
        rol_docente = Roles.objects.filter(nombre_rol='Docente').first()
        if not rol_docente:
            messages.error(request, 'No existe el rol "Docente" en el sistema.')
            return redirect('gestion_carga_masiva_director')
        
        # Obtener área del director para asignar a los docentes
        area_director = perfil_director.area
        
        creados = 0
        actualizados = 0
        errores = []
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    rut = str(row[col_idx.get('rut', 0)] or '').strip()
                    nombres = str(row[col_idx.get('nombres', 1)] or '').strip()
                    apellidos = str(row[col_idx.get('apellidos', 2)] or '').strip()
                    email = str(row[col_idx.get('email', 3)] or '').strip()
                    password = str(row[col_idx.get('password', -1)] or '') if 'password' in col_idx else None
                    
                    if not rut or not nombres or not apellidos or not email:
                        errores.append(f'Fila {row_num}: Datos incompletos')
                        continue
                    
                    # Verificar si el usuario ya existe
                    usuario_existente = Usuario.objects.filter(Q(rut=rut) | Q(email=email)).first()
                    
                    if usuario_existente:
                        # Actualizar datos existentes
                        usuario_existente.first_name = nombres
                        usuario_existente.last_name = apellidos
                        usuario_existente.save()
                        
                        # Asegurar que tenga perfil de docente
                        perfil, _ = PerfilUsuario.objects.get_or_create(
                            usuario=usuario_existente,
                            defaults={'rol': rol_docente, 'area': area_director}
                        )
                        if perfil.rol != rol_docente:
                            perfil.rol = rol_docente
                            perfil.save()
                        
                        actualizados += 1
                    else:
                        # Crear nuevo usuario
                        default_password = password if password else f'{rut[:4]}Docente!'
                        
                        usuario = Usuario.objects.create_user(
                            email=email,
                            password=default_password,
                            first_name=nombres,
                            last_name=apellidos,
                            rut=rut
                        )
                        
                        # Crear perfil de docente
                        PerfilUsuario.objects.create(
                            usuario=usuario,
                            rol=rol_docente,
                            area=area_director
                        )
                        
                        creados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {row_num}: {str(e)}')
        
        msg = f'Proceso completado: {creados} docentes creados, {actualizados} actualizados.'
        if errores:
            msg += f' {len(errores)} errores encontrados.'
            for error in errores[:5]:
                messages.warning(request, error)
        messages.success(request, msg)
        
    except Exception as e:
        messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return redirect('gestion_carga_masiva_director')


@login_required
@require_POST
def cargar_asignaturas_excel(request):
    """
    Procesa un archivo Excel con datos de asignaturas.
    Columnas esperadas: Nombre, Seccion, Carrera_ID, Docente_RUT (o Docente_Email)
    """
    import openpyxl
    from django.db import transaction
    
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Debe seleccionar un archivo Excel.')
        return redirect('gestion_carga_masiva_director')
    
    if not archivo.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'El archivo debe ser un Excel (.xlsx o .xls).')
        return redirect('gestion_carga_masiva_director')
    
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    carreras_ids = list(carreras_del_director.values_list('id', flat=True))
    
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
        required_headers = ['nombre', 'seccion']
        
        for h in required_headers:
            if h not in headers:
                messages.error(request, f'El archivo debe contener la columna: {h.upper()}')
                return redirect('gestion_carga_masiva_director')
        
        col_idx = {h: headers.index(h) for h in headers if h}
        
        creados = 0
        actualizados = 0
        errores = []
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    nombre = str(row[col_idx.get('nombre', 0)] or '').strip()
                    seccion = str(row[col_idx.get('seccion', 1)] or '').strip()
                    carrera_id = row[col_idx.get('carrera_id', -1)] if 'carrera_id' in col_idx else None
                    docente_rut = str(row[col_idx.get('docente_rut', -1)] or '').strip() if 'docente_rut' in col_idx else None
                    docente_email = str(row[col_idx.get('docente_email', -1)] or '').strip() if 'docente_email' in col_idx else None
                    
                    if not nombre or not seccion:
                        errores.append(f'Fila {row_num}: Datos incompletos')
                        continue
                    
                    # Determinar carrera
                    carrera = None
                    if carrera_id:
                        try:
                            carrera_id_int = int(carrera_id)
                            if carrera_id_int not in carreras_ids:
                                errores.append(f'Fila {row_num}: Carrera {carrera_id} no pertenece a tus carreras')
                                continue
                            carrera = Carreras.objects.get(id=carrera_id_int)
                        except (ValueError, Carreras.DoesNotExist):
                            errores.append(f'Fila {row_num}: Carrera ID inválido')
                            continue
                    else:
                        carrera = carreras_del_director.first()
                    
                    if not carrera:
                        errores.append(f'Fila {row_num}: No se pudo determinar la carrera')
                        continue
                    
                    # Buscar docente
                    docente_perfil = None
                    if docente_rut:
                        docente_perfil = PerfilUsuario.objects.filter(
                            usuario__rut=docente_rut,
                            rol__nombre_rol='Docente'
                        ).first()
                    elif docente_email:
                        docente_perfil = PerfilUsuario.objects.filter(
                            usuario__email=docente_email,
                            rol__nombre_rol='Docente'
                        ).first()
                    
                    if not docente_perfil:
                        errores.append(f'Fila {row_num}: No se encontró el docente especificado')
                        continue
                    
                    # Determinar semestre y año actual
                    hoy = timezone.localtime().date()
                    anio_actual = hoy.year
                    mes_actual = hoy.month
                    if mes_actual >= 3 and mes_actual <= 7:
                        semestre_actual = 'otono'
                    else:
                        semestre_actual = 'primavera'
                    
                    # Crear o actualizar asignatura
                    asignatura, created = Asignaturas.objects.update_or_create(
                        nombre=nombre,
                        seccion=seccion,
                        carreras=carrera,
                        defaults={
                            'docente': docente_perfil,
                            'semestre': semestre_actual,
                            'anio': anio_actual,
                            'is_active': True
                        }
                    )
                    
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {row_num}: {str(e)}')
        
        msg = f'Proceso completado: {creados} asignaturas creadas, {actualizados} actualizadas.'
        if errores:
            msg += f' {len(errores)} errores encontrados.'
            for error in errores[:5]:
                messages.warning(request, error)
        messages.success(request, msg)
        
    except Exception as e:
        messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return redirect('gestion_carga_masiva_director')


@login_required
@require_POST
def cargar_inscripciones_excel(request):
    """
    Procesa un archivo Excel para inscribir estudiantes en asignaturas.
    Columnas esperadas: Estudiante_RUT, Asignatura_Nombre, Asignatura_Seccion
    """
    import openpyxl
    from django.db import transaction
    
    try:
        perfil_director = request.user.perfil
        if perfil_director.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    archivo = request.FILES.get('archivo_excel')
    if not archivo:
        messages.error(request, 'Debe seleccionar un archivo Excel.')
        return redirect('gestion_carga_masiva_director')
    
    if not archivo.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'El archivo debe ser un Excel (.xlsx o .xls).')
        return redirect('gestion_carga_masiva_director')
    
    carreras_del_director = Carreras.objects.filter(director=perfil_director)
    
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
        
        # Validar columnas requeridas
        columnas_requeridas = ['estudiante_rut', 'asignatura_nombre', 'asignatura_seccion']
        columnas_faltantes = [col for col in columnas_requeridas if col not in headers]
        if columnas_faltantes:
            messages.error(request, f'El archivo debe contener las columnas: {", ".join([c.upper() for c in columnas_faltantes])}')
            return redirect('gestion_carga_masiva_director')
        
        col_idx = {h: headers.index(h) for h in headers if h}
        
        creados = 0
        ya_existentes = 0
        errores = []
        
        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Obtener valores de las celdas, manejando None y valores vacíos
                    estudiante_rut_celda = row[col_idx.get('estudiante_rut', 0)]
                    asignatura_nombre_celda = row[col_idx.get('asignatura_nombre', 1)]
                    asignatura_seccion_celda = row[col_idx.get('asignatura_seccion', 2)]
                    
                    # Convertir a string y limpiar (manejar números del Excel)
                    if estudiante_rut_celda is None:
                        estudiante_rut_raw = ''
                    elif isinstance(estudiante_rut_celda, (int, float)):
                        # Si viene como número, convertirlo a string sin decimales
                        estudiante_rut_raw = str(int(estudiante_rut_celda)).strip()
                    else:
                        estudiante_rut_raw = str(estudiante_rut_celda).strip()
                    
                    asignatura_nombre = str(asignatura_nombre_celda).strip() if asignatura_nombre_celda is not None else ''
                    asignatura_seccion = str(asignatura_seccion_celda).strip() if asignatura_seccion_celda is not None else ''
                    
                    # Validar que no estén vacíos
                    if not estudiante_rut_raw or estudiante_rut_raw.lower() in ['none', 'nan', '']:
                        errores.append(f'Fila {row_num}: RUT del estudiante requerido')
                        continue
                    
                    if not asignatura_nombre:
                        errores.append(f'Fila {row_num}: Nombre de asignatura requerido')
                        continue
                    
                    if not asignatura_seccion:
                        errores.append(f'Fila {row_num}: Sección de asignatura requerida')
                        continue
                    
                    # Normalizar RUT antes de validar (corregir dígitos verificadores de dos dígitos)
                    import re
                    rut_limpio = re.sub(r'[.\s]', '', estudiante_rut_raw).upper()
                    
                    # Si tiene guión, separar número y dígito verificador
                    if '-' in rut_limpio:
                        partes = rut_limpio.split('-')
                        numero_rut = partes[0]
                        digito_verificador = partes[1] if len(partes) > 1 else ''
                    else:
                        # Si no tiene guión, asumir que el último carácter es el dígito verificador
                        if len(rut_limpio) >= 2:
                            numero_rut = rut_limpio[:-1]
                            digito_verificador = rut_limpio[-1]
                        else:
                            numero_rut = rut_limpio
                            digito_verificador = ''
                    
                    # Corregir dígitos verificadores de dos dígitos (10 -> K, 11 -> 0)
                    if digito_verificador == '10':
                        digito_verificador = 'K'
                    elif digito_verificador == '11':
                        digito_verificador = '0'
                    
                    # Reconstruir RUT normalizado
                    estudiante_rut_normalizado = f"{numero_rut}-{digito_verificador}" if digito_verificador else numero_rut
                    
                    # Validar RUT
                    es_valido, mensaje_error = validar_rut_chileno(estudiante_rut_normalizado)
                    if not es_valido:
                        # Intentar también con el formato original por si acaso
                        es_valido, mensaje_error = validar_rut_chileno(estudiante_rut_raw)
                        if not es_valido:
                            mensaje = mensaje_error if mensaje_error else "RUT inválido"
                            errores.append(f'Fila {row_num}: {mensaje} (RUT recibido: "{estudiante_rut_raw}")')
                            continue
                        else:
                            estudiante_rut_normalizado = estudiante_rut_raw
                    else:
                        estudiante_rut_raw = estudiante_rut_normalizado
                    
                    # Normalizar RUT para búsqueda en BD
                    # Los RUTs pueden estar guardados en diferentes formatos en la BD
                    # Intentar múltiples variaciones
                    # Crear todas las variaciones posibles del RUT
                    rut_variaciones = []
                    
                    # 1. Original
                    rut_variaciones.append(estudiante_rut_raw)
                    
                    # 2. Sin puntos ni espacios, con guión
                    rut_sin_puntos = re.sub(r'[.\s]', '', estudiante_rut_raw).upper()
                    if '-' not in rut_sin_puntos and len(rut_sin_puntos) >= 2:
                        rut_con_guion = rut_sin_puntos[:-1] + '-' + rut_sin_puntos[-1]
                        rut_variaciones.append(rut_con_guion)
                    else:
                        rut_variaciones.append(rut_sin_puntos)
                    
                    # 3. Sin guión
                    rut_sin_guion = rut_sin_puntos.replace('-', '')
                    if rut_sin_guion != rut_sin_puntos:
                        rut_variaciones.append(rut_sin_guion)
                    
                    # 4. Con puntos y guión (formato estándar)
                    if len(rut_sin_guion) >= 2:
                        numero = rut_sin_guion[:-1]
                        dv = rut_sin_guion[-1]
                        # Agregar puntos cada 3 dígitos desde la derecha
                        numero_formateado = ""
                        for i, digito in enumerate(reversed(numero)):
                            if i > 0 and i % 3 == 0:
                                numero_formateado = '.' + numero_formateado
                            numero_formateado = digito + numero_formateado
                        rut_con_puntos = f"{numero_formateado}-{dv}"
                        rut_variaciones.append(rut_con_puntos)
                    
                    # Eliminar duplicados manteniendo el orden
                    rut_variaciones = list(dict.fromkeys(rut_variaciones))
                    
                    # Buscar estudiante con todas las variaciones
                    estudiante = None
                    for rut_var in rut_variaciones:
                        estudiante = Estudiantes.objects.filter(rut=rut_var).first()
                        if estudiante:
                            break
                    if not estudiante:
                        errores.append(f'Fila {row_num}: Estudiante con RUT {estudiante_rut_raw} no encontrado (se intentó con formatos: {", ".join(rut_variaciones[:3])})')
                        continue
                    
                    # Verificar que el estudiante pertenece a una carrera del director
                    if estudiante.carreras not in carreras_del_director:
                        errores.append(f'Fila {row_num}: El estudiante no pertenece a tus carreras')
                        continue
                    
                    # Buscar asignatura por nombre Y sección (dentro de las carreras del director, solo activas)
                    asignatura = Asignaturas.objects.filter(
                        nombre=asignatura_nombre,
                        seccion=asignatura_seccion,
                        carreras__in=carreras_del_director,
                        is_active=True
                    ).first()
                    
                    if not asignatura:
                        # Verificar si existe pero está inactiva
                        asignatura_inactiva = Asignaturas.objects.filter(
                            nombre=asignatura_nombre,
                            seccion=asignatura_seccion,
                            carreras__in=carreras_del_director,
                            is_active=False
                        ).first()
                        if asignatura_inactiva:
                            errores.append(f'Fila {row_num}: La asignatura "{asignatura_nombre}" - "{asignatura_seccion}" está inactiva')
                        else:
                            errores.append(f'Fila {row_num}: No se encontró asignatura con nombre "{asignatura_nombre}" y sección "{asignatura_seccion}"')
                        continue
                    
                    # Crear inscripción si no existe
                    inscripcion, created = AsignaturasEnCurso.objects.get_or_create(
                        estudiantes=estudiante,
                        asignaturas=asignatura,
                        defaults={'estado': True}
                    )
                    
                    if created:
                        creados += 1
                    else:
                        ya_existentes += 1
                        
                except Exception as e:
                    errores.append(f'Fila {row_num}: {str(e)}')
        
        msg = f'Proceso completado: {creados} inscripciones creadas, {ya_existentes} ya existían.'
        if errores:
            msg += f' {len(errores)} errores encontrados.'
            for error in errores[:5]:
                messages.warning(request, error)
        messages.success(request, msg)
        
    except Exception as e:
        messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    return redirect('gestion_carga_masiva_director')


@login_required
def descargar_plantilla_excel(request, tipo):
    """
    Genera y descarga una plantilla Excel vacía según el tipo solicitado.
    Tipos: estudiantes, docentes, asignaturas, inscripciones
    """
    import openpyxl
    from django.http import HttpResponse
    
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_DIRECTOR:
            messages.error(request, 'No tienes permisos para esta acción.')
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if tipo == 'estudiantes':
        ws.title = 'Estudiantes'
        ws.append(['RUT', 'Nombres', 'Apellidos', 'Email', 'Telefono', 'Carrera_ID', 'Semestre_Actual'])
        ws.append(['12345678-9', 'Juan', 'Pérez González', 'juan.perez@email.com', '912345678', '1', '3'])
        filename = 'plantilla_estudiantes.xlsx'
        
    elif tipo == 'docentes':
        ws.title = 'Docentes'
        ws.append(['RUT', 'Nombres', 'Apellidos', 'Email', 'Password'])
        ws.append(['98765432-1', 'María', 'González López', 'maria.gonzalez@email.com', 'MiPassword123'])
        filename = 'plantilla_docentes.xlsx'
        
    elif tipo == 'asignaturas':
        ws.title = 'Asignaturas'
        ws.append(['Nombre', 'Seccion', 'Carrera_ID', 'Docente_RUT', 'Docente_Email'])
        ws.append(['Cálculo I', 'A-001', '1', '98765432-1', ''])
        filename = 'plantilla_asignaturas.xlsx'
        
    elif tipo == 'inscripciones':
        ws.title = 'Inscripciones'
        ws.append(['Estudiante_RUT', 'Asignatura_Nombre', 'Asignatura_Seccion'])
        ws.append(['12345678-9', 'Cálculo I', 'A-001'])
        filename = 'plantilla_inscripciones.xlsx'
        
    else:
        messages.error(request, 'Tipo de plantilla no válido.')
        return redirect('gestion_carga_masiva_director')
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    
    return response


# ----------------------------------------------------
#                VISTA ASESORA TÉCNICA PEDAGÓGICA
# ----------------------------------------------------

@login_required
def dashboard_coordinador_tecnico_pedagogico(request):
    """
    Dashboard principal para el Coordinador Técnico Pedagógico.
    Muestra KPIs de casos pendientes de formulación y estadísticas.
    """
    
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para acceder a este panel.')
            return redirect('home')
    except AttributeError:
        logger.warning(f"Usuario {request.user.email} sin perfil/rol intentó acceder al dashboard de asesora técnica.")
        return redirect('home')

    # 2. --- Configuración de Fechas ---
    now = timezone.localtime(timezone.now())
    today = now.date()
    
    # Cálculo de la semana (Lunes a Domingo)
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_of_week_dt = timezone.make_aware(datetime.combine(start_of_week, datetime.min.time()))
    end_of_week_dt = timezone.make_aware(datetime.combine(end_of_week, datetime.max.time()))
    
    # 3. --- Obtener Datos para KPIs ---
    
    # Base de solicitudes para todas las asesoras técnicas (filtramos por rol)
    todas_las_asesoras_tecnicas = PerfilUsuario.objects.filter(rol__nombre_rol=ROL_COORDINADOR_TECNICO_PEDAGOGICO)
    
    # KPI 1: Casos nuevos (pendiente_formulacion_ajustes esta semana)
    # Casos que cambiaron a pendiente_formulacion_ajustes esta semana
    casos_nuevos_semana = Solicitudes.objects.filter(
        estado='pendiente_formulacion_ajustes',
        updated_at__range=(start_of_week_dt, end_of_week_dt)
    ).select_related('estudiantes', 'estudiantes__carreras').count()
    
    # KPI 2: Casos pendientes de formulación de ajustes en total
    casos_pendientes_formulacion = Solicitudes.objects.filter(
        estado='pendiente_formulacion_ajustes'
    ).select_related('estudiantes', 'estudiantes__carreras')
    kpi_casos_pendientes_total = casos_pendientes_formulacion.count()
    
    # KPI 3: Casos devueltos desde Asesora Pedagógica
    casos_devueltos = Solicitudes.objects.filter(
        estado='pendiente_formulacion_ajustes',
        ajusteasignado__isnull=False
    ).distinct().count()
    
    # KPI 4: Total de ajustes formulados por este coordinador
    casos_asignados = Solicitudes.objects.filter(
        coordinador_tecnico_pedagogico_asignado=perfil
    )
    total_ajustes_formulados = AjusteAsignado.objects.filter(
        solicitudes__in=casos_asignados
    ).count()
    
    # KPI 5: Ajustes aprobados
    ajustes_aprobados = AjusteAsignado.objects.filter(
        solicitudes__in=casos_asignados,
        estado_aprobacion='aprobado'
    ).count()
    
    # KPI 6: Casos enviados a Asesor Pedagógico esta semana
    casos_enviados_semana = Solicitudes.objects.filter(
        estado='pendiente_preaprobacion',
        coordinador_tecnico_pedagogico_asignado=perfil,
        updated_at__range=(start_of_week_dt, end_of_week_dt)
    ).count()
    
    # 4. --- Obtener Lista de Casos Pendientes de Formulación ---
    casos_pendientes_list = casos_pendientes_formulacion.order_by('-updated_at')[:10]
    
    # 5. --- Preparar Contexto ---
    context = {
        'nombre_usuario': request.user.first_name,
        'kpis': {
            'casos_nuevos_semana': casos_nuevos_semana,
            'casos_pendientes_total': kpi_casos_pendientes_total,
            'casos_devueltos': casos_devueltos,
            'total_ajustes_formulados': total_ajustes_formulados,
            'ajustes_aprobados': ajustes_aprobados,
            'casos_enviados_semana': casos_enviados_semana,
        },
        'casos_pendientes_list': casos_pendientes_list,
    }
    
    # 6. --- Renderizar Template ---
    return render(request, 'SIAPE/dashboard_coordinador_tecnico_pedagogico.html', context)


# ----------------------------------------------------
#           GESTIÓN DE CATEGORÍAS DE AJUSTES
#           (Coordinador Técnico Pedagógico)
# ----------------------------------------------------

@login_required
def gestion_categorias_ajustes(request):
    """
    Vista para que el Coordinador Técnico Pedagógico gestione las categorías de ajustes.
    Permite crear, editar y eliminar categorías.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
    
    # 2. --- Obtener todas las categorías con conteo de uso ---
    categorias = CategoriasAjustes.objects.annotate(
        total_ajustes=Count('ajusterazonable')
    ).order_by('nombre_categoria')
    
    # 3. --- Procesar formularios ---
    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        if accion == 'crear':
            nombre = request.POST.get('nombre', '').strip()
            if nombre:
                # Verificar si ya existe
                if CategoriasAjustes.objects.filter(nombre_categoria__iexact=nombre).exists():
                    messages.error(request, f'La categoría "{nombre}" ya existe.')
                else:
                    CategoriasAjustes.objects.create(nombre_categoria=nombre.capitalize())
                    messages.success(request, f'Categoría "{nombre}" creada exitosamente.')
            else:
                messages.error(request, 'El nombre de la categoría es requerido.')
        
        elif accion == 'editar':
            categoria_id = request.POST.get('categoria_id')
            nuevo_nombre = request.POST.get('nuevo_nombre', '').strip()
            if categoria_id and nuevo_nombre:
                try:
                    categoria = CategoriasAjustes.objects.get(id=categoria_id)
                    # Verificar si el nuevo nombre ya existe (excepto la misma categoría)
                    if CategoriasAjustes.objects.filter(nombre_categoria__iexact=nuevo_nombre).exclude(id=categoria_id).exists():
                        messages.error(request, f'La categoría "{nuevo_nombre}" ya existe.')
                    else:
                        categoria.nombre_categoria = nuevo_nombre.capitalize()
                        categoria.save()
                        messages.success(request, 'Categoría actualizada exitosamente.')
                except CategoriasAjustes.DoesNotExist:
                    messages.error(request, 'Categoría no encontrada.')
            else:
                messages.error(request, 'Datos incompletos.')
        
        elif accion == 'eliminar':
            categoria_id = request.POST.get('categoria_id')
            if categoria_id:
                try:
                    categoria = CategoriasAjustes.objects.get(id=categoria_id)
                    # Verificar si tiene ajustes asociados
                    total_ajustes = AjusteRazonable.objects.filter(categorias_ajustes=categoria).count()
                    if total_ajustes > 0:
                        messages.error(request, f'No se puede eliminar la categoría "{categoria.nombre_categoria}" porque tiene {total_ajustes} ajuste(s) asociado(s).')
                    else:
                        nombre_categoria = categoria.nombre_categoria
                        categoria.delete()
                        messages.success(request, f'Categoría "{nombre_categoria}" eliminada exitosamente.')
                except CategoriasAjustes.DoesNotExist:
                    messages.error(request, 'Categoría no encontrada.')
        
        return redirect('gestion_categorias_ajustes')
    
    # 4. --- Preparar Contexto ---
    context = {
        'categorias': categorias,
    }
    
    return render(request, 'SIAPE/gestion_categorias_ajustes.html', context)


@login_required
def estadisticas_ajustes_coordinador_tecnico(request):
    """
    Vista de estadísticas de ajustes formulados por el Coordinador Técnico Pedagógico.
    """
    # 1. --- Verificación de Permisos ---
    try:
        perfil = request.user.perfil
        if perfil.rol.nombre_rol != ROL_COORDINADOR_TECNICO_PEDAGOGICO:
            messages.error(request, 'No tienes permisos para acceder a esta página.')
            return redirect('home')
    except AttributeError:
        if not request.user.is_superuser:
            return redirect('home')
    
    # 2. --- Obtener ajustes formulados por este coordinador ---
    # Aproximación: ajustes en casos asignados a este coordinador
    casos_asignados = Solicitudes.objects.filter(
        coordinador_tecnico_pedagogico_asignado=perfil
    )
    
    ajustes_formulados = AjusteAsignado.objects.filter(
        solicitudes__in=casos_asignados
    ).select_related(
        'ajuste_razonable__categorias_ajustes',
        'solicitudes__estudiantes',
        'solicitudes__estudiantes__carreras'
    )
    
    # 3. --- Estadísticas por categoría ---
    estadisticas_categoria = {}
    for ajuste in ajustes_formulados:
        categoria = ajuste.ajuste_razonable.categorias_ajustes
        if categoria:
            if categoria.nombre_categoria not in estadisticas_categoria:
                estadisticas_categoria[categoria.nombre_categoria] = {
                    'total': 0,
                    'aprobados': 0,
                    'rechazados': 0,
                    'pendientes': 0
                }
            estadisticas_categoria[categoria.nombre_categoria]['total'] += 1
            if ajuste.estado_aprobacion == 'aprobado':
                estadisticas_categoria[categoria.nombre_categoria]['aprobados'] += 1
            elif ajuste.estado_aprobacion == 'rechazado':
                estadisticas_categoria[categoria.nombre_categoria]['rechazados'] += 1
            else:
                estadisticas_categoria[categoria.nombre_categoria]['pendientes'] += 1
    
    # 4. --- Estadísticas generales ---
    total_ajustes = ajustes_formulados.count()
    ajustes_aprobados = ajustes_formulados.filter(estado_aprobacion='aprobado').count()
    ajustes_rechazados = ajustes_formulados.filter(estado_aprobacion='rechazado').count()
    ajustes_pendientes = ajustes_formulados.filter(estado_aprobacion='pendiente').count()
    
    # 5. --- Ajustes recientes ---
    ajustes_recientes = ajustes_formulados.order_by('-created_at')[:10]
    
    # 6. --- Preparar Contexto ---
    context = {
        'total_ajustes': total_ajustes,
        'ajustes_aprobados': ajustes_aprobados,
        'ajustes_rechazados': ajustes_rechazados,
        'ajustes_pendientes': ajustes_pendientes,
        'estadisticas_categoria': estadisticas_categoria,
        'ajustes_recientes': ajustes_recientes,
    }
    
    return render(request, 'SIAPE/estadisticas_ajustes_coordinador_tecnico.html', context)


# ----------- Vistas para Docente ------------  


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def obtener_datos_caso_docente(request, solicitud_id):
    """
    Endpoint API para obtener los datos del caso para mostrar en el modal del docente.
    """
    try:
        if request.user.perfil.rol.nombre_rol != ROL_DOCENTE:
            return Response({'error': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
    except AttributeError:
        return Response({'error': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
    
    perfil_docente = request.user.perfil
    mis_asignaturas = Asignaturas.objects.filter(docente=perfil_docente)
    
    # Obtener la solicitud
    try:
        solicitud = Solicitudes.objects.get(id=solicitud_id, estado='aprobado')
    except Solicitudes.DoesNotExist:
        return Response({'error': 'Caso no encontrado'}, status=status.HTTP_404_NOT_FOUND)
    
    # Verificar que el estudiante está en las clases del docente
    estudiante_en_clases = AsignaturasEnCurso.objects.filter(
        estudiantes=solicitud.estudiantes,
        asignaturas__in=mis_asignaturas
    ).exists()
    
    if not estudiante_en_clases:
        return Response({'error': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
    
    # Obtener solo los ajustes aprobados para el docente
    ajustes_asignados = AjusteAsignado.objects.filter(
        solicitudes=solicitud,
        estado_aprobacion='aprobado'
    ).select_related('ajuste_razonable__categorias_ajustes', 'docente_comentador__usuario')
    
    ajustes_data = []
    for ajuste in ajustes_asignados:
        ajustes_data.append({
            'id': ajuste.id,
            'categoria': ajuste.ajuste_razonable.categorias_ajustes.nombre_categoria,
            'descripcion': ajuste.ajuste_razonable.descripcion,
            'estado_aprobacion': ajuste.estado_aprobacion,
            'comentarios_docente': ajuste.comentarios_docente or '',
            'fecha_comentario_docente': ajuste.fecha_comentario_docente.strftime('%d/%m/%Y %H:%M') if ajuste.fecha_comentario_docente else None,
            'docente_comentador': ajuste.docente_comentador.usuario.get_full_name() if ajuste.docente_comentador else None
        })
    
    # Preparar respuesta
    data = {
        'estudiante': {
            'nombres': solicitud.estudiantes.nombres,
            'apellidos': solicitud.estudiantes.apellidos,
            'rut': solicitud.estudiantes.rut,
            'foto': None  # Si hay campo de foto, agregarlo aquí
        },
        'descripcion_caso': solicitud.descripcion or 'No hay descripción disponible.',
        'ajustes': ajustes_data,
        'tiene_ajustes': len(ajustes_data) > 0
    }
    
    return Response(data, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def agregar_comentario_ajuste_docente(request, ajuste_asignado_id):
    """
    Endpoint API para que el docente agregue o actualice un comentario sobre un ajuste.
    """
    try:
        if request.user.perfil.rol.nombre_rol != ROL_DOCENTE:
            return Response({'error': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
    except AttributeError:
        return Response({'error': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
    
    perfil_docente = request.user.perfil
    
    # Obtener el ajuste asignado
    try:
        ajuste = AjusteAsignado.objects.get(
            id=ajuste_asignado_id,
            estado_aprobacion='aprobado'
        )
    except AjusteAsignado.DoesNotExist:
        return Response({'error': 'Ajuste no encontrado'}, status=status.HTTP_404_NOT_FOUND)
    
    # Verificar que el docente tiene acceso a este ajuste (el estudiante debe estar en sus clases)
    mis_asignaturas = Asignaturas.objects.filter(docente=perfil_docente)
    estudiante_en_clases = AsignaturasEnCurso.objects.filter(
        estudiantes=ajuste.solicitudes.estudiantes,
        asignaturas__in=mis_asignaturas
    ).exists()
    
    if not estudiante_en_clases:
        return Response({'error': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
    
    # Obtener el comentario del request
    comentario = request.data.get('comentario', '').strip()
    
    if not comentario:
        return Response({'error': 'El comentario no puede estar vacío'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Guardar el comentario
    ajuste.comentarios_docente = comentario
    ajuste.docente_comentador = perfil_docente
    ajuste.fecha_comentario_docente = timezone.now()
    ajuste.save()
    
    return Response({
        'success': True,
        'message': 'Comentario guardado exitosamente',
        'comentario': ajuste.comentarios_docente,
        'fecha_comentario': ajuste.fecha_comentario_docente.strftime('%d/%m/%Y %H:%M'),
        'docente_comentador': perfil_docente.usuario.get_full_name()
    }, status=status.HTTP_200_OK)

@login_required
def dashboard_docente(request):
    """
    Dashboard para los docentes.
    Muestra sus asignaturas y alumnos asociados con ajustes aprobados.
    """
    try:
        if request.user.perfil.rol.nombre_rol != ROL_DOCENTE:
            return redirect('home')
    except AttributeError:
        return redirect('home')

    perfil_docente = request.user.perfil

    # 1. Obtener las asignaturas del docente y contar el total de alumnos
    mis_asignaturas = Asignaturas.objects.filter(
        docente=perfil_docente
    ).annotate(
        total_alumnos=Count('asignaturasencurso', distinct=True) # Total de alumnos en la clase
    )

    # 2. Obtener IDs de estudiantes únicos en todas las clases del docente
    estudiantes_ids = AsignaturasEnCurso.objects.filter(
        asignaturas__in=mis_asignaturas
    ).values_list('estudiantes_id', flat=True).distinct()
    
    # 3. Obtener solo solicitudes de estudiantes que tienen AJUSTES APROBADOS
    #    Filtrar estudiantes que cursan las asignaturas del docente Y tienen ajustes aprobados
    ajustes_aprobados = AjusteAsignado.objects.filter(
        solicitudes__estudiantes_id__in=estudiantes_ids,
        estado_aprobacion='aprobado'
    ).select_related(
        'solicitudes',
        'solicitudes__estudiantes'
    ).prefetch_related(
        'ajuste_razonable__categorias_ajustes',
        'solicitudes__asignaturas_solicitadas'
    ).distinct()
    
    # Obtener IDs únicos de solicitudes con ajustes aprobados
    solicitudes_ids_con_ajustes_aprobados = ajustes_aprobados.values_list('solicitudes_id', flat=True).distinct()
    
    # Obtener las solicitudes completas
    solicitudes_aprobadas = Solicitudes.objects.filter(
        id__in=solicitudes_ids_con_ajustes_aprobados,
        estado='aprobado'
    ).select_related(
        'estudiantes'
    ).prefetch_related(
        'ajusteasignado_set__ajuste_razonable__categorias_ajustes',
        'asignaturas_solicitadas' 
    ).distinct()

    # 4. Crear un mapa de { asignatura_id -> [lista de detalles de caso] }
    # Solo mostrar estudiantes con ajustes aprobados
    mapa_casos_por_asignatura = {}
    total_estudiantes_con_caso = set() 
    mis_asignaturas_ids = set(mis_asignaturas.values_list('id', flat=True))

    for sol in solicitudes_aprobadas:
        # Obtener ajustes aprobados (ya filtrados arriba, solo aprobados)
        ajustes_aprobados = sol.ajusteasignado_set.filter(estado_aprobacion='aprobado')
        
        # Solo agregar el estudiante si tiene ajustes aprobados
        if ajustes_aprobados.exists():
            detalle_para_tabla = {
                'estudiante': sol.estudiantes,
                'ajustes': ajustes_aprobados,
                'solicitud_id': sol.id  # ID de la solicitud original
            }
            
            total_estudiantes_con_caso.add(sol.estudiantes.id)
            
            # Asignar este detalle a TODAS las asignaturas del docente donde el estudiante está inscrito
            # (no solo las asignaturas de la solicitud)
            for asig in mis_asignaturas:
                # Verificar si el estudiante está inscrito en esta asignatura
                estudiante_en_asignatura = AsignaturasEnCurso.objects.filter(
                    estudiantes=sol.estudiantes,
                    asignaturas=asig
                ).exists()
                
                if estudiante_en_asignatura:
                    if asig.id not in mapa_casos_por_asignatura:
                        mapa_casos_por_asignatura[asig.id] = []
                    
                    # Evitar duplicados de estudiantes por asignatura
                    existe = False
                    for item in mapa_casos_por_asignatura[asig.id]:
                        if item['estudiante'].id == sol.estudiantes.id:
                            existe = True
                            break
                    if not existe:
                        mapa_casos_por_asignatura[asig.id].append(detalle_para_tabla)

    # 4. Construir el contexto final 'casos_por_asignatura' que espera la plantilla
    casos_por_asignatura = []
    for asig in mis_asignaturas:
        detalles = mapa_casos_por_asignatura.get(asig.id, [])
        casos_por_asignatura.append({
            'asignatura': asig,
            'total_alumnos': asig.total_alumnos,
            'total_ajustes_aprobados': len(detalles), # N° de estudiantes con caso aprobado
            'ajustes_aprobados_detalle': detalles
        })

    context = {
        'casos_por_asignatura': casos_por_asignatura,
        'asignaturas_docente': mis_asignaturas, 
        'total_estudiantes_con_ajuste': len(total_estudiantes_con_caso)  # Cambiado para reflejar casos aprobados
    }
    
    return render(request, 'SIAPE/dashboard_docente.html', context)

@login_required
def mis_asignaturas_docente(request):
    """
    Muestra al docente la lista de asignaturas que imparte.
    """
    try:
        # Verificar que el usuario sea docente
        if request.user.perfil.rol.nombre_rol != ROL_DOCENTE:
            return redirect('home')
    except AttributeError:
        return redirect('home')
    
    perfil_docente = request.user.perfil
    
    # Obtener asignaturas del docente
    asignaturas_docente = Asignaturas.objects.filter(
        docente=perfil_docente
    ).order_by('nombre')
    
    # Para cada asignatura, contar estudiantes con ajustes aprobados
    asignaturas_con_contador = []
    for asignatura in asignaturas_docente:
        # Obtener estudiantes inscritos en esta asignatura
        estudiantes_ids = AsignaturasEnCurso.objects.filter(
            asignaturas=asignatura
        ).values_list('estudiantes_id', flat=True)
        
        # Contar estudiantes con ajustes aprobados
        estudiantes_con_ajustes_aprobados = AjusteAsignado.objects.filter(
            solicitudes__estudiantes_id__in=estudiantes_ids,
            estado_aprobacion='aprobado'
        ).values('solicitudes__estudiantes_id').distinct().count()
        
        # Agregar el contador a la asignatura
        asignatura.total_estudiantes = estudiantes_con_ajustes_aprobados
        asignaturas_con_contador.append(asignatura)

    context = {
        'asignaturas': asignaturas_con_contador
    }
    
    return render(request, 'SIAPE/mis_asignaturas_docente.html', context)
    



@login_required
def mis_alumnos_docente(request):
    """
    Muestra una lista de TODOS los alumnos únicos que el docente
    tiene en todas sus asignaturas.
    """
    try:
        # Verificar que el usuario sea docente
        if request.user.perfil.rol.nombre_rol != ROL_DOCENTE:
            return redirect('home')
    except AttributeError:

        return redirect('home')

    perfil_docente = request.user.perfil
    mis_asignaturas = Asignaturas.objects.filter(docente=perfil_docente)
    mis_asignaturas_ids = list(mis_asignaturas.values_list('id', flat=True))

    # 1. Obtener IDs de estudiantes únicos en todas las clases del docente
    estudiantes_ids = AsignaturasEnCurso.objects.filter(
        asignaturas__in=mis_asignaturas
    ).values_list('estudiantes_id', flat=True).distinct()
    
    # 2. Obtener los objetos de esos estudiantes
    mis_estudiantes = Estudiantes.objects.filter(
        id__in=estudiantes_ids
    ).select_related('carreras').order_by('apellidos', 'nombres')

    # 3. Obtener solo estudiantes que tienen AJUSTES APROBADOS
    #    Filtrar estudiantes que cursan las asignaturas del docente Y tienen ajustes aprobados
    estudiantes_con_ajustes_aprobados_ids = set()
    solicitudes_por_estudiante = {}
    
    # Obtener ajustes aprobados de estudiantes que están en las clases del docente
    ajustes_aprobados = AjusteAsignado.objects.filter(
        solicitudes__estudiantes_id__in=estudiantes_ids,
        estado_aprobacion='aprobado'
    ).select_related(
        'solicitudes',
        'solicitudes__estudiantes'
    ).distinct()
    
    # Para cada ajuste aprobado, agregar el estudiante a la lista
    for ajuste in ajustes_aprobados:
        estudiante_id = ajuste.solicitudes.estudiantes_id
        estudiantes_con_ajustes_aprobados_ids.add(estudiante_id)
        # Guardar la primera solicitud encontrada para cada estudiante
        if estudiante_id not in solicitudes_por_estudiante:
            solicitudes_por_estudiante[estudiante_id] = ajuste.solicitudes_id
    
    # 4. Filtrar estudiantes: solo aquellos con ajustes aprobados
    estudiantes_filtrados = mis_estudiantes.filter(
        id__in=estudiantes_con_ajustes_aprobados_ids
    )
    
    # 5. Preparar la lista final para la plantilla
    lista_alumnos_final = []
    for est in estudiantes_filtrados:
        solicitud_id = solicitudes_por_estudiante.get(est.id)
        
        lista_alumnos_final.append({
            'estudiante': est,
            'tiene_caso_aprobado': True,  # Todos tienen ajustes aprobados
            'solicitud_id': solicitud_id
        })

    context = {
        'lista_alumnos': lista_alumnos_final,
        'total_alumnos': len(lista_alumnos_final),
        # Debug temporal
        'debug_estudiantes_con_caso': len(estudiantes_con_ajustes_aprobados_ids),
        'debug_total_solicitudes': ajustes_aprobados.count()
    }

    return render(request, 'SIAPE/mis_alumnos_docente.html', context)



@login_required
def detalle_asignatura_docente(request, asignatura_id):
    """
    Muestra el listado de alumnos de una asignatura específica.
    """
    try:
        # Verificar que el usuario sea docente
        if request.user.perfil.rol.nombre_rol != ROL_DOCENTE:
            return redirect('home')
    except AttributeError:
        return redirect('home')

    # 1. Obtener la asignatura (con chequeo de seguridad de que le pertenece)
    asignatura = get_object_or_404(
        Asignaturas, 
        id=asignatura_id, 
        docente=request.user.perfil
    )

    # 2. Obtener todos los estudiantes inscritos en esa asignatura
    estudiantes_en_curso = AsignaturasEnCurso.objects.filter(
        asignaturas=asignatura
    ).select_related('estudiantes')

    # 3. Obtener solo estudiantes que tienen AJUSTES APROBADOS
    #    Filtrar estudiantes que cursan esta asignatura Y tienen ajustes aprobados
    estudiantes_ids_en_asignatura = list(estudiantes_en_curso.values_list('estudiantes_id', flat=True))
    
    # Obtener ajustes aprobados de estudiantes que están en esta asignatura
    ajustes_aprobados = AjusteAsignado.objects.filter(
        solicitudes__estudiantes_id__in=estudiantes_ids_en_asignatura,
        estado_aprobacion='aprobado'
    ).select_related(
        'solicitudes',
        'solicitudes__estudiantes'
    ).distinct()
    
    estudiantes_con_ajustes_aprobados_ids = set()
    solicitudes_por_estudiante = {}
    for ajuste in ajustes_aprobados:
        estudiante_id = ajuste.solicitudes.estudiantes_id
        estudiantes_con_ajustes_aprobados_ids.add(estudiante_id)
        if estudiante_id not in solicitudes_por_estudiante:
            solicitudes_por_estudiante[estudiante_id] = ajuste.solicitudes_id

    
    # 4. Filtrar estudiantes: solo aquellos con ajustes aprobados
    estudiantes_filtrados = estudiantes_en_curso.filter(
        estudiantes_id__in=estudiantes_con_ajustes_aprobados_ids
    )
    
    # 5. Preparar la lista final de alumnos para la plantilla
    lista_alumnos = []
    for ec in estudiantes_filtrados:
        solicitud_id = solicitudes_por_estudiante.get(ec.estudiantes.id)
        lista_alumnos.append({
            'estudiante': ec.estudiantes,
            'tiene_caso_aprobado': True,  # Todos tienen ajustes aprobados
            'solicitud_id': solicitud_id
        })

    context = {
        'asignatura': asignatura,
        'lista_alumnos': lista_alumnos,
        'total_alumnos': len(lista_alumnos)
    }
    

    return render(request, 'SIAPE/detalle_asignatura_docente.html', context)
    


@login_required
def detalle_ajuste_docente(request, estudiante_id):
    """
    Muesta el detalle de los ajustes aporbados de un estudiante
    especifico, pero solo los relevantes a las asignaturas
    que imparte el docente logueado
    """
    try:
        if request.user.perfil.rol.nombre_rol != ROL_DOCENTE:
            return redirect('home')

    except (AttributeError, PerfilUsuario.DoesNotExist):
        return redirect('home')
    
    perfil_docente = request.user.perfil
    estudiante = get_object_or_404(Estudiantes, id=estudiante_id)
    
    #1. Obtener las asignaturas que imparte este docente 
    mis_asignaturas = Asignaturas.objects.filter(docente=perfil_docente)

    # 2. Buscar solicitudes aprobadas de ese estudiante
    # y que apliquen a cualquiera de las asignaturas del docente
    solicitudes_relevantes = Solicitudes.objects.filter(
        estudiantes=estudiante,
        estado='aprobado',
        asignaturas_solicitadas__in=mis_asignaturas
    ).prefetch_related(
        'ajusteasignado_set__ajuste_razonable'
    ).distinct()

    # 3. Juntar todos los ajustes aprobados en una sola lista 
    # (un alumno puede tener varios ajustes de varias solicitudes)
    ajustes = []
    ajustes_ids = set() # para evitar duplicados
    for solicitud in solicitudes_relevantes:
        # Filtrar solo los ajustes que están aprobados
        for ajuste in solicitud.ajusteasignado_set.filter(estado_aprobacion='aprobado'):
            if ajuste.id not in ajustes_ids:
                ajustes.append(ajuste)
                ajustes_ids.add(ajuste.id)

    context = {
        'estudiante': estudiante,
        'ajustes': ajustes
    }

    return render(request, 'SIAPE/detalle_ajuste_docente.html', context)



# ----------- Vistas de los modelos (API) ------------
# ViewSets con controles de acceso mejorados
class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminUser]  # Solo administradores pueden gestionar usuarios
    
    def get_queryset(self):
        """
        Los usuarios solo pueden ver su propio perfil, excepto administradores.
        """
        queryset = Usuario.objects.all()
        if self.request.user.is_superuser or self.request.user.is_staff:
            return queryset
        # Usuario normal solo ve su propio perfil
        return queryset.filter(id=self.request.user.id)
class RolesViewSet(viewsets.ModelViewSet):
    queryset = Roles.objects.all()
    serializer_class = RolesSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminOrReadOnly]  # Lectura para autenticados, escritura solo admin
class AreasViewSet(viewsets.ModelViewSet):
    queryset = Areas.objects.all()
    serializer_class = AreasSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminOrReadOnly]  # Lectura para autenticados, escritura solo admin
class CategoriasAjustesViewSet(viewsets.ModelViewSet):
    queryset = CategoriasAjustes.objects.all()
    serializer_class = CategoriasAjustesSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminOrReadOnly]  # Lectura para autenticados, escritura solo admin
class CarrerasViewSet(viewsets.ModelViewSet):
    queryset = Carreras.objects.all()
    serializer_class = CarrerasSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminOrReadOnly]  # Lectura para autenticados, escritura solo admin
class EstudiantesViewSet(viewsets.ModelViewSet):
    queryset = Estudiantes.objects.all()
    serializer_class = EstudiantesSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Filtrar estudiantes según el rol del usuario:
        - Admin: ve todos
        - Director: ve estudiantes de sus carreras
        - Docente: ve estudiantes de sus asignaturas
        - Otros: acceso limitado
        """
        queryset = Estudiantes.objects.all()
        user = self.request.user
        
        if user.is_superuser or user.is_staff:
            return queryset
        
        try:
            perfil = user.perfil
            rol = perfil.rol.nombre_rol if perfil.rol else None
            
            if rol == 'Director de Carrera':
                # Director ve estudiantes de sus carreras
                carreras_dirigidas = Carreras.objects.filter(director=perfil)
                return queryset.filter(carreras__in=carreras_dirigidas).distinct()
            
            elif rol == 'Docente':
                # Docente ve estudiantes de sus asignaturas
                asignaturas_docente = Asignaturas.objects.filter(docente=perfil)
                return queryset.filter(
                    asignaturasencurso__asignaturas__in=asignaturas_docente
                ).distinct()
            
            # Otros roles (Coordinadora, Asesores) pueden ver todos los estudiantes
            # pero solo en lectura
            return queryset
        except AttributeError:
            # Si no tiene perfil, no puede ver nada
            return Estudiantes.objects.none()
class SolicitudesViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin, mixins.DestroyModelMixin, viewsets.GenericViewSet):
    queryset = Solicitudes.objects.all().order_by('-created_at')
    serializer_class = SolicitudesSerializer
    permission_classes = [IsAsesorPedagogico | IsAdminUser | IsCoordinadora | IsAsesorTecnico | IsDirectorCarrera]
    
    def get_queryset(self):
        """
        Filtrar solicitudes según el rol del usuario.
        """
        queryset = Solicitudes.objects.all().order_by('-created_at')
        user = self.request.user
        
        if user.is_superuser or user.is_staff:
            return queryset
        
        try:
            perfil = user.perfil
            rol = perfil.rol.nombre_rol if perfil.rol else None
            
            if rol == 'Encargado de Inclusión':
                # Ve solicitudes asignadas a ella
                return queryset.filter(coordinadora_asignada=perfil)
            
            elif rol == 'Coordinador Técnico Pedagógico':
                # Ve solicitudes asignadas a él
                return queryset.filter(coordinador_tecnico_pedagogico_asignado=perfil)
            
            elif rol == 'Asesor Pedagógico':
                # Ve solicitudes asignadas a él
                return queryset.filter(asesor_pedagogico_asignado=perfil)
            
            elif rol == 'Director de Carrera':
                # Ve solicitudes de estudiantes de sus carreras
                carreras_dirigidas = Carreras.objects.filter(director=perfil)
                return queryset.filter(estudiantes__carreras__in=carreras_dirigidas).distinct()
            
            # Si no tiene un rol válido, no puede ver nada
            return Solicitudes.objects.none()
        except AttributeError:
            return Solicitudes.objects.none()
class EvidenciasViewSet(viewsets.ModelViewSet):
    queryset = Evidencias.objects.all()
    serializer_class = EvidenciasSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Filtrar evidencias según el rol del usuario.
        Solo pueden ver evidencias de solicitudes a las que tienen acceso.
        """
        queryset = Evidencias.objects.all()
        user = self.request.user
        
        if user.is_superuser or user.is_staff:
            return queryset
        
        try:
            perfil = user.perfil
            rol = perfil.rol.nombre_rol if perfil.rol else None
            
            # Obtener solicitudes accesibles según el rol
            solicitudes_accesibles = Solicitudes.objects.none()
            
            if rol == 'Encargado de Inclusión':
                solicitudes_accesibles = Solicitudes.objects.filter(coordinadora_asignada=perfil)
            elif rol == 'Coordinador Técnico Pedagógico':
                solicitudes_accesibles = Solicitudes.objects.filter(coordinador_tecnico_pedagogico_asignado=perfil)
            elif rol == 'Asesor Pedagógico':
                solicitudes_accesibles = Solicitudes.objects.filter(asesor_pedagogico_asignado=perfil)
            elif rol == 'Director de Carrera':
                carreras_dirigidas = Carreras.objects.filter(director=perfil)
                solicitudes_accesibles = Solicitudes.objects.filter(estudiantes__carreras__in=carreras_dirigidas).distinct()
            elif rol == 'Docente':
                asignaturas_docente = Asignaturas.objects.filter(docente=perfil)
                estudiantes_docente = Estudiantes.objects.filter(
                    asignaturasencurso__asignaturas__in=asignaturas_docente
                ).distinct()
                solicitudes_accesibles = Solicitudes.objects.filter(estudiantes__in=estudiantes_docente)
            
            return queryset.filter(solicitudes__in=solicitudes_accesibles)
        except AttributeError:
            return Evidencias.objects.none()
class AsignaturasViewSet(viewsets.ModelViewSet):
    queryset = Asignaturas.objects.all()
    serializer_class = AsignaturasSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Filtrar asignaturas según el rol del usuario.
        """
        queryset = Asignaturas.objects.all()
        user = self.request.user
        
        if user.is_superuser or user.is_staff:
            return queryset
        
        try:
            perfil = user.perfil
            rol = perfil.rol.nombre_rol if perfil.rol else None
            
            if rol == 'Docente':
                # Docente solo ve sus propias asignaturas
                return queryset.filter(docente=perfil)
            
            elif rol == 'Director de Carrera':
                # Director ve asignaturas de sus carreras
                carreras_dirigidas = Carreras.objects.filter(director=perfil)
                return queryset.filter(carreras__in=carreras_dirigidas).distinct()
            
            # Otros roles pueden ver todas las asignaturas (solo lectura)
            return queryset
        except AttributeError:
            return Asignaturas.objects.none()
class AsignaturasEnCursoViewSet(viewsets.ModelViewSet):
    queryset = AsignaturasEnCurso.objects.all()
    serializer_class = AsignaturasEnCursoSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Filtrar asignaturas en curso según el rol del usuario.
        """
        queryset = AsignaturasEnCurso.objects.all()
        user = self.request.user
        
        if user.is_superuser or user.is_staff:
            return queryset
        
        try:
            perfil = user.perfil
            rol = perfil.rol.nombre_rol if perfil.rol else None
            
            if rol == 'Docente':
                # Docente ve asignaturas en curso de sus asignaturas
                asignaturas_docente = Asignaturas.objects.filter(docente=perfil)
                return queryset.filter(asignaturas__in=asignaturas_docente)
            
            elif rol == 'Director de Carrera':
                # Director ve asignaturas en curso de estudiantes de sus carreras
                carreras_dirigidas = Carreras.objects.filter(director=perfil)
                return queryset.filter(estudiantes__carreras__in=carreras_dirigidas).distinct()
            
            # Otros roles pueden ver todas (solo lectura)
            return queryset
        except AttributeError:
            return AsignaturasEnCurso.objects.none()
class EntrevistasViewSet(viewsets.ModelViewSet):
    queryset = Entrevistas.objects.all()
    serializer_class = EntrevistasSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Filtrar entrevistas según el rol del usuario.
        """
        queryset = Entrevistas.objects.all()
        user = self.request.user
        
        if user.is_superuser or user.is_staff:
            return queryset
        
        try:
            perfil = user.perfil
            rol = perfil.rol.nombre_rol if perfil.rol else None
            
            if rol == 'Encargado de Inclusión':
                # Ve entrevistas donde es coordinadora
                return queryset.filter(coordinadora=perfil)
            
            # Otros roles ven entrevistas de solicitudes a las que tienen acceso
            solicitudes_accesibles = Solicitudes.objects.none()
            
            if rol == 'Coordinador Técnico Pedagógico':
                solicitudes_accesibles = Solicitudes.objects.filter(coordinador_tecnico_pedagogico_asignado=perfil)
            elif rol == 'Asesor Pedagógico':
                solicitudes_accesibles = Solicitudes.objects.filter(asesor_pedagogico_asignado=perfil)
            elif rol == 'Director de Carrera':
                carreras_dirigidas = Carreras.objects.filter(director=perfil)
                solicitudes_accesibles = Solicitudes.objects.filter(estudiantes__carreras__in=carreras_dirigidas).distinct()
            
            return queryset.filter(solicitudes__in=solicitudes_accesibles)
        except AttributeError:
            return Entrevistas.objects.none()
class AjusteRazonableViewSet(viewsets.ModelViewSet):
    queryset = AjusteRazonable.objects.all()
    serializer_class = AjusteRazonableSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminOrReadOnly]  # Lectura para autenticados, escritura solo admin
class AjusteAsignadoViewSet(viewsets.ModelViewSet):
    queryset = AjusteAsignado.objects.all()
    serializer_class = AjusteAsignadoSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Filtrar ajustes asignados según el rol del usuario.
        """
        queryset = AjusteAsignado.objects.all()
        user = self.request.user
        
        if user.is_superuser or user.is_staff:
            return queryset
        
        try:
            perfil = user.perfil
            rol = perfil.rol.nombre_rol if perfil.rol else None
            
            # Obtener solicitudes accesibles según el rol
            solicitudes_accesibles = Solicitudes.objects.none()
            
            if rol == 'Encargado de Inclusión':
                solicitudes_accesibles = Solicitudes.objects.filter(coordinadora_asignada=perfil)
            elif rol == 'Coordinador Técnico Pedagógico':
                solicitudes_accesibles = Solicitudes.objects.filter(coordinador_tecnico_pedagogico_asignado=perfil)
            elif rol == 'Asesor Pedagógico':
                solicitudes_accesibles = Solicitudes.objects.filter(asesor_pedagogico_asignado=perfil)
            elif rol == 'Director de Carrera':
                carreras_dirigidas = Carreras.objects.filter(director=perfil)
                solicitudes_accesibles = Solicitudes.objects.filter(estudiantes__carreras__in=carreras_dirigidas).distinct()
            
            return queryset.filter(solicitudes__in=solicitudes_accesibles)
        except AttributeError:
            return AjusteAsignado.objects.none()
class PerfilUsuarioViewSet(viewsets.ModelViewSet):
    queryset = PerfilUsuario.objects.all()
    serializer_class = PerfilUsuarioSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Los usuarios solo pueden ver su propio perfil, excepto administradores.
        """
        queryset = PerfilUsuario.objects.all()
        if self.request.user.is_superuser or self.request.user.is_staff:
            return queryset
        # Usuario normal solo ve su propio perfil
        try:
            return queryset.filter(usuario=self.request.user)
        except AttributeError:
            return PerfilUsuario.objects.none()


@login_required
def opciones_usuario(request):
    """
    Vista para que el usuario pueda ver y editar sus datos personales
    y cambiar su contraseña.
    """
    usuario = request.user
    
    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        if accion == 'actualizar_datos':
            # Actualizar datos personales
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            numero = request.POST.get('numero', '').strip()
            
            # Validaciones básicas
            if not first_name or not last_name or not email:
                messages.error(request, 'Nombre, apellido y correo electrónico son obligatorios.')
                return redirect('opciones_usuario')
            
            # Validar que el email no esté en uso por otro usuario
            if email != usuario.email and Usuario.objects.filter(email=email).exclude(id=usuario.id).exists():
                messages.error(request, 'Este correo electrónico ya está en uso por otro usuario.')
                return redirect('opciones_usuario')
            
            # Actualizar datos
            usuario.first_name = first_name
            usuario.last_name = last_name
            usuario.email = email
            if numero:
                try:
                    usuario.numero = int(numero)
                except ValueError:
                    messages.error(request, 'El número de teléfono debe ser un número válido.')
                    return redirect('opciones_usuario')
            else:
                usuario.numero = None
            
            usuario.save()
            messages.success(request, 'Datos actualizados correctamente.')
            return redirect('opciones_usuario')
        
        elif accion == 'cambiar_password':
            # Cambiar contraseña
            password_actual = request.POST.get('password_actual', '')
            password_nueva = request.POST.get('password_nueva', '')
            password_confirmar = request.POST.get('password_confirmar', '')
            
            # Validar que se ingresó la contraseña actual
            if not password_actual:
                messages.error(request, 'Debe ingresar su contraseña actual.')
                return redirect('opciones_usuario')
            
            # Verificar que la contraseña actual sea correcta
            if not usuario.check_password(password_actual):
                messages.error(request, 'La contraseña actual es incorrecta.')
                return redirect('opciones_usuario')
            
            # Validar que se ingresó la nueva contraseña
            if not password_nueva:
                messages.error(request, 'Debe ingresar una nueva contraseña.')
                return redirect('opciones_usuario')
            
            # Validar que las contraseñas nuevas coincidan
            if password_nueva != password_confirmar:
                messages.error(request, 'Las contraseñas nuevas no coinciden.')
                return redirect('opciones_usuario')
            
            # Validar la nueva contraseña
            es_valida, mensaje_error = validar_contraseña(password_nueva)
            if not es_valida:
                messages.error(request, mensaje_error)
                return redirect('opciones_usuario')
            
            # Cambiar la contraseña
            usuario.set_password(password_nueva)
            usuario.save()
            
            # Actualizar la sesión para que el usuario no se desloguee
            update_session_auth_hash(request, usuario)
            
            messages.success(request, 'Contraseña cambiada correctamente.')
            return redirect('opciones_usuario')
    
    # GET: Mostrar formulario con datos actuales
    context = {
        'usuario': usuario,
        'rol': usuario.perfil.rol.nombre_rol if hasattr(usuario, 'perfil') and usuario.perfil.rol else 'Sin rol asignado',
    }
    
    return render(request, 'SIAPE/opciones_usuario.html', context)

